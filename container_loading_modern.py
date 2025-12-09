# -*- coding: utf-8 -*-
"""
集装箱配载软件 (Container Loading Software) v0.5
使用 PyQt6 + OpenGL 实现可拖动旋转的3D视图
支持多集装箱、装载图导出、拖拽调整等高级功能

作者: Henry Xue
版本: 0.5
"""

import sys
import json
import math
import numpy as np
from dataclasses import dataclass, asdict, field
from typing import List, Optional, Tuple, Dict
import copy
import io
import os

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# 图片导出支持
try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_SUPPORT = True
except ImportError:
    PIL_SUPPORT = False

# PDF导出支持
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_SUPPORT = True
    # 注册中文字体
    try:
        # Windows 系统字体路径
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",  # 微软雅黑
            "C:/Windows/Fonts/simsun.ttc",  # 宋体
            "C:/Windows/Fonts/simhei.ttf",  # 黑体
        ]
        font_registered = False
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                font_registered = True
                break
        if not font_registered:
            PDF_SUPPORT = False
    except:
        PDF_SUPPORT = False
except ImportError:
    PDF_SUPPORT = False

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar,
    QFileDialog, QMessageBox, QSplitter, QFrame, QSpinBox,
    QDoubleSpinBox, QStyle, QStyleFactory, QScrollArea,
    QDialog, QGridLayout, QFormLayout, QListWidget, QTabWidget,
    QProgressDialog
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QColor, QPalette, QIcon

from OpenGL.GL import *
from OpenGL.GLU import *
from PyQt6.QtOpenGLWidgets import QOpenGLWidget


@dataclass
class PalletContent:
    """托盘内的货物位置信息"""
    cargo: 'Cargo'  # 原始货物
    x: float  # 在托盘内的X位置
    y: float  # 在托盘内的Y位置
    z: float  # 在托盘内的Z位置（从托盘面开始）
    rotated: bool = False  # 是否旋转
    quantity: int = 1  # 该货物的数量
    
    @property
    def actual_length(self) -> float:
        return self.cargo.width if self.rotated else self.cargo.length
    
    @property
    def actual_width(self) -> float:
        return self.cargo.length if self.rotated else self.cargo.width


@dataclass
class Cargo:
    """货物类"""
    id: str = ""  # 货物唯一ID
    name: str = ""
    length: float = 0  # 长度 (cm)
    width: float = 0   # 宽度 (cm)
    height: float = 0  # 高度 (cm)
    weight: float = 0  # 重量 (kg)
    quantity: int = 1  # 数量
    stackable: bool = True  # 是否可堆叠
    max_stack: int = 3  # 最大堆叠层数
    color: Tuple[float, float, float] = (0.3, 0.7, 0.3)  # RGB颜色
    group_id: str = ""  # 组ID，同组货物锁定在一起
    allow_rotate: bool = True  # 是否允许旋转
    bottom_only: bool = False  # 是否只能放在底层
    priority: int = 0  # 装载优先级（数字越大越优先）
    # 组托相关字段
    is_pallet: bool = False  # 是否是组托后的托盘
    pallet_base_height: float = 15  # 托盘底座高度 (cm)
    pallet_contents: List['PalletContent'] = field(default_factory=list)  # 托盘内的货物列表
    original_cargos: List['Cargo'] = field(default_factory=list)  # 原始货物列表（未展开）
    
    def __post_init__(self):
        if not self.id:
            import uuid
            self.id = str(uuid.uuid4())[:8]
    
    @property
    def volume(self) -> float:
        return self.length * self.width * self.height
    
    @property
    def total_volume(self) -> float:
        return self.volume * self.quantity
    
    @property
    def total_weight(self) -> float:
        return self.weight * self.quantity
    
    @property
    def content_height(self) -> float:
        """货物实际高度（不含托盘底座）"""
        if self.is_pallet:
            return self.height - self.pallet_base_height
        return self.height


@dataclass
class CargoGroup:
    """货物组 - 多个货物锁定在一起"""
    id: str
    name: str
    cargo_ids: List[str] = field(default_factory=list)
    # 组合后的整体尺寸（自动计算或手动指定）
    combined_length: float = 0
    combined_width: float = 0
    combined_height: float = 0
    combined_weight: float = 0


@dataclass
class Container:
    """容器类（集装箱/货车/托盘）"""
    name: str
    length: float  # 内部长度 (cm)
    width: float   # 内部宽度 (cm)
    height: float  # 内部高度 (cm)
    max_weight: float  # 最大载重 (kg)
    container_type: str = "container"  # container/truck/pallet
    description: str = ""
    
    @property
    def volume(self) -> float:
        return self.length * self.width * self.height
    
    @property
    def volume_cbm(self) -> float:
        return self.volume / 1000000


@dataclass
class PlacedCargo:
    """已放置的货物"""
    cargo: Cargo
    x: float
    y: float
    z: float
    rotated: bool = False
    step_number: int = 0  # 装箱步骤编号
    container_index: int = 0  # 所属集装箱索引（多集装箱时使用）
    
    @property
    def actual_length(self) -> float:
        return self.cargo.width if self.rotated else self.cargo.length
    
    @property
    def actual_width(self) -> float:
        return self.cargo.length if self.rotated else self.cargo.width
    
    @property
    def center_x(self) -> float:
        return self.x + self.actual_length / 2
    
    @property
    def center_y(self) -> float:
        return self.y + self.actual_width / 2
    
    @property
    def center_z(self) -> float:
        return self.z + self.cargo.height / 2


@dataclass
class ContainerLoadingResult:
    """单个集装箱的装载结果"""
    container: Container
    container_index: int
    placed_cargos: List[PlacedCargo] = field(default_factory=list)
    
    @property
    def total_volume(self) -> float:
        return sum(p.cargo.volume for p in self.placed_cargos)
    
    @property
    def total_weight(self) -> float:
        return sum(p.cargo.weight for p in self.placed_cargos)
    
    @property
    def volume_utilization(self) -> float:
        if self.container.volume == 0:
            return 0
        return (self.total_volume / self.container.volume) * 100
    
    @property
    def weight_utilization(self) -> float:
        if self.container.max_weight == 0:
            return 0
        return (self.total_weight / self.container.max_weight) * 100


# ==================== 容器预设 ====================

# 标准集装箱
CONTAINERS_SHIPPING = {
    "20英尺标准箱 (20' GP)": Container("20英尺标准箱", 589, 234, 238, 21770, "container", "标准20尺海运集装箱"),
    "40英尺标准箱 (40' GP)": Container("40英尺标准箱", 1203, 234, 238, 26680, "container", "标准40尺海运集装箱"),
    "40英尺高箱 (40' HC)": Container("40英尺高箱", 1203, 234, 269, 26460, "container", "40尺高柜海运集装箱"),
    "45英尺高箱 (45' HC)": Container("45英尺高箱", 1351, 234, 269, 25600, "container", "45尺高柜海运集装箱"),
}

# 货车类型
CONTAINERS_TRUCK = {
    "4.2米厢式货车": Container("4.2米厢式货车", 420, 180, 180, 2000, "truck", "轻型厢式货车"),
    "6.8米平板车": Container("6.8米平板车", 680, 235, 230, 10000, "truck", "中型平板货车"),
    "7.7米厢式货车": Container("7.7米厢式货车", 770, 235, 240, 12000, "truck", "中型厢式货车"),
    "9.6米厢式货车": Container("9.6米厢式货车", 960, 235, 250, 18000, "truck", "大型厢式货车"),
    "9.6米飞翼车": Container("9.6米飞翼车", 960, 235, 260, 18000, "truck", "侧开式飞翼货车"),
    "13米平板车": Container("13米平板车", 1300, 245, 260, 32000, "truck", "重型平板货车"),
    "13米厢式货车": Container("13米厢式货车", 1300, 245, 270, 32000, "truck", "重型厢式货车"),
    "17.5米高低板车": Container("17.5米高低板车", 1750, 300, 300, 35000, "truck", "超长高低板挂车"),
    "17.5米平板车": Container("17.5米平板车", 1750, 300, 280, 35000, "truck", "超长平板挂车"),
}

# 托盘类型
CONTAINERS_PALLET = {
    "标准托盘 (1200×1000)": Container("标准托盘", 120, 100, 150, 1000, "pallet", "欧标托盘1200×1000mm"),
    "标准托盘 (1200×800)": Container("标准托盘", 120, 80, 150, 800, "pallet", "欧标托盘1200×800mm"),
    "美标托盘 (1219×1016)": Container("美标托盘", 122, 102, 150, 1000, "pallet", "美标托盘48×40英寸"),
    "日标托盘 (1100×1100)": Container("日标托盘", 110, 110, 150, 1000, "pallet", "日标方形托盘"),
    "仓储笼 (1200×1000×890)": Container("仓储笼", 120, 100, 89, 1500, "pallet", "标准仓储笼箱"),
    "周转箱 (600×400×280)": Container("周转箱", 60, 40, 28, 50, "pallet", "标准物流周转箱"),
}

# 合并所有容器类型
STANDARD_CONTAINERS = {
    **CONTAINERS_SHIPPING,
    **CONTAINERS_TRUCK,
    **CONTAINERS_PALLET,
}

# 容器分类
CONTAINER_CATEGORIES = {
    "海运集装箱": list(CONTAINERS_SHIPPING.keys()),
    "公路货车": list(CONTAINERS_TRUCK.keys()),
    "托盘/周转箱": list(CONTAINERS_PALLET.keys()),
    "自定义": [],
}

# 预设颜色 (RGB 0-1)
CARGO_COLORS = [
    (0.30, 0.69, 0.31),  # 绿色
    (0.13, 0.59, 0.95),  # 蓝色
    (1.00, 0.60, 0.00),  # 橙色
    (0.91, 0.12, 0.39),  # 粉红
    (0.61, 0.15, 0.69),  # 紫色
    (0.00, 0.74, 0.83),  # 青色
    (1.00, 0.92, 0.23),  # 黄色
    (0.47, 0.33, 0.28),  # 棕色
    (0.38, 0.49, 0.55),  # 灰蓝
    (0.96, 0.26, 0.21),  # 红色
    (0.55, 0.76, 0.29),  # 浅绿
    (0.01, 0.66, 0.96),  # 浅蓝
    (0.80, 0.86, 0.22),  # 黄绿
    (0.40, 0.23, 0.72),  # 深紫
    (0.00, 0.59, 0.53),  # 深青
]


# ==================== 配载规则 ====================

@dataclass
class LoadingRule:
    """配载规则"""
    id: str
    name: str
    description: str
    enabled: bool = True
    priority: int = 0  # 优先级，数字越大越优先
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        """应用规则对货物排序，子类重写"""
        return cargos


class RuleSameSizeFirst(LoadingRule):
    """相同尺寸优先配载规则"""
    def __init__(self):
        super().__init__("same_size", "相同尺寸优先", "相同或相近尺寸的货物优先放在一起", True, 50)
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        if not cargos:
            return cargos
        # 按尺寸分组排序
        def size_key(c):
            return (round(c.length / 10) * 10, round(c.width / 10) * 10, round(c.height / 10) * 10)
        return sorted(cargos, key=size_key, reverse=True)


class RuleHeavyBottom(LoadingRule):
    """重物下沉规则"""
    def __init__(self, weight_threshold: float = 100):
        super().__init__("heavy_bottom", "重物下沉", f"重量超过{weight_threshold}kg的货物优先放在底层", True, 80)
        self.weight_threshold = weight_threshold
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        heavy = [c for c in cargos if c.weight >= self.weight_threshold]
        light = [c for c in cargos if c.weight < self.weight_threshold]
        # 重物优先，按重量降序
        heavy.sort(key=lambda c: c.weight, reverse=True)
        return heavy + light


class RuleSimilarSizeStack(LoadingRule):
    """相近尺寸堆叠规则"""
    def __init__(self, tolerance: float = 50):
        super().__init__("similar_stack", "相近尺寸堆叠", f"长度差{tolerance}mm以内的货物可堆叠", True, 60)
        self.tolerance = tolerance  # mm
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        # 按长度排序，便于相近尺寸的货物放在一起
        return sorted(cargos, key=lambda c: c.length, reverse=True)


class RuleVolumeFirst(LoadingRule):
    """体积优先规则（默认）"""
    def __init__(self):
        super().__init__("volume_first", "体积优先", "按体积从大到小装载", True, 40)
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        return sorted(cargos, key=lambda c: c.volume, reverse=True)


class RulePriorityFirst(LoadingRule):
    """优先级规则"""
    def __init__(self):
        super().__init__("priority_first", "按优先级", "按货物设定的优先级装载", True, 100)
    
    def apply(self, cargos: List[Cargo], placed: List[PlacedCargo]) -> List[Cargo]:
        return sorted(cargos, key=lambda c: c.priority, reverse=True)


# 默认规则集
DEFAULT_RULES = [
    RulePriorityFirst(),
    RuleHeavyBottom(100),
    RuleSimilarSizeStack(50),
    RuleSameSizeFirst(),
    RuleVolumeFirst(),
]


class LoadingAlgorithm:
    """装载算法类 - 优化版，目标是最高装载率"""
    
    def __init__(self, container: Container, rules: List[LoadingRule] = None, 
                 cargo_groups: List[CargoGroup] = None):
        self.container = container
        self.placed_cargos: List[PlacedCargo] = []
        self.rules = rules or DEFAULT_RULES.copy()
        self.cargo_groups = cargo_groups or []
        self.step_counter = 0
        self.similar_size_tolerance = 50  # mm，相近尺寸容差
        # 空间网格步长，用于更精细的位置搜索
        self.grid_step = 10  # cm
    
    def can_place(self, cargo: Cargo, x: float, y: float, z: float, rotated: bool) -> bool:
        # 检查是否允许旋转
        if rotated and not cargo.allow_rotate:
            return False
        
        length = cargo.width if rotated else cargo.length
        width = cargo.length if rotated else cargo.width
        height = cargo.height
        
        # 检查是否只能放底层
        if cargo.bottom_only and z > 0.01:
            return False
        
        # 严格的边界检查 - 不允许任何超出
        if x < -0.01 or y < -0.01 or z < -0.01:
            return False
        if x + length > self.container.length + 0.01:
            return False
        if y + width > self.container.width + 0.01:
            return False
        if z + height > self.container.height + 0.01:
            return False
        
        # 碰撞检测
        for placed in self.placed_cargos:
            pl = placed.actual_length
            pw = placed.actual_width
            ph = placed.cargo.height
            
            # 使用严格的碰撞检测
            if (x < placed.x + pl - 0.01 and x + length > placed.x + 0.01 and
                y < placed.y + pw - 0.01 and y + width > placed.y + 0.01 and
                z < placed.z + ph - 0.01 and z + height > placed.z + 0.01):
                return False
        
        # 堆叠支撑检查
        if z > 0.01:
            support_area = 0
            required_support = length * width * 0.7
            
            for placed in self.placed_cargos:
                if abs(placed.z + placed.cargo.height - z) < 0.1:
                    pl = placed.actual_length
                    pw = placed.actual_width
                    
                    overlap_x = max(0, min(x + length, placed.x + pl) - max(x, placed.x))
                    overlap_y = max(0, min(y + width, placed.y + pw) - max(y, placed.y))
                    support_area += overlap_x * overlap_y
            
            if support_area < required_support:
                return False
        
        return True
    
    def calculate_best_rotation_for_layer(self, cargo: Cargo) -> bool:
        """计算在当前层最优的旋转方向，目标是最大化可放置数量"""
        if not cargo.allow_rotate:
            return False
        
        # 计算不旋转时每层可放数量
        cols_no_rotate = int(self.container.width // cargo.width)
        rows_no_rotate = int(self.container.length // cargo.length)
        count_no_rotate = cols_no_rotate * rows_no_rotate
        
        # 计算旋转时每层可放数量
        cols_rotated = int(self.container.width // cargo.length)
        rows_rotated = int(self.container.length // cargo.width)
        count_rotated = cols_rotated * rows_rotated
        
        # 选择能放更多货物的方向
        return count_rotated > count_no_rotate
    
    def get_candidate_positions(self, cargo: Cargo, rotated: bool) -> List[Tuple[float, float, float]]:
        """获取所有候选放置位置 - 优化版"""
        positions = set()
        
        length = cargo.width if rotated else cargo.length
        width = cargo.length if rotated else cargo.width
        
        # 基础位置：原点
        positions.add((0, 0, 0))
        
        # 基于已放置货物生成候选位置
        for placed in self.placed_cargos:
            pl = placed.actual_length
            pw = placed.actual_width
            ph = placed.cargo.height
            
            # 货物右侧
            positions.add((placed.x + pl, placed.y, placed.z))
            # 货物后方
            positions.add((placed.x, placed.y + pw, placed.z))
            # 货物顶部（如果可堆叠）
            if placed.cargo.stackable and not cargo.bottom_only:
                positions.add((placed.x, placed.y, placed.z + ph))
            
            # 额外的紧凑位置 - 靠近已放置货物的边缘
            # 右侧对齐
            if placed.x + pl + length <= self.container.length:
                positions.add((placed.x + pl, 0, placed.z))
                positions.add((placed.x + pl, 0, 0))
            # 后方对齐
            if placed.y + pw + width <= self.container.width:
                positions.add((0, placed.y + pw, placed.z))
                positions.add((0, placed.y + pw, 0))
        
        # 沿着容器边缘的位置
        for placed in self.placed_cargos:
            pl = placed.actual_length
            pw = placed.actual_width
            # 尝试贴着左边界
            positions.add((0, placed.y, placed.z))
            positions.add((0, placed.y + pw, placed.z))
            # 尝试贴着前边界
            positions.add((placed.x, 0, placed.z))
            positions.add((placed.x + pl, 0, placed.z))
        
        return list(positions)
    
    def calculate_placement_score(self, cargo: Cargo, x: float, y: float, z: float, rotated: bool) -> float:
        """计算放置位置的得分 - 分数越低越好"""
        length = cargo.width if rotated else cargo.length
        width = cargo.length if rotated else cargo.width
        
        # 基础得分：优先填充角落和边缘
        # 越靠近原点越好
        distance_score = x * 1.0 + y * 1.5 + z * 2.0
        
        # 紧凑性得分：与已有货物的贴合度
        contact_score = 0
        for placed in self.placed_cargos:
            pl = placed.actual_length
            pw = placed.actual_width
            ph = placed.cargo.height
            
            # 检查是否紧贴（在X、Y或Z方向上相邻）
            # X方向贴合
            if abs(x - (placed.x + pl)) < 0.1 or abs(placed.x - (x + length)) < 0.1:
                overlap_y = max(0, min(y + width, placed.y + pw) - max(y, placed.y))
                overlap_z = max(0, min(z + cargo.height, placed.z + ph) - max(z, placed.z))
                contact_score -= overlap_y * overlap_z * 0.01
            
            # Y方向贴合
            if abs(y - (placed.y + pw)) < 0.1 or abs(placed.y - (y + width)) < 0.1:
                overlap_x = max(0, min(x + length, placed.x + pl) - max(x, placed.x))
                overlap_z = max(0, min(z + cargo.height, placed.z + ph) - max(z, placed.z))
                contact_score -= overlap_x * overlap_z * 0.01
            
            # Z方向贴合（底部支撑）
            if abs(z - (placed.z + ph)) < 0.1:
                overlap_x = max(0, min(x + length, placed.x + pl) - max(x, placed.x))
                overlap_y = max(0, min(y + width, placed.y + pw) - max(y, placed.y))
                contact_score -= overlap_x * overlap_y * 0.02  # 底部支撑更重要
        
        # 边界贴合加分
        if x < 0.1:  # 贴左边界
            contact_score -= width * cargo.height * 0.005
        if y < 0.1:  # 贴前边界
            contact_score -= length * cargo.height * 0.005
        if z < 0.1:  # 贴底部
            contact_score -= length * width * 0.01
        
        # 空间利用率：优先选择能更好利用剩余空间的位置
        remaining_x = self.container.length - (x + length)
        remaining_y = self.container.width - (y + width)
        
        # 如果剩余空间太小无法放置其他货物，给予惩罚
        waste_penalty = 0
        if 0 < remaining_x < 30:  # 剩余空间太小
            waste_penalty += remaining_x * 0.5
        if 0 < remaining_y < 30:
            waste_penalty += remaining_y * 0.5
        
        return distance_score + contact_score + waste_penalty
    
    def find_best_rotation(self, cargo: Cargo, x: float, y: float, z: float) -> Tuple[bool, float]:
        """找到最佳旋转方向，返回(是否旋转, 得分)"""
        best_rotated = False
        best_score = float('inf')
        
        rotations = [False]
        if cargo.allow_rotate:
            rotations.append(True)
        
        for rotated in rotations:
            if self.can_place(cargo, x, y, z, rotated):
                score = self.calculate_placement_score(cargo, x, y, z, rotated)
                if score < best_score:
                    best_score = score
                    best_rotated = rotated
        
        return best_rotated, best_score

    def find_position(self, cargo: Cargo) -> Optional[Tuple[float, float, float, bool]]:
        """寻找最佳放置位置 - 优化版，优先考虑最大化装载率的旋转方向"""
        best_position = None
        best_score = float('inf')
        
        # 首先计算最优旋转方向（基于能放置更多货物）
        optimal_rotation = self.calculate_best_rotation_for_layer(cargo)
        
        # 按优先级尝试旋转方向：先尝试最优方向
        if cargo.allow_rotate:
            rotations = [optimal_rotation, not optimal_rotation]
        else:
            rotations = [False]
        
        for rotated in rotations:
            # 获取候选位置
            positions = self.get_candidate_positions(cargo, rotated)
            
            for x, y, z in positions:
                if self.can_place(cargo, x, y, z, rotated):
                    score = self.calculate_placement_score(cargo, x, y, z, rotated)
                    # 如果使用最优旋转方向，给予额外优势
                    if rotated == optimal_rotation:
                        score -= 100  # 奖励最优旋转
                    if score < best_score:
                        best_score = score
                        best_position = (x, y, z, rotated)
        
        # 如果常规位置找不到，尝试更细粒度的搜索
        if best_position is None:
            # 网格搜索
            z_levels = [0]
            for placed in self.placed_cargos:
                z_levels.append(placed.z + placed.cargo.height)
            z_levels = sorted(set(z_levels))
            
            for z in z_levels:
                for x in range(0, int(self.container.length), self.grid_step):
                    for y in range(0, int(self.container.width), self.grid_step):
                        for rotated in rotations:
                            if self.can_place(cargo, x, y, z, rotated):
                                score = self.calculate_placement_score(cargo, x, y, z, rotated)
                                if rotated == optimal_rotation:
                                    score -= 100
                                if score < best_score:
                                    best_score = score
                                    best_position = (x, y, z, rotated)
        
        return best_position
        
        return best_position
    
    def place_cargo(self, cargo: Cargo) -> bool:
        position = self.find_position(cargo)
        if position:
            x, y, z, rotated = position
            self.step_counter += 1
            placed = PlacedCargo(cargo, x, y, z, rotated, self.step_counter)
            self.placed_cargos.append(placed)
            return True
        return False
    
    def apply_rules(self, cargos: List[Cargo]) -> List[Cargo]:
        """应用所有启用的规则 - 使用复合排序实现多规则联合作用"""
        # 获取启用的规则，按优先级降序
        enabled_rules = sorted([r for r in self.rules if r.enabled], 
                              key=lambda r: r.priority, reverse=True)
        
        if not enabled_rules:
            return cargos
        
        # 使用复合排序键：每个规则产生一个排序分数
        def composite_key(cargo):
            scores = []
            for rule in enabled_rules:
                if rule.id == "priority_first":
                    # 优先级规则：优先级越高分数越高
                    scores.append(-cargo.priority)  # 负号使高优先级排前面
                elif rule.id == "heavy_bottom":
                    # 重物下沉：重货优先
                    threshold = getattr(rule, 'weight_threshold', 100)
                    if cargo.weight >= threshold:
                        scores.append(0)  # 重货排前
                    else:
                        scores.append(1)  # 轻货排后
                    scores.append(-cargo.weight)  # 同类按重量降序
                elif rule.id == "volume_first":
                    # 体积优先：体积越大排越前
                    scores.append(-cargo.volume)
                elif rule.id == "similar_stack":
                    # 相近尺寸：按长度分组
                    scores.append(-cargo.length)
                elif rule.id == "same_size":
                    # 相同尺寸优先：按尺寸分组
                    size_group = (round(cargo.length / 10) * 10, 
                                 round(cargo.width / 10) * 10, 
                                 round(cargo.height / 10) * 10)
                    scores.append((-size_group[0], -size_group[1], -size_group[2]))
            return tuple(scores)
        
        return sorted(cargos, key=composite_key)
    
    def expand_groups(self, cargos: List[Cargo]) -> List[Cargo]:
        """处理货物组，将组合货物合并为单个虚拟货物"""
        if not self.cargo_groups:
            return cargos
        
        result = []
        grouped_ids = set()
        
        for group in self.cargo_groups:
            group_cargos = [c for c in cargos if c.id in group.cargo_ids]
            if group_cargos:
                # 计算组合后的尺寸（取最大包围盒）
                if group.combined_length > 0:
                    combined = Cargo(
                        name=group.name,
                        length=group.combined_length,
                        width=group.combined_width,
                        height=group.combined_height,
                        weight=group.combined_weight or sum(c.weight for c in group_cargos),
                        quantity=1,
                        stackable=all(c.stackable for c in group_cargos),
                        color=group_cargos[0].color if group_cargos else (0.5, 0.5, 0.5),
                        group_id=group.id
                    )
                else:
                    # 自动计算组合尺寸
                    combined = Cargo(
                        name=group.name,
                        length=max(c.length for c in group_cargos),
                        width=max(c.width for c in group_cargos),
                        height=sum(c.height for c in group_cargos),
                        weight=sum(c.weight for c in group_cargos),
                        quantity=1,
                        stackable=all(c.stackable for c in group_cargos),
                        color=group_cargos[0].color if group_cargos else (0.5, 0.5, 0.5),
                        group_id=group.id
                    )
                result.append(combined)
                grouped_ids.update(group.cargo_ids)
        
        # 添加未分组的货物
        for cargo in cargos:
            if cargo.id not in grouped_ids:
                result.append(cargo)
        
        return result
    
    def load_all(self, cargos: List[Cargo]) -> Tuple[List[PlacedCargo], List[Cargo]]:
        """装载所有货物"""
        # 处理货物组
        processed_cargos = self.expand_groups(cargos)
        
        # 展开数量
        sorted_cargos = []
        for cargo in processed_cargos:
            for i in range(cargo.quantity):
                single_cargo = copy.copy(cargo)
                single_cargo.quantity = 1
                single_cargo.id = f"{cargo.id}_{i}"
                sorted_cargos.append(single_cargo)
        
        # 应用配载规则
        sorted_cargos = self.apply_rules(sorted_cargos)
        
        loaded = []
        not_loaded = []
        
        for cargo in sorted_cargos:
            if self.place_cargo(cargo):
                loaded.append(self.placed_cargos[-1])
            else:
                not_loaded.append(cargo)
        
        return loaded, not_loaded
    
    def calculate_center_of_gravity(self) -> Tuple[float, float, float]:
        """计算重心位置"""
        if not self.placed_cargos:
            return (0, 0, 0)
        
        total_weight = sum(p.cargo.weight for p in self.placed_cargos)
        if total_weight == 0:
            return (0, 0, 0)
        
        cx = sum(p.center_x * p.cargo.weight for p in self.placed_cargos) / total_weight
        cy = sum(p.center_y * p.cargo.weight for p in self.placed_cargos) / total_weight
        cz = sum(p.center_z * p.cargo.weight for p in self.placed_cargos) / total_weight
        
        return (cx, cy, cz)
    
    def calculate_center_offset(self) -> Tuple[float, float, float]:
        """计算重心偏移量（相对于容器中心）"""
        cx, cy, cz = self.calculate_center_of_gravity()
        container_cx = self.container.length / 2
        container_cy = self.container.width / 2
        container_cz = self.container.height / 2
        
        return (cx - container_cx, cy - container_cy, cz - container_cz)
    
    def get_loading_steps(self) -> List[dict]:
        """获取装箱步骤"""
        steps = []
        sorted_placements = sorted(self.placed_cargos, key=lambda p: p.step_number)
        
        for p in sorted_placements:
            position_desc = []
            if p.x < self.container.length * 0.33:
                position_desc.append("柜头")
            elif p.x > self.container.length * 0.66:
                position_desc.append("柜尾")
            else:
                position_desc.append("中部")
            
            if p.y < self.container.width * 0.5:
                position_desc.append("左侧")
            else:
                position_desc.append("右侧")
            
            if p.z < 1:
                position_desc.append("底层")
            elif p.z > self.container.height * 0.5:
                position_desc.append("上层")
            else:
                position_desc.append("中层")
            
            steps.append({
                "step": p.step_number,
                "container": "-",  # 单集装箱模式
                "cargo_name": p.cargo.name,
                "dimensions": f"{p.actual_length:.0f}×{p.actual_width:.0f}×{p.cargo.height:.0f}",
                "position": f"({p.x:.0f}, {p.y:.0f}, {p.z:.0f})",
                "position_desc": " ".join(position_desc),
                "rotated": "是" if p.rotated else "否",
                "securing": "标准"
            })
        
        return steps
    
    def get_statistics(self) -> dict:
        total_cargo_volume = sum(p.cargo.volume for p in self.placed_cargos)
        total_cargo_weight = sum(p.cargo.weight for p in self.placed_cargos)
        
        # 计算重心偏移
        offset_x, offset_y, offset_z = self.calculate_center_offset()
        
        # 计算偏移百分比
        offset_x_pct = (offset_x / (self.container.length / 2)) * 100 if self.container.length > 0 else 0
        offset_y_pct = (offset_y / (self.container.width / 2)) * 100 if self.container.width > 0 else 0
        
        return {
            "loaded_count": len(self.placed_cargos),
            "total_volume": total_cargo_volume,
            "volume_utilization": (total_cargo_volume / self.container.volume) * 100 if self.container.volume > 0 else 0,
            "total_weight": total_cargo_weight,
            "weight_utilization": (total_cargo_weight / self.container.max_weight) * 100 if self.container.max_weight > 0 else 0,
            "center_of_gravity": self.calculate_center_of_gravity(),
            "center_offset": (offset_x, offset_y, offset_z),
            "center_offset_pct": (offset_x_pct, offset_y_pct),
        }


class Container3DView(QOpenGLWidget):
    """OpenGL 3D视图组件 - 支持拖拽选择和多集装箱"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.container: Optional[Container] = None
        self.placed_cargos: List[PlacedCargo] = []
        self.all_container_results: List[ContainerLoadingResult] = []  # 多集装箱结果
        self.current_container_index: int = -1  # -1表示显示全部概览
        
        # 视角控制
        self.rotation_x = 25
        self.rotation_y = 45
        self.zoom = 1.0
        self.pan_x = 0
        self.pan_y = 0
        
        # 鼠标控制
        self.last_mouse_pos = None
        self.mouse_button = None
        
        # 拖拽选择
        self.drag_mode = False  # 是否处于拖拽调整模式
        self.selected_cargo_index = -1  # 当前选中的货物索引
        self.dragging = False
        self.drag_start_pos = None
        self.drag_axis = None  # 'x', 'y', 'z'
        
        # 吸附和碰撞检测
        self.snap_distance = 5  # 吸附距离 (cm)
        self.collision_enabled = True  # 是否启用碰撞检测
        
        # 选择回调
        self.on_cargo_selected = None  # 选中货物时的回调
        self.on_cargo_moved = None  # 移动货物后的回调
        self.on_cargo_rotated = None  # 旋转货物后的回调
        
        self.setMinimumSize(600, 400)
    
    def set_drag_mode(self, enabled: bool):
        """设置拖拽模式"""
        self.drag_mode = enabled
        if not enabled:
            self.selected_cargo_index = -1
            self.dragging = False
        self.update()
    
    def check_collision(self, placed: 'PlacedCargo', new_x: float, new_y: float, new_z: float, exclude_index: int = -1) -> bool:
        """检查货物在新位置是否与其他货物碰撞
        返回 True 表示有碰撞"""
        length = placed.actual_length
        width = placed.actual_width
        height = placed.cargo.height
        
        for i, other in enumerate(self.placed_cargos):
            if i == exclude_index:
                continue
            
            ol = other.actual_length
            ow = other.actual_width
            oh = other.cargo.height
            
            # 碰撞检测（带微小容差）
            if (new_x < other.x + ol - 0.5 and new_x + length > other.x + 0.5 and
                new_y < other.y + ow - 0.5 and new_y + width > other.y + 0.5 and
                new_z < other.z + oh - 0.5 and new_z + height > other.z + 0.5):
                return True
        return False
    
    def find_snap_position(self, placed: 'PlacedCargo', new_x: float, new_y: float, new_z: float, exclude_index: int = -1) -> tuple:
        """找到吸附位置
        返回 (snapped_x, snapped_y, snapped_z)"""
        snap_dist = self.snap_distance
        length = placed.actual_length
        width = placed.actual_width
        height = placed.cargo.height
        
        best_x, best_y, best_z = new_x, new_y, new_z
        
        # 吸附到容器边界
        if abs(new_x) < snap_dist:
            best_x = 0
        if abs(new_x + length - self.container.length) < snap_dist:
            best_x = self.container.length - length
        if abs(new_y) < snap_dist:
            best_y = 0
        if abs(new_y + width - self.container.width) < snap_dist:
            best_y = self.container.width - width
        if abs(new_z) < snap_dist:
            best_z = 0
        
        # 吸附到其他货物边缘
        for i, other in enumerate(self.placed_cargos):
            if i == exclude_index:
                continue
            
            ol = other.actual_length
            ow = other.actual_width
            oh = other.cargo.height
            
            # X方向吸附
            if abs(new_x - (other.x + ol)) < snap_dist:  # 左边缘对齐右边缘
                best_x = other.x + ol
            if abs(new_x + length - other.x) < snap_dist:  # 右边缘对齐左边缘
                best_x = other.x - length
            
            # Y方向吸附
            if abs(new_y - (other.y + ow)) < snap_dist:
                best_y = other.y + ow
            if abs(new_y + width - other.y) < snap_dist:
                best_y = other.y - width
            
            # Z方向吸附（堆叠）
            if abs(new_z - (other.z + oh)) < snap_dist:
                best_z = other.z + oh
            if abs(new_z + height - other.z) < snap_dist:
                best_z = other.z - height
        
        return best_x, best_y, best_z
    
    def rotate_selected_cargo(self):
        """旋转选中的货物（水平方向，长宽互换）"""
        if self.selected_cargo_index < 0 or self.selected_cargo_index >= len(self.placed_cargos):
            return False
        
        placed = self.placed_cargos[self.selected_cargo_index]
        cargo = placed.cargo
        
        # 检查是否允许旋转
        if not cargo.allow_rotate:
            return False
        
        # 保存原始状态
        original_rotated = placed.rotated
        original_x, original_y = placed.x, placed.y
        
        # 尝试旋转
        placed.rotated = not placed.rotated
        new_length = placed.actual_length
        new_width = placed.actual_width
        
        # 调整位置使货物中心保持不变
        center_x = original_x + (cargo.length if not original_rotated else cargo.width) / 2
        center_y = original_y + (cargo.width if not original_rotated else cargo.length) / 2
        new_x = center_x - new_length / 2
        new_y = center_y - new_width / 2
        
        # 确保在容器边界内
        new_x = max(0, min(self.container.length - new_length, new_x))
        new_y = max(0, min(self.container.width - new_width, new_y))
        
        # 检查碰撞
        if self.collision_enabled and self.check_collision(placed, new_x, new_y, placed.z, self.selected_cargo_index):
            # 如果有碰撞，尝试找附近的有效位置
            found_valid = False
            for dx in range(-50, 51, 10):
                for dy in range(-50, 51, 10):
                    test_x = new_x + dx
                    test_y = new_y + dy
                    test_x = max(0, min(self.container.length - new_length, test_x))
                    test_y = max(0, min(self.container.width - new_width, test_y))
                    if not self.check_collision(placed, test_x, test_y, placed.z, self.selected_cargo_index):
                        new_x, new_y = test_x, test_y
                        found_valid = True
                        break
                if found_valid:
                    break
            
            if not found_valid:
                # 无法旋转，恢复原状
                placed.rotated = original_rotated
                return False
        
        placed.x = new_x
        placed.y = new_y
        self.update()
        
        if self.on_cargo_rotated:
            self.on_cargo_rotated(self.selected_cargo_index)
        
        return True
    
    def move_selected_cargo(self, dx: float, dy: float, dz: float) -> bool:
        """移动选中的货物指定距离（用于微调）
        返回是否移动成功"""
        if self.selected_cargo_index < 0 or self.selected_cargo_index >= len(self.placed_cargos):
            return False
        
        placed = self.placed_cargos[self.selected_cargo_index]
        length = placed.actual_length
        width = placed.actual_width
        height = placed.cargo.height
        
        # 计算新位置
        new_x = placed.x + dx
        new_y = placed.y + dy
        new_z = placed.z + dz
        
        # 边界检查
        new_x = max(0, min(self.container.length - length, new_x))
        new_y = max(0, min(self.container.width - width, new_y))
        new_z = max(0, min(self.container.height - height, new_z))
        
        # 碰撞检测
        if self.collision_enabled and self.check_collision(placed, new_x, new_y, new_z, self.selected_cargo_index):
            return False
        
        placed.x = new_x
        placed.y = new_y
        placed.z = new_z
        self.update()
        
        if self.on_cargo_moved:
            self.on_cargo_moved(self.selected_cargo_index)
        
        return True

    def set_multi_container_results(self, results: List[ContainerLoadingResult]):
        """设置多集装箱结果"""
        self.all_container_results = results
        if results:
            self.current_container_index = -1  # 默认显示全部概览
            self.update_display()
        else:
            self.current_container_index = -1
            self.placed_cargos = []
        self.update()
    
    def show_container(self, index: int):
        """切换显示特定集装箱 (-1 显示全部概览)"""
        self.current_container_index = index
        self.update_display()
        self.reset_view()
    
    def update_display(self):
        """更新显示内容"""
        if not self.all_container_results:
            self.update()
            return
        
        if self.current_container_index >= 0 and self.current_container_index < len(self.all_container_results):
            # 显示单个集装箱
            result = self.all_container_results[self.current_container_index]
            self.container = result.container
            self.placed_cargos = result.placed_cargos
        else:
            # 全部概览模式 - 使用第一个集装箱作为参考
            if self.all_container_results:
                self.container = self.all_container_results[0].container
                # 合并所有货物用于统计，但实际绘制在 paintGL 中单独处理
                self.placed_cargos = []
                for result in self.all_container_results:
                    self.placed_cargos.extend(result.placed_cargos)
        
        self.update()
    
    def is_overview_mode(self) -> bool:
        """是否处于全局概览模式"""
        return self.current_container_index < 0 and len(self.all_container_results) >= 1
    
    def initializeGL(self):
        """初始化OpenGL"""
        glClearColor(0.15, 0.15, 0.18, 1.0)
        glEnable(GL_DEPTH_TEST)
        glEnable(GL_LIGHTING)
        glEnable(GL_LIGHT0)
        glEnable(GL_COLOR_MATERIAL)
        glColorMaterial(GL_FRONT_AND_BACK, GL_AMBIENT_AND_DIFFUSE)
        
        # 光源设置
        glLightfv(GL_LIGHT0, GL_POSITION, [1, 1, 1, 0])
        glLightfv(GL_LIGHT0, GL_AMBIENT, [0.3, 0.3, 0.3, 1])
        glLightfv(GL_LIGHT0, GL_DIFFUSE, [0.8, 0.8, 0.8, 1])
        
        glEnable(GL_BLEND)
        glBlendFunc(GL_SRC_ALPHA, GL_ONE_MINUS_SRC_ALPHA)
        
        # 仅启用线条抗锯齿，不启用多边形抗锯齿（会产生斜线）
        glEnable(GL_LINE_SMOOTH)
        glHint(GL_LINE_SMOOTH_HINT, GL_NICEST)
    
    def resizeGL(self, w, h):
        """调整视口"""
        glViewport(0, 0, w, h)
        glMatrixMode(GL_PROJECTION)
        glLoadIdentity()
        aspect = w / h if h > 0 else 1
        gluPerspective(45, aspect, 0.1, 10000)
        glMatrixMode(GL_MODELVIEW)
    
    def paintGL(self):
        """渲染场景"""
        glClear(GL_COLOR_BUFFER_BIT | GL_DEPTH_BUFFER_BIT)
        glLoadIdentity()
        
        if not self.container:
            return
        
        # 判断是否为全局概览模式
        if self.is_overview_mode():
            self.paintGL_overview()
        else:
            self.paintGL_single()
    
    def paintGL_single(self):
        """渲染单个集装箱场景"""
        # 计算观察距离 - 使用1.8让视图更近
        max_dim = max(self.container.length, self.container.width, self.container.height)
        distance = max_dim * 1.8 / self.zoom
        
        # 设置相机
        glTranslatef(self.pan_x, self.pan_y, -distance)
        glRotatef(self.rotation_x, 1, 0, 0)
        glRotatef(self.rotation_y, 0, 1, 0)
        
        # 将原点移到集装箱中心
        glTranslatef(-self.container.length/2, -self.container.height/2, -self.container.width/2)
        
        # 绘制地面网格
        self.draw_grid()
        
        # 绘制集装箱
        self.draw_container_wireframe()
        
        # 绘制已放置的货物（带索引用于选择）
        for i, placed in enumerate(self.placed_cargos):
            self.draw_cargo(placed, i)
        
        # 绘制坐标轴
        self.draw_axes()
        
        # 如果处于拖拽模式且有选中货物，显示拖拽轴
        if self.drag_mode and 0 <= self.selected_cargo_index < len(self.placed_cargos):
            self.draw_drag_axes(self.placed_cargos[self.selected_cargo_index])
    
    def paintGL_overview(self):
        """渲染全局概览模式 - 显示所有集装箱并排"""
        num_containers = len(self.all_container_results)
        if num_containers == 0:
            return
        
        # 计算所有集装箱的布局
        # 集装箱并排放置，中间留有间隙
        spacing = 200  # 集装箱之间的间隙 (cm) - 增加间距便于区分
        
        # 计算总宽度和最大尺寸
        total_length = 0
        max_height = 0
        max_width = 0
        
        for result in self.all_container_results:
            c = result.container
            total_length += c.length
            max_height = max(max_height, c.height)
            max_width = max(max_width, c.width)
        
        total_length += spacing * (num_containers - 1)
        
        # 计算观察距离 - 需要能看到所有集装箱
        max_dim = max(total_length, max_width * 2, max_height * 2)
        distance = max_dim * 1.8 / self.zoom
        
        # 设置相机
        glTranslatef(self.pan_x, self.pan_y, -distance)
        glRotatef(self.rotation_x, 1, 0, 0)
        glRotatef(self.rotation_y, 0, 1, 0)
        
        # 将原点移到所有集装箱的中心
        glTranslatef(-total_length/2, -max_height/2, -max_width/2)
        
        # 绘制扩展的地面网格
        self.draw_overview_grid(total_length, max_width)
        
        # 依次绘制每个集装箱
        x_offset = 0
        for idx, result in enumerate(self.all_container_results):
            glPushMatrix()
            glTranslatef(x_offset, 0, 0)
            
            # 直接使用result中的数据绘制，不修改self的属性
            container = result.container
            placed_cargos = result.placed_cargos
            
            # 绘制集装箱线框
            self.draw_container_wireframe_at(container)
            
            # 绘制货物
            for i, placed in enumerate(placed_cargos):
                self.draw_cargo(placed, i)
            
            # 绘制集装箱编号标签
            self.draw_container_label(idx + 1, container)
            
            glPopMatrix()
            
            x_offset += container.length + spacing
        
        # 绘制坐标轴
        self.draw_axes()
    
    def draw_overview_grid(self, total_length: float, max_width: float):
        """绘制全局概览的地面网格"""
        glDisable(GL_LIGHTING)
        glColor4f(0.3, 0.3, 0.35, 0.5)
        glLineWidth(1)
        
        padding = 100  # 网格边距
        step = 100  # 100cm 网格间距
        
        glBegin(GL_LINES)
        x = -padding
        while x <= total_length + padding:
            glVertex3f(x, 0, -padding)
            glVertex3f(x, 0, max_width + padding)
            x += step
        
        z = -padding
        while z <= max_width + padding:
            glVertex3f(-padding, 0, z)
            glVertex3f(total_length + padding, 0, z)
            z += step
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def draw_container_label(self, index: int, container):
        """绘制集装箱编号标签（使用简单的3D位置标记）"""
        # 在集装箱顶部中心位置绘制标记
        glDisable(GL_LIGHTING)
        
        cx = container.length / 2
        cy = container.height + 20  # 在顶部上方
        cz = container.width / 2
        
        # 绘制标记点
        glPointSize(15)
        glBegin(GL_POINTS)
        # 使用不同颜色区分不同集装箱
        colors = [
            (1.0, 0.4, 0.4),   # 红
            (0.4, 1.0, 0.4),   # 绿
            (0.4, 0.4, 1.0),   # 蓝
            (1.0, 1.0, 0.4),   # 黄
            (1.0, 0.4, 1.0),   # 紫
            (0.4, 1.0, 1.0),   # 青
        ]
        color = colors[(index - 1) % len(colors)]
        glColor3f(*color)
        glVertex3f(cx, cy, cz)
        glEnd()
        
        # 绘制编号指示线
        glLineWidth(2)
        glBegin(GL_LINES)
        glVertex3f(cx, container.height, cz)
        glVertex3f(cx, cy, cz)
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def capture_image(self, width: int = 800, height: int = 600) -> 'QImage':
        """捕获当前3D视图为图片"""
        from PyQt6.QtCore import QSize
        from PyQt6.QtGui import QImage
        
        # 保存当前尺寸
        old_size = self.size()
        
        # 调整到目标尺寸并渲染
        self.resize(width, height)
        self.makeCurrent()
        self.resizeGL(width, height)
        self.paintGL()
        
        # 捕获帧缓冲
        image = self.grabFramebuffer()
        
        # 恢复尺寸
        self.resize(old_size)
        self.makeCurrent()
        self.resizeGL(old_size.width(), old_size.height())
        self.update()
        
        return image
    
    def capture_isometric_image(self, width: int = 800, height: int = 600) -> 'QImage':
        """捕获等轴测视角的图片"""
        from PyQt6.QtGui import QImage
        
        # 保存当前视角
        old_rot_x = self.rotation_x
        old_rot_y = self.rotation_y
        old_zoom = self.zoom
        old_pan_x = self.pan_x
        old_pan_y = self.pan_y
        
        # 设置等轴测视角 (30度俯视, 45度侧视)
        self.rotation_x = 30
        self.rotation_y = 45
        self.zoom = 1.2  # 稍微拉近
        self.pan_x = 0
        self.pan_y = 0
        
        # 捕获图片
        image = self.capture_image(width, height)
        
        # 恢复视角
        self.rotation_x = old_rot_x
        self.rotation_y = old_rot_y
        self.zoom = old_zoom
        self.pan_x = old_pan_x
        self.pan_y = old_pan_y
        
        self.update()
        return image
    
    def draw_grid(self):
        """绘制地面网格"""
        glDisable(GL_LIGHTING)
        glColor4f(0.3, 0.3, 0.35, 0.5)
        glLineWidth(1)
        
        grid_size = max(self.container.length, self.container.width) * 1.5
        step = 50  # 50cm 网格
        
        glBegin(GL_LINES)
        x = -grid_size / 4
        while x <= self.container.length + grid_size / 4:
            glVertex3f(x, 0, -grid_size / 4)
            glVertex3f(x, 0, self.container.width + grid_size / 4)
            x += step
        
        z = -grid_size / 4
        while z <= self.container.width + grid_size / 4:
            glVertex3f(-grid_size / 4, 0, z)
            glVertex3f(self.container.length + grid_size / 4, 0, z)
            z += step
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def draw_container_wireframe(self):
        """绘制集装箱（半透明面+线框）"""
        l, w, h = self.container.length, self.container.width, self.container.height
        
        glDisable(GL_LIGHTING)
        glEnable(GL_BLEND)
        glBlendFunc(GL_SRC_ALPHA, GL_ONE_MINUS_SRC_ALPHA)
        glDepthMask(GL_FALSE)  # 禁用深度写入，让透明面正确显示
        
        # 绘制半透明的所有面
        glBegin(GL_QUADS)
        
        # 底面 - 稍深一点
        glColor4f(0.5, 0.5, 0.55, 0.35)
        glVertex3f(0, 0, 0)
        glVertex3f(l, 0, 0)
        glVertex3f(l, 0, w)
        glVertex3f(0, 0, w)
        
        # 顶面 - 很透明
        glColor4f(0.4, 0.4, 0.45, 0.15)
        glVertex3f(0, h, 0)
        glVertex3f(0, h, w)
        glVertex3f(l, h, w)
        glVertex3f(l, h, 0)
        
        # 前面 (z=0) - 半透明
        glColor4f(0.45, 0.45, 0.5, 0.2)
        glVertex3f(0, 0, 0)
        glVertex3f(0, h, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, 0, 0)
        
        # 后面 (z=w) - 半透明
        glColor4f(0.45, 0.45, 0.5, 0.2)
        glVertex3f(0, 0, w)
        glVertex3f(l, 0, w)
        glVertex3f(l, h, w)
        glVertex3f(0, h, w)
        
        # 左面 (x=0) - 半透明
        glColor4f(0.4, 0.4, 0.45, 0.2)
        glVertex3f(0, 0, 0)
        glVertex3f(0, 0, w)
        glVertex3f(0, h, w)
        glVertex3f(0, h, 0)
        
        # 右面 (x=l) - 半透明
        glColor4f(0.4, 0.4, 0.45, 0.2)
        glVertex3f(l, 0, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, h, w)
        glVertex3f(l, 0, w)
        
        glEnd()
        
        glDepthMask(GL_TRUE)  # 恢复深度写入
        
        # 绘制边框线
        glColor4f(0.8, 0.8, 0.85, 1.0)
        glLineWidth(2)
        
        # 底面边框
        glBegin(GL_LINE_LOOP)
        glVertex3f(0, 0, 0)
        glVertex3f(l, 0, 0)
        glVertex3f(l, 0, w)
        glVertex3f(0, 0, w)
        glEnd()
        
        # 顶面边框
        glBegin(GL_LINE_LOOP)
        glVertex3f(0, h, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, h, w)
        glVertex3f(0, h, w)
        glEnd()
        
        # 竖直边
        glBegin(GL_LINES)
        for x, z in [(0, 0), (l, 0), (l, w), (0, w)]:
            glVertex3f(x, 0, z)
            glVertex3f(x, h, z)
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def draw_container_wireframe_at(self, container):
        """绘制指定集装箱（半透明面+线框）- 用于概览模式"""
        l, w, h = container.length, container.width, container.height
        
        glDisable(GL_LIGHTING)
        glEnable(GL_BLEND)
        glBlendFunc(GL_SRC_ALPHA, GL_ONE_MINUS_SRC_ALPHA)
        glDepthMask(GL_FALSE)
        
        # 绘制半透明的所有面
        glBegin(GL_QUADS)
        
        # 底面
        glColor4f(0.5, 0.5, 0.55, 0.35)
        glVertex3f(0, 0, 0)
        glVertex3f(l, 0, 0)
        glVertex3f(l, 0, w)
        glVertex3f(0, 0, w)
        
        # 顶面
        glColor4f(0.4, 0.4, 0.45, 0.15)
        glVertex3f(0, h, 0)
        glVertex3f(0, h, w)
        glVertex3f(l, h, w)
        glVertex3f(l, h, 0)
        
        # 前面
        glColor4f(0.45, 0.45, 0.5, 0.2)
        glVertex3f(0, 0, 0)
        glVertex3f(0, h, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, 0, 0)
        
        # 后面
        glColor4f(0.45, 0.45, 0.5, 0.2)
        glVertex3f(0, 0, w)
        glVertex3f(l, 0, w)
        glVertex3f(l, h, w)
        glVertex3f(0, h, w)
        
        # 左面
        glColor4f(0.4, 0.4, 0.45, 0.2)
        glVertex3f(0, 0, 0)
        glVertex3f(0, 0, w)
        glVertex3f(0, h, w)
        glVertex3f(0, h, 0)
        
        # 右面
        glColor4f(0.4, 0.4, 0.45, 0.2)
        glVertex3f(l, 0, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, h, w)
        glVertex3f(l, 0, w)
        
        glEnd()
        
        glDepthMask(GL_TRUE)
        
        # 绘制边框线
        glColor4f(0.8, 0.8, 0.85, 1.0)
        glLineWidth(2)
        
        # 底面边框
        glBegin(GL_LINE_LOOP)
        glVertex3f(0, 0, 0)
        glVertex3f(l, 0, 0)
        glVertex3f(l, 0, w)
        glVertex3f(0, 0, w)
        glEnd()
        
        # 顶面边框
        glBegin(GL_LINE_LOOP)
        glVertex3f(0, h, 0)
        glVertex3f(l, h, 0)
        glVertex3f(l, h, w)
        glVertex3f(0, h, w)
        glEnd()
        
        # 竖直边
        glBegin(GL_LINES)
        for x, z in [(0, 0), (l, 0), (l, w), (0, w)]:
            glVertex3f(x, 0, z)
            glVertex3f(x, h, z)
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def draw_cargo(self, placed: PlacedCargo, index: int = -1):
        """绘制货物"""
        x, y, z = placed.x, placed.z, placed.y
        l = placed.actual_length
        h = placed.cargo.height
        w = placed.actual_width
        
        r, g, b = placed.cargo.color
        
        # 如果是选中状态，增加亮度
        is_selected = self.drag_mode and index == self.selected_cargo_index
        if is_selected:
            r = min(1.0, r + 0.3)
            g = min(1.0, g + 0.3)
            b = min(1.0, b + 0.3)
        
        # 定义顶点
        vertices = [
            (x, y, z), (x+l, y, z), (x+l, y, z+w), (x, y, z+w),
            (x, y+h, z), (x+l, y+h, z), (x+l, y+h, z+w), (x, y+h, z+w)
        ]
        
        glColor3f(r, g, b)
        
        # 绘制面
        glBegin(GL_QUADS)
        # 底面
        glNormal3f(0, -1, 0)
        glVertex3f(*vertices[0]); glVertex3f(*vertices[1]); glVertex3f(*vertices[2]); glVertex3f(*vertices[3])
        # 顶面
        glNormal3f(0, 1, 0)
        glVertex3f(*vertices[4]); glVertex3f(*vertices[7]); glVertex3f(*vertices[6]); glVertex3f(*vertices[5])
        # 前面
        glNormal3f(0, 0, -1)
        glVertex3f(*vertices[0]); glVertex3f(*vertices[4]); glVertex3f(*vertices[5]); glVertex3f(*vertices[1])
        # 后面
        glNormal3f(0, 0, 1)
        glVertex3f(*vertices[2]); glVertex3f(*vertices[6]); glVertex3f(*vertices[7]); glVertex3f(*vertices[3])
        # 左面
        glNormal3f(-1, 0, 0)
        glVertex3f(*vertices[0]); glVertex3f(*vertices[3]); glVertex3f(*vertices[7]); glVertex3f(*vertices[4])
        # 右面
        glNormal3f(1, 0, 0)
        glVertex3f(*vertices[1]); glVertex3f(*vertices[5]); glVertex3f(*vertices[6]); glVertex3f(*vertices[2])
        glEnd()
        
        # 绘制边框 - 使用更柔和的颜色和适当的线宽
        glDisable(GL_LIGHTING)
        
        # 轻微偏移避免z-fighting
        glEnable(GL_POLYGON_OFFSET_LINE)
        glPolygonOffset(-1.0, -1.0)
        
        if is_selected:
            glColor4f(1.0, 1.0, 0.0, 1.0)  # 选中时用黄色边框
            glLineWidth(2.5)
        else:
            # 使用货物颜色的深色版本作为边框，更自然
            glColor4f(r * 0.3, g * 0.3, b * 0.3, 0.8)
            glLineWidth(1.0)
        
        edges = [
            (0, 1), (1, 2), (2, 3), (3, 0),
            (4, 5), (5, 6), (6, 7), (7, 4),
            (0, 4), (1, 5), (2, 6), (3, 7)
        ]
        
        glBegin(GL_LINES)
        for i, j in edges:
            glVertex3f(*vertices[i])
            glVertex3f(*vertices[j])
        glEnd()
        
        glDisable(GL_POLYGON_OFFSET_LINE)
        glEnable(GL_LIGHTING)
    
    def draw_axes(self):
        """绘制坐标轴"""
        glDisable(GL_LIGHTING)
        glLineWidth(3)
        
        axis_length = min(self.container.length, self.container.width, self.container.height) * 0.2
        
        glBegin(GL_LINES)
        # X轴 - 红色
        glColor3f(1, 0.3, 0.3)
        glVertex3f(0, 0, 0)
        glVertex3f(axis_length, 0, 0)
        # Y轴 - 绿色 (高度)
        glColor3f(0.3, 1, 0.3)
        glVertex3f(0, 0, 0)
        glVertex3f(0, axis_length, 0)
        # Z轴 - 蓝色 (宽度)
        glColor3f(0.3, 0.3, 1)
        glVertex3f(0, 0, 0)
        glVertex3f(0, 0, axis_length)
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def draw_drag_axes(self, placed: PlacedCargo):
        """绘制拖拽轴"""
        glDisable(GL_LIGHTING)
        glLineWidth(4)
        
        # 货物中心位置
        cx = placed.x + placed.actual_length / 2
        cy = placed.z + placed.cargo.height / 2
        cz = placed.y + placed.actual_width / 2
        
        axis_len = max(placed.actual_length, placed.actual_width, placed.cargo.height) * 0.7
        
        glBegin(GL_LINES)
        # X轴 - 红色（长度方向）
        glColor3f(1, 0, 0)
        glVertex3f(cx - axis_len/2, cy, cz)
        glVertex3f(cx + axis_len/2, cy, cz)
        
        # Y轴 - 绿色（高度方向）
        glColor3f(0, 1, 0)
        glVertex3f(cx, cy - axis_len/2, cz)
        glVertex3f(cx, cy + axis_len/2, cz)
        
        # Z轴 - 蓝色（宽度方向）
        glColor3f(0, 0, 1)
        glVertex3f(cx, cy, cz - axis_len/2)
        glVertex3f(cx, cy, cz + axis_len/2)
        glEnd()
        
        glEnable(GL_LIGHTING)
    
    def hit_test(self, mouse_x: int, mouse_y: int) -> int:
        """碰撞检测 - 使用颜色拾取返回点击位置的货物索引，-1表示未命中"""
        if not self.placed_cargos or not self.container:
            return -1
        
        # 使用颜色拾取方法 - 更可靠
        self.makeCurrent()
        
        # 保存当前状态
        glPushAttrib(GL_ALL_ATTRIB_BITS)
        
        # 清除缓冲区
        glClearColor(0, 0, 0, 1)
        glClear(GL_COLOR_BUFFER_BIT | GL_DEPTH_BUFFER_BIT)
        glDisable(GL_LIGHTING)
        glDisable(GL_BLEND)
        glDisable(GL_DITHER)
        glDisable(GL_TEXTURE_2D)
        
        # 设置视图变换 (与 paintGL_single 保持完全一致！)
        glLoadIdentity()
        max_dim = max(self.container.length, self.container.width, self.container.height)
        distance = max_dim * 1.8 / self.zoom  # 必须与 paintGL_single 一致
        
        glTranslatef(self.pan_x, self.pan_y, -distance)
        glRotatef(self.rotation_x, 1, 0, 0)
        glRotatef(self.rotation_y, 0, 1, 0)
        glTranslatef(-self.container.length/2, -self.container.height/2, -self.container.width/2)
        
        # 用唯一颜色绘制每个货物
        for i, placed in enumerate(self.placed_cargos):
            # 将索引编码为颜色 (索引+1，因为0是背景)
            idx = i + 1
            r = (idx & 0xFF) / 255.0
            g = ((idx >> 8) & 0xFF) / 255.0
            b = ((idx >> 16) & 0xFF) / 255.0
            glColor3f(r, g, b)
            self.draw_cargo_for_picking(placed)
        
        glFlush()
        glFinish()
        
        # 读取鼠标位置的像素颜色
        viewport = glGetIntegerv(GL_VIEWPORT)
        pixel_y = viewport[3] - mouse_y  # OpenGL Y轴翻转
        
        pixel = glReadPixels(mouse_x, pixel_y, 1, 1, GL_RGB, GL_UNSIGNED_BYTE)
        
        # 恢复状态
        glPopAttrib()
        
        # 重新绘制正常场景
        self.update()
        
        # 解码颜色为索引
        if pixel:
            r, g, b = pixel[0], pixel[1], pixel[2]
            idx = r + (g << 8) + (b << 16)
            if idx > 0 and idx <= len(self.placed_cargos):
                return idx - 1
        
        return -1
    
    def draw_cargo_for_picking(self, placed: PlacedCargo):
        """绘制用于拾取的货物（简化版）"""
        x, y, z = placed.x, placed.z, placed.y
        l = placed.actual_length
        h = placed.cargo.height
        w = placed.actual_width
        
        glBegin(GL_QUADS)
        # 简单绘制六个面
        # 底面
        glVertex3f(x, y, z); glVertex3f(x+l, y, z); glVertex3f(x+l, y, z+w); glVertex3f(x, y, z+w)
        # 顶面
        glVertex3f(x, y+h, z); glVertex3f(x, y+h, z+w); glVertex3f(x+l, y+h, z+w); glVertex3f(x+l, y+h, z)
        # 前面
        glVertex3f(x, y, z); glVertex3f(x, y+h, z); glVertex3f(x+l, y+h, z); glVertex3f(x+l, y, z)
        # 后面
        glVertex3f(x, y, z+w); glVertex3f(x+l, y, z+w); glVertex3f(x+l, y+h, z+w); glVertex3f(x, y+h, z+w)
        # 左面
        glVertex3f(x, y, z); glVertex3f(x, y, z+w); glVertex3f(x, y+h, z+w); glVertex3f(x, y+h, z)
        # 右面
        glVertex3f(x+l, y, z); glVertex3f(x+l, y+h, z); glVertex3f(x+l, y+h, z+w); glVertex3f(x+l, y, z+w)
        glEnd()
    
    def mousePressEvent(self, event):
        """鼠标按下"""
        self.last_mouse_pos = event.pos()
        self.mouse_button = event.button()
        
        # 左键点击尝试选择货物（无论是否在拖拽模式）
        if event.button() == Qt.MouseButton.LeftButton:
            try:
                hit_index = self.hit_test(event.pos().x(), event.pos().y())
                if hit_index >= 0:
                    self.selected_cargo_index = hit_index
                    # 拖拽模式下才启用拖动
                    if self.drag_mode:
                        self.dragging = True
                        self.drag_start_pos = event.pos()
                    # 无论是否拖拽模式都触发选中回调
                    if self.on_cargo_selected:
                        self.on_cargo_selected(hit_index)
                    self.update()
                else:
                    self.selected_cargo_index = -1
                    self.update()
            except Exception:
                # 如果选择失败，忽略错误
                pass
    
    def mouseMoveEvent(self, event):
        """鼠标移动"""
        if self.last_mouse_pos is None:
            return
        
        dx = event.pos().x() - self.last_mouse_pos.x()
        dy = event.pos().y() - self.last_mouse_pos.y()
        
        # 拖拽模式下的移动逻辑
        if self.drag_mode and self.dragging and self.selected_cargo_index >= 0:
            if self.selected_cargo_index < len(self.placed_cargos):
                placed = self.placed_cargos[self.selected_cargo_index]
                # 简单的移动：水平移动改变X，垂直移动按Shift键改变Z，否则改变Y
                move_scale = self.container.length / 500  # 移动比例
                
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.KeyboardModifier.ShiftModifier:
                    # Shift + 拖动改变高度
                    new_z = placed.z - dy * move_scale
                    new_z = max(0, min(self.container.height - placed.cargo.height, new_z))
                    
                    # 吸附和碰撞检测
                    snap_x, snap_y, snap_z = self.find_snap_position(placed, placed.x, placed.y, new_z, self.selected_cargo_index)
                    if not self.collision_enabled or not self.check_collision(placed, placed.x, placed.y, snap_z, self.selected_cargo_index):
                        placed.z = snap_z
                else:
                    # 正常拖动改变X和Y
                    new_x = placed.x + dx * move_scale
                    new_y = placed.y + dy * move_scale
                    
                    # 边界检查
                    new_x = max(0, min(self.container.length - placed.actual_length, new_x))
                    new_y = max(0, min(self.container.width - placed.actual_width, new_y))
                    
                    # 吸附
                    snap_x, snap_y, snap_z = self.find_snap_position(placed, new_x, new_y, placed.z, self.selected_cargo_index)
                    
                    # 碰撞检测
                    if not self.collision_enabled or not self.check_collision(placed, snap_x, snap_y, placed.z, self.selected_cargo_index):
                        placed.x = snap_x
                        placed.y = snap_y
                    elif not self.check_collision(placed, new_x, new_y, placed.z, self.selected_cargo_index):
                        # 如果吸附位置有碰撞，使用原始位置
                        placed.x = new_x
                        placed.y = new_y
                    # 如果都有碰撞，不移动
                
                self.last_mouse_pos = event.pos()
                self.update()
                return
        
        if self.mouse_button == Qt.MouseButton.LeftButton and not self.drag_mode:
            # 左键拖动 - 旋转（更平滑）
            self.rotation_y += dx * 0.3
            self.rotation_x += dy * 0.3
            self.rotation_x = max(-90, min(90, self.rotation_x))
        elif self.mouse_button == Qt.MouseButton.RightButton:
            # 右键拖动 - 平移（根据缩放级别调整速度）
            pan_speed = 0.5 / self.zoom
            self.pan_x += dx * pan_speed
            self.pan_y -= dy * pan_speed
        elif self.mouse_button == Qt.MouseButton.MiddleButton:
            # 中键拖动 - 缩放
            self.zoom *= 1 + dy * 0.003
            self.zoom = max(0.1, min(10, self.zoom))
        
        self.last_mouse_pos = event.pos()
        self.update()
    
    def mouseReleaseEvent(self, event):
        """鼠标释放"""
        if self.dragging and self.on_cargo_moved:
            self.on_cargo_moved(self.selected_cargo_index)
        
        self.last_mouse_pos = None
        self.mouse_button = None
        self.dragging = False
    
    def keyPressEvent(self, event):
        """键盘事件处理"""
        if not self.drag_mode:
            return
        
        # R键旋转货物
        if event.key() == Qt.Key.Key_R:
            if self.rotate_selected_cargo():
                self.update()
            return
        
        # 方向键微调 (1cm)
        step = 1.0
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            step = 10.0  # Ctrl + 方向键 10cm 步进
        
        moved = False
        if event.key() == Qt.Key.Key_Left:
            moved = self.move_selected_cargo(-step, 0, 0)
        elif event.key() == Qt.Key.Key_Right:
            moved = self.move_selected_cargo(step, 0, 0)
        elif event.key() == Qt.Key.Key_Up:
            moved = self.move_selected_cargo(0, -step, 0)
        elif event.key() == Qt.Key.Key_Down:
            moved = self.move_selected_cargo(0, step, 0)
        elif event.key() == Qt.Key.Key_PageUp:
            moved = self.move_selected_cargo(0, 0, step)
        elif event.key() == Qt.Key.Key_PageDown:
            moved = self.move_selected_cargo(0, 0, -step)
        
        if moved:
            # 更新选中货物信息显示
            if self.on_cargo_selected:
                self.on_cargo_selected(self.selected_cargo_index)

    def wheelEvent(self, event):
        """鼠标滚轮 - 平滑缩放"""
        delta = event.angleDelta().y()
        # 使用更平滑的缩放因子
        zoom_factor = 1 + delta * 0.0008
        self.zoom *= zoom_factor
        self.zoom = max(0.1, min(10, self.zoom))  # 允许更大的缩放范围
        self.update()
    
    def reset_view(self):
        """重置视角"""
        self.rotation_x = 25
        self.rotation_y = 45
        self.zoom = 1.0
        self.pan_x = 0
        self.pan_y = 0
        self.update()
    
    def set_view(self, preset: str):
        """设置预设视角"""
        views = {
            "front": (0, 0),
            "back": (0, 180),
            "left": (0, -90),
            "right": (0, 90),
            "top": (90, 0),
            "iso": (25, 45),
        }
        if preset in views:
            self.rotation_x, self.rotation_y = views[preset]
            self.update()


class ModernButton(QPushButton):
    """现代风格按钮"""
    def __init__(self, text, primary=False, parent=None):
        super().__init__(text, parent)
        self.setMinimumHeight(36)
        if primary:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background-color: #1976D2;
                }
                QPushButton:pressed {
                    background-color: #1565C0;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #37474F;
                    color: white;
                    border: 1px solid #546E7A;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background-color: #455A64;
                    border-color: #78909C;
                }
                QPushButton:pressed {
                    background-color: #263238;
                }
            """)


class LoadingImageGenerator:
    """装载图生成器 - 支持中文和多视角"""
    
    def __init__(self, container: Container, placed_cargos: List[PlacedCargo], view_3d: 'Container3DView' = None):
        self.container = container
        self.placed_cargos = placed_cargos
        self.view_3d = view_3d  # 3D视图引用，用于截图
        self.margin = 60  # 边距
        self.scale = 1.0  # 比例尺
        self.font = None
        self.title_font = None
        self._load_fonts()
    
    def _load_fonts(self):
        """加载中文字体"""
        if not PIL_SUPPORT:
            return
        
        # 尝试加载中文字体
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",  # 微软雅黑
            "C:/Windows/Fonts/simsun.ttc",  # 宋体
            "C:/Windows/Fonts/simhei.ttf",  # 黑体
            "/System/Library/Fonts/PingFang.ttc",  # macOS
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",  # Linux
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    self.font = ImageFont.truetype(font_path, 12)
                    self.title_font = ImageFont.truetype(font_path, 16)
                    return
                except:
                    continue
        
        # 使用默认字体
        try:
            self.font = ImageFont.truetype("arial.ttf", 12)
            self.title_font = ImageFont.truetype("arial.ttf", 16)
        except:
            self.font = ImageFont.load_default()
            self.title_font = self.font
    
    def calculate_scale(self, max_width: int, max_height: int, container_dim1: float, container_dim2: float):
        """计算适合图像尺寸的比例尺"""
        available_width = max_width - 2 * self.margin
        available_height = max_height - 2 * self.margin
        
        scale_x = available_width / container_dim1 if container_dim1 > 0 else 1
        scale_y = available_height / container_dim2 if container_dim2 > 0 else 1
        
        return min(scale_x, scale_y)
    
    def generate_top_view(self, width: int = 800, height: int = 600) -> Optional['Image.Image']:
        """生成俯视图（X-Y平面，从上往下看）"""
        if not PIL_SUPPORT:
            return None
        
        self.scale = self.calculate_scale(width, height, self.container.length, self.container.width)
        
        img = Image.new('RGB', (width, height), color=(240, 240, 245))
        draw = ImageDraw.Draw(img)
        
        # 绘制容器轮廓
        container_x = self.margin
        container_y = self.margin
        container_w = int(self.container.length * self.scale)
        container_h = int(self.container.width * self.scale)
        
        draw.rectangle([container_x, container_y, container_x + container_w, container_y + container_h],
                      outline=(100, 100, 100), width=3)
        
        # 绘制货物（按高度排序，底层的先画）
        sorted_cargos = sorted(self.placed_cargos, key=lambda p: p.z)
        
        for placed in sorted_cargos:
            x = container_x + int(placed.x * self.scale)
            y = container_y + int(placed.y * self.scale)
            w = int(placed.actual_length * self.scale)
            h = int(placed.actual_width * self.scale)
            
            r, g, b = placed.cargo.color
            color = (int(r * 255), int(g * 255), int(b * 255))
            
            # 绘制货物矩形
            draw.rectangle([x, y, x + w, y + h], fill=color, outline=(50, 50, 50), width=1)
            
            # 添加货物名称（如果空间足够）
            if w > 40 and h > 20:
                text = placed.cargo.name[:6]
                draw.text((x + 3, y + 3), text, fill=(255, 255, 255), font=self.font)
        
        # 添加标题
        draw.text((10, 10), "俯视图 (Top View)", fill=(50, 50, 50), font=self.title_font)
        
        # 添加尺寸标注
        draw.text((container_x, height - 30), f"长度: {self.container.length}cm", fill=(80, 80, 80), font=self.font)
        draw.text((width - 180, container_y + container_h + 10), f"宽度: {self.container.width}cm", fill=(80, 80, 80), font=self.font)
        
        return img
    
    def generate_front_view(self, width: int = 800, height: int = 600) -> Optional['Image.Image']:
        """生成正视图（X-Z平面，从前往后看）"""
        if not PIL_SUPPORT:
            return None
        
        self.scale = self.calculate_scale(width, height, self.container.length, self.container.height)
        
        img = Image.new('RGB', (width, height), color=(240, 240, 245))
        draw = ImageDraw.Draw(img)
        
        # 绘制容器轮廓
        container_x = self.margin
        container_y = height - self.margin - int(self.container.height * self.scale)
        container_w = int(self.container.length * self.scale)
        container_h = int(self.container.height * self.scale)
        
        draw.rectangle([container_x, container_y, container_x + container_w, container_y + container_h],
                      outline=(100, 100, 100), width=3)
        
        # 绘制货物
        for placed in self.placed_cargos:
            x = container_x + int(placed.x * self.scale)
            y = container_y + container_h - int((placed.z + placed.cargo.height) * self.scale)
            w = int(placed.actual_length * self.scale)
            h = int(placed.cargo.height * self.scale)
            
            r, g, b = placed.cargo.color
            color = (int(r * 255), int(g * 255), int(b * 255))
            
            draw.rectangle([x, y, x + w, y + h], fill=color, outline=(50, 50, 50), width=1)
        
        # 添加标题
        draw.text((10, 10), "正视图 (Front View)", fill=(50, 50, 50), font=self.title_font)
        draw.text((container_x, height - 30), f"长度: {self.container.length}cm", fill=(80, 80, 80), font=self.font)
        draw.text((10, container_y - 25), f"高度: {self.container.height}cm", fill=(80, 80, 80), font=self.font)
        
        return img
    
    def generate_side_view(self, width: int = 800, height: int = 600) -> Optional['Image.Image']:
        """生成侧视图（Y-Z平面，从左往右看）"""
        if not PIL_SUPPORT:
            return None
        
        self.scale = self.calculate_scale(width, height, self.container.width, self.container.height)
        
        img = Image.new('RGB', (width, height), color=(240, 240, 245))
        draw = ImageDraw.Draw(img)
        
        # 绘制容器轮廓
        container_x = self.margin
        container_y = height - self.margin - int(self.container.height * self.scale)
        container_w = int(self.container.width * self.scale)
        container_h = int(self.container.height * self.scale)
        
        draw.rectangle([container_x, container_y, container_x + container_w, container_y + container_h],
                      outline=(100, 100, 100), width=3)
        
        # 绘制货物
        for placed in self.placed_cargos:
            x = container_x + int(placed.y * self.scale)
            y = container_y + container_h - int((placed.z + placed.cargo.height) * self.scale)
            w = int(placed.actual_width * self.scale)
            h = int(placed.cargo.height * self.scale)
            
            r, g, b = placed.cargo.color
            color = (int(r * 255), int(g * 255), int(b * 255))
            
            draw.rectangle([x, y, x + w, y + h], fill=color, outline=(50, 50, 50), width=1)
        
        # 添加标题
        draw.text((10, 10), "侧视图 (Side View)", fill=(50, 50, 50), font=self.title_font)
        draw.text((container_x, height - 30), f"宽度: {self.container.width}cm", fill=(80, 80, 80), font=self.font)
        draw.text((10, container_y - 25), f"高度: {self.container.height}cm", fill=(80, 80, 80), font=self.font)
        
        return img
    
    def generate_isometric_view(self, width: int = 800, height: int = 600) -> Optional['Image.Image']:
        """生成等轴测视图（使用OpenGL截图）"""
        if not PIL_SUPPORT:
            return None
        
        # 如果有3D视图引用，使用OpenGL截图
        if self.view_3d is not None:
            try:
                # 使用OpenGL截图
                qimage = self.view_3d.capture_isometric_image(width, height)
                
                # 将QImage转换为PIL Image
                qimage = qimage.convertToFormat(qimage.Format.Format_RGB888)
                ptr = qimage.bits()
                ptr.setsize(qimage.sizeInBytes())
                
                img = Image.frombytes('RGB', (qimage.width(), qimage.height()), bytes(ptr))
                
                # 添加标题和尺寸信息
                draw = ImageDraw.Draw(img)
                
                # 绘制半透明背景
                title_bg = Image.new('RGBA', (width, 40), (240, 240, 245, 220))
                img_rgba = img.convert('RGBA')
                img_rgba.paste(title_bg, (0, 0), title_bg)
                
                footer_bg = Image.new('RGBA', (width, 35), (240, 240, 245, 220))
                img_rgba.paste(footer_bg, (0, height - 35), footer_bg)
                
                img = img_rgba.convert('RGB')
                draw = ImageDraw.Draw(img)
                
                L, W, H = self.container.length, self.container.width, self.container.height
                draw.text((10, 10), "等轴测视图 (Isometric View)", fill=(50, 50, 50), font=self.title_font)
                draw.text((10, height - 30), f"尺寸: {L} × {W} × {H} cm", fill=(80, 80, 80), font=self.font)
                
                return img
            except Exception as e:
                print(f"OpenGL截图失败，回退到PIL绘制: {e}")
        
        # 回退到PIL手动绘制
        return self._generate_isometric_view_pil(width, height)
    
    def _generate_isometric_view_pil(self, width: int = 800, height: int = 600) -> Optional['Image.Image']:
        """使用PIL手动绘制等轴测视图（备用方法）"""
        img = Image.new('RGB', (width, height), color=(240, 240, 245))
        draw = ImageDraw.Draw(img)
        
        # 等轴测角度
        angle = math.radians(30)
        cos_a = math.cos(angle)
        sin_a = math.sin(angle)
        
        # 容器尺寸
        L, W, H = self.container.length, self.container.width, self.container.height
        
        # 计算投影后的边界框
        # 等轴测投影: px = (x - y) * cos(30), py = -(x + y) * sin(30) - z
        corners = [
            (0, 0, 0), (L, 0, 0), (L, W, 0), (0, W, 0),
            (0, 0, H), (L, 0, H), (L, W, H), (0, W, H)
        ]
        
        proj_x = [(c[0] - c[1]) * cos_a for c in corners]
        proj_y = [-(c[0] + c[1]) * sin_a - c[2] for c in corners]
        
        min_px, max_px = min(proj_x), max(proj_x)
        min_py, max_py = min(proj_y), max(proj_y)
        
        proj_width = max_px - min_px
        proj_height = max_py - min_py
        
        # 计算缩放比例，留边距
        margin = 60
        scale = min((width - 2 * margin) / proj_width, (height - 2 * margin) / proj_height)
        
        # 计算居中偏移
        cx = width / 2 - (min_px + max_px) / 2 * scale
        cy = height / 2 - (min_py + max_py) / 2 * scale
        
        def project(x, y, z):
            """等轴测投影"""
            px = (x - y) * cos_a * scale + cx
            py = -(x + y) * sin_a * scale - z * scale + cy
            return int(px), int(py)
        
        # 绘制容器边框（线框）
        container_color = (100, 100, 110)
        
        # 底面
        p0 = project(0, 0, 0)
        p1 = project(L, 0, 0)
        p2 = project(L, W, 0)
        p3 = project(0, W, 0)
        draw.line([p0, p1], fill=container_color, width=2)
        draw.line([p1, p2], fill=container_color, width=2)
        draw.line([p2, p3], fill=container_color, width=2)
        draw.line([p3, p0], fill=container_color, width=2)
        
        # 顶面
        p4 = project(0, 0, H)
        p5 = project(L, 0, H)
        p6 = project(L, W, H)
        p7 = project(0, W, H)
        draw.line([p4, p5], fill=container_color, width=2)
        draw.line([p5, p6], fill=container_color, width=2)
        draw.line([p6, p7], fill=container_color, width=2)
        draw.line([p7, p4], fill=container_color, width=2)
        
        # 竖直边
        draw.line([p0, p4], fill=container_color, width=2)
        draw.line([p1, p5], fill=container_color, width=2)
        draw.line([p2, p6], fill=container_color, width=2)
        draw.line([p3, p7], fill=container_color, width=2)
        
        # 绘制货物（按深度排序 - painter's algorithm）
        # 等轴测视角从右前上方看，需要先画左后下的货物
        # 排序依据：x小、y小的在后面先画；同位置时z小的先画
        sorted_cargos = sorted(self.placed_cargos, 
                               key=lambda p: (p.x + p.y + p.z * 0.5))
        
        for placed in sorted_cargos:
            x, y, z = placed.x, placed.y, placed.z
            l = placed.actual_length
            w = placed.actual_width
            h = placed.cargo.height
            
            r, g, b = placed.cargo.color
            color = (int(r * 255), int(g * 255), int(b * 255))
            darker = (int(r * 200), int(g * 200), int(b * 200))
            darkest = (int(r * 160), int(g * 160), int(b * 160))
            
            # 货物的8个顶点
            # 底面四点
            v0 = project(x, y, z)          # 左后下
            v1 = project(x + l, y, z)      # 右后下
            v2 = project(x + l, y + w, z)  # 右前下
            v3 = project(x, y + w, z)      # 左前下
            # 顶面四点
            v4 = project(x, y, z + h)          # 左后上
            v5 = project(x + l, y, z + h)      # 右后上
            v6 = project(x + l, y + w, z + h)  # 右前上
            v7 = project(x, y + w, z + h)      # 左前上
            
            # 从右前上方看，可见三个面：顶面、右面(x=x+l)、前面(y=y+w)
            # 按painter算法，先画被遮挡的面
            
            # 右面 (x = x+l 那一面) - 中等亮度
            draw.polygon([v1, v2, v6, v5], fill=darker, outline=(30, 30, 30))
            # 前面 (y = y+w 那一面) - 最暗
            draw.polygon([v3, v2, v6, v7], fill=darkest, outline=(30, 30, 30))
            # 顶面 (z = z+h 那一面) - 最亮，最后画
            draw.polygon([v4, v5, v6, v7], fill=color, outline=(30, 30, 30))
        
        # 添加标题
        draw.text((10, 10), "等轴测视图 (Isometric View)", fill=(50, 50, 50), font=self.title_font)
        draw.text((10, height - 30), f"尺寸: {L} × {W} × {H} cm", fill=(80, 80, 80), font=self.font)
        
        return img
    
    def generate_combined_view(self, width: int = 1200, height: int = 900) -> Optional['Image.Image']:
        """生成组合视图（四视图合一：俯视、正视、侧视、等轴测）"""
        if not PIL_SUPPORT:
            return None
        
        # 计算子图尺寸
        sub_width = width // 2 - 20
        sub_height = height // 2 - 20
        
        combined = Image.new('RGB', (width, height), color=(255, 255, 255))
        
        # 生成四个视图
        top_view = self.generate_top_view(sub_width, sub_height)
        front_view = self.generate_front_view(sub_width, sub_height)
        side_view = self.generate_side_view(sub_width, sub_height)
        iso_view = self.generate_isometric_view(sub_width, sub_height)
        
        # 拼接
        if top_view:
            combined.paste(top_view, (10, 10))
        if front_view:
            combined.paste(front_view, (sub_width + 20, 10))
        if side_view:
            combined.paste(side_view, (10, sub_height + 20))
        if iso_view:
            combined.paste(iso_view, (sub_width + 20, sub_height + 20))
        
        return combined
    
    def generate_summary_image(self, width: int = 1200, height: int = 800) -> Optional['Image.Image']:
        """生成带统计信息的综合图"""
        if not PIL_SUPPORT:
            return None
        
        img = Image.new('RGB', (width, height), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)
        
        # 左侧放等轴测视图
        iso_width = width * 2 // 3 - 20
        iso_height = height - 40
        iso_view = self.generate_isometric_view(iso_width, iso_height)
        if iso_view:
            img.paste(iso_view, (10, 20))
        
        # 右侧放统计信息
        stats_x = iso_width + 30
        stats_y = 30
        stats_w = width - stats_x - 20
        
        # 绘制统计信息背景
        draw.rectangle([stats_x, stats_y, width - 20, height - 20],
                      fill=(248, 248, 250), outline=(200, 200, 210), width=2)
        
        # 标题
        y_offset = stats_y + 20
        draw.text((stats_x + 15, y_offset), "装载统计", fill=(50, 50, 50), font=self.title_font)
        y_offset += 40
        
        # 分隔线
        draw.line([(stats_x + 10, y_offset), (width - 30, y_offset)], fill=(200, 200, 210), width=1)
        y_offset += 15
        
        # 统计数据
        total_volume = sum(p.cargo.volume for p in self.placed_cargos)
        total_weight = sum(p.cargo.weight for p in self.placed_cargos)
        vol_util = (total_volume / self.container.volume) * 100 if self.container.volume > 0 else 0
        wt_util = (total_weight / self.container.max_weight) * 100 if self.container.max_weight > 0 else 0
        
        stats_items = [
            ("容器类型", self.container.name),
            ("容器尺寸", f"{self.container.length}×{self.container.width}×{self.container.height} cm"),
            ("容积", f"{self.container.volume_cbm:.1f} m³"),
            ("最大载重", f"{self.container.max_weight:,} kg"),
            ("", ""),  # 空行
            ("装载件数", f"{len(self.placed_cargos)} 件"),
            ("已用体积", f"{total_volume/1000000:.2f} m³"),
            ("空间利用率", f"{vol_util:.1f}%"),
            ("总重量", f"{total_weight:.1f} kg"),
            ("载重利用率", f"{wt_util:.1f}%"),
        ]
        
        for label, value in stats_items:
            if label:
                draw.text((stats_x + 15, y_offset), f"{label}:", fill=(100, 100, 100), font=self.font)
                draw.text((stats_x + 100, y_offset), str(value), fill=(50, 50, 50), font=self.font)
            y_offset += 28
        
        return img
    
    def save_images(self, base_path: str) -> List[str]:
        """保存所有视图图片"""
        saved_files = []
        
        views = [
            ('top', self.generate_top_view),
            ('front', self.generate_front_view),
            ('side', self.generate_side_view),
            ('isometric', self.generate_isometric_view),
            ('combined', self.generate_combined_view),
            ('summary', self.generate_summary_image),
        ]
        
        for name, generator in views:
            img = generator()
            if img:
                file_path = f"{base_path}_{name}.png"
                img.save(file_path)
                saved_files.append(file_path)
        
        return saved_files


class ContainerLoadingApp(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("集装箱配载软件 v0.5 - by Henry Xue")
        self.setMinimumSize(1500, 900)
        self.resize(1600, 1000)
        
        self.cargos: List[Cargo] = []
        self.cargo_groups: List[CargoGroup] = []
        self.container: Optional[Container] = None
        self.placed_cargos: List[PlacedCargo] = []
        self.color_index = 0
        self.loading_rules = DEFAULT_RULES.copy()
        self.custom_containers: dict = {}
        self.last_statistics: dict = {}
        
        # 多集装箱支持
        self.multi_container_mode = False
        self.container_results: List[ContainerLoadingResult] = []
        self.container_count = 1
        
        self.setup_style()
        self.setup_ui()
        self.setup_default_container()
    
    def setup_style(self):
        """设置应用样式"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1e1e1e;
            }
            QWidget {
                background-color: #1e1e1e;
                color: #e0e0e0;
                font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
            }
            QGroupBox {
                border: 1px solid #3d3d3d;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 10px;
                font-weight: bold;
                font-size: 13px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
                color: #81D4FA;
            }
            QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
                background-color: #2d2d2d;
                border: 1px solid #3d3d3d;
                border-radius: 6px;
                padding: 8px;
                color: #e0e0e0;
                font-size: 13px;
            }
            QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus {
                border-color: #2196F3;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #9e9e9e;
                margin-right: 10px;
            }
            QTableWidget {
                background-color: #252525;
                border: 1px solid #3d3d3d;
                border-radius: 6px;
                gridline-color: #3d3d3d;
            }
            QTableWidget::item {
                padding: 4px 2px;
            }
            QTableWidget::item:selected {
                background-color: #2196F3;
            }
            QTableWidget QLineEdit {
                background-color: #1e1e1e;
                color: #ffffff;
                border: 1px solid #2196F3;
                padding: 2px;
                selection-background-color: #2196F3;
            }
            QHeaderView::section {
                background-color: #2d2d2d;
                color: #81D4FA;
                padding: 10px;
                border: none;
                border-bottom: 1px solid #3d3d3d;
                font-weight: bold;
            }
            QProgressBar {
                border: none;
                border-radius: 6px;
                background-color: #2d2d2d;
                height: 20px;
                text-align: center;
            }
            QProgressBar::chunk {
                border-radius: 6px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2196F3, stop:1 #21CBF3);
            }
            QCheckBox {
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border-radius: 4px;
                border: 2px solid #546E7A;
            }
            QCheckBox::indicator:checked {
                background-color: #2196F3;
                border-color: #2196F3;
            }
            QLabel {
                font-size: 13px;
            }
            QScrollBar:vertical {
                background-color: #1e1e1e;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #3d3d3d;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #4d4d4d;
            }
        """)
    
    def setup_ui(self):
        """设置界面"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 左侧面板
        left_panel = QWidget()
        left_panel.setMinimumWidth(520)
        left_panel.setMaximumWidth(580)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(12)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        # 使用滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(12)
        
        # ==================== 容器选择 ====================
        container_group = QGroupBox("📦 容器选择")
        container_layout = QVBoxLayout(container_group)
        
        # 容器类别
        cat_layout = QHBoxLayout()
        cat_layout.addWidget(QLabel("类别:"))
        self.container_category = QComboBox()
        self.container_category.addItems(list(CONTAINER_CATEGORIES.keys()))
        self.container_category.currentTextChanged.connect(self.on_category_changed)
        cat_layout.addWidget(self.container_category)
        container_layout.addLayout(cat_layout)
        
        # 容器型号
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("型号:"))
        self.container_combo = QComboBox()
        self.container_combo.currentTextChanged.connect(self.on_container_selected)
        type_layout.addWidget(self.container_combo)
        container_layout.addLayout(type_layout)
        
        # 自定义容器按钮
        custom_btn_layout = QHBoxLayout()
        custom_btn = ModernButton("➕ 自定义容器")
        custom_btn.clicked.connect(self.show_custom_container_dialog)
        custom_btn_layout.addWidget(custom_btn)
        container_layout.addLayout(custom_btn_layout)
        
        # 容器信息
        self.container_info = QLabel()
        self.container_info.setStyleSheet("color: #9e9e9e; font-size: 12px;")
        self.container_info.setWordWrap(True)
        container_layout.addWidget(self.container_info)
        
        scroll_layout.addWidget(container_group)
        
        # ==================== 货物添加 ====================
        cargo_group = QGroupBox("📋 添加货物")
        cargo_layout = QVBoxLayout(cargo_group)
        
        # 货物名称
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("名称:"))
        self.cargo_name = QLineEdit("货物1")
        name_layout.addWidget(self.cargo_name)
        cargo_layout.addLayout(name_layout)
        
        # 尺寸输入
        size_layout = QHBoxLayout()
        size_layout.addWidget(QLabel("尺寸:"))
        self.cargo_length = QDoubleSpinBox()
        self.cargo_length.setRange(1, 10000)
        self.cargo_length.setValue(100)
        self.cargo_length.setSuffix(" cm")
        size_layout.addWidget(self.cargo_length)
        size_layout.addWidget(QLabel("×"))
        self.cargo_width = QDoubleSpinBox()
        self.cargo_width.setRange(1, 10000)
        self.cargo_width.setValue(80)
        self.cargo_width.setSuffix(" cm")
        size_layout.addWidget(self.cargo_width)
        size_layout.addWidget(QLabel("×"))
        self.cargo_height = QDoubleSpinBox()
        self.cargo_height.setRange(1, 10000)
        self.cargo_height.setValue(60)
        self.cargo_height.setSuffix(" cm")
        size_layout.addWidget(self.cargo_height)
        cargo_layout.addLayout(size_layout)
        
        # 重量和数量
        weight_layout = QHBoxLayout()
        weight_layout.addWidget(QLabel("重量:"))
        self.cargo_weight = QDoubleSpinBox()
        self.cargo_weight.setRange(0.1, 100000)
        self.cargo_weight.setValue(50)
        self.cargo_weight.setSuffix(" kg")
        weight_layout.addWidget(self.cargo_weight)
        weight_layout.addWidget(QLabel("数量:"))
        self.cargo_quantity = QSpinBox()
        self.cargo_quantity.setRange(1, 10000)
        self.cargo_quantity.setValue(10)
        weight_layout.addWidget(self.cargo_quantity)
        cargo_layout.addLayout(weight_layout)
        
        # 货物选项
        options_layout = QHBoxLayout()
        self.cargo_stackable = QCheckBox("可堆叠")
        self.cargo_stackable.setChecked(True)
        options_layout.addWidget(self.cargo_stackable)
        self.cargo_rotatable = QCheckBox("可旋转")
        self.cargo_rotatable.setChecked(True)
        options_layout.addWidget(self.cargo_rotatable)
        self.cargo_bottom_only = QCheckBox("仅底层")
        options_layout.addWidget(self.cargo_bottom_only)
        cargo_layout.addLayout(options_layout)
        
        # 优先级
        priority_layout = QHBoxLayout()
        priority_layout.addWidget(QLabel("优先级:"))
        self.cargo_priority = QSpinBox()
        self.cargo_priority.setRange(0, 100)
        self.cargo_priority.setValue(0)
        self.cargo_priority.setToolTip("数字越大优先级越高")
        priority_layout.addWidget(self.cargo_priority)
        priority_layout.addStretch()
        cargo_layout.addLayout(priority_layout)
        
        # 添加按钮
        add_btn = ModernButton("➕ 添加货物", primary=True)
        add_btn.clicked.connect(self.add_cargo)
        cargo_layout.addWidget(add_btn)
        
        scroll_layout.addWidget(cargo_group)
        
        # ==================== 货物列表 ====================
        list_group = QGroupBox("📜 货物列表")
        list_layout = QVBoxLayout(list_group)
        
        self.cargo_table = QTableWidget()
        self.cargo_table.setColumnCount(6)
        self.cargo_table.setHorizontalHeaderLabels(["名称", "尺寸(cm)", "重量", "数量", "选项", "体积"])
        # 设置各列宽度 - 全部固定宽度
        self.cargo_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.cargo_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # 尺寸列自动拉伸
        self.cargo_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.cargo_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.cargo_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.cargo_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.cargo_table.setColumnWidth(0, 60)   # 名称
        self.cargo_table.setColumnWidth(2, 60)   # 重量
        self.cargo_table.setColumnWidth(3, 35)   # 数量
        self.cargo_table.setColumnWidth(4, 50)   # 选项
        self.cargo_table.setColumnWidth(5, 45)   # 体积
        self.cargo_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.cargo_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.cargo_table.setAlternatingRowColors(True)
        self.cargo_table.setMinimumHeight(180)
        # 连接单元格编辑信号
        self.cargo_table.cellChanged.connect(self.on_cargo_table_cell_changed)
        list_layout.addWidget(self.cargo_table)
        
        # 列表操作按钮
        list_btn_layout = QHBoxLayout()
        del_btn = ModernButton("🗑 删除")
        del_btn.clicked.connect(self.delete_cargo)
        clear_btn = ModernButton("清空")
        clear_btn.clicked.connect(self.clear_cargos)
        import_btn = ModernButton("📥 导入")
        import_btn.clicked.connect(self.import_cargos)
        export_btn = ModernButton("📤 导出")
        export_btn.clicked.connect(self.export_cargos)
        
        list_btn_layout.addWidget(del_btn)
        list_btn_layout.addWidget(clear_btn)
        list_btn_layout.addWidget(import_btn)
        list_btn_layout.addWidget(export_btn)
        list_layout.addLayout(list_btn_layout)
        
        # 货物组操作
        group_btn_layout = QHBoxLayout()
        create_group_btn = ModernButton("🔗 创建组")
        create_group_btn.clicked.connect(self.create_cargo_group)
        create_group_btn.setToolTip("将选中的货物组合为一组")
        group_btn_layout.addWidget(create_group_btn)
        ungroup_btn = ModernButton("解除组")
        ungroup_btn.clicked.connect(self.ungroup_cargo)
        group_btn_layout.addWidget(ungroup_btn)
        list_layout.addLayout(group_btn_layout)
        
        scroll_layout.addWidget(list_group)
        
        # ==================== 配载规则 ====================
        rules_group = QGroupBox("📐 配载规则")
        rules_layout = QVBoxLayout(rules_group)
        
        # 规则列表
        self.rules_list = QTableWidget()
        self.rules_list.setColumnCount(3)
        self.rules_list.setHorizontalHeaderLabels(["启用", "规则", "优先级"])
        self.rules_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.rules_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.rules_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.rules_list.setColumnWidth(0, 50)
        self.rules_list.setColumnWidth(2, 60)
        self.rules_list.setMaximumHeight(150)
        self.setup_rules_table()
        rules_layout.addWidget(self.rules_list)
        
        scroll_layout.addWidget(rules_group)
        
        # ==================== 配载操作 ====================
        action_group = QGroupBox("⚙️ 配载操作")
        action_layout = QVBoxLayout(action_group)
        
        # 多集装箱模式
        multi_layout = QHBoxLayout()
        self.multi_container_check = QCheckBox("多集装箱模式")
        self.multi_container_check.setChecked(False)
        self.multi_container_check.stateChanged.connect(self.toggle_multi_container_mode)
        multi_layout.addWidget(self.multi_container_check)
        
        multi_layout.addWidget(QLabel("数量:"))
        self.container_count_spin = QSpinBox()
        self.container_count_spin.setRange(1, 100)
        self.container_count_spin.setValue(1)
        self.container_count_spin.setEnabled(False)
        multi_layout.addWidget(self.container_count_spin)
        action_layout.addLayout(multi_layout)
        
        start_btn = ModernButton("🚀 开始配载", primary=True)
        start_btn.clicked.connect(self.start_loading)
        action_layout.addWidget(start_btn)
        
        # 拖拽调整模式
        drag_layout = QHBoxLayout()
        self.drag_mode_btn = ModernButton("🎯 拖拽调整模式")
        self.drag_mode_btn.setCheckable(True)
        self.drag_mode_btn.clicked.connect(self.toggle_drag_mode)
        self.drag_mode_btn.setToolTip("开启后可在3D视图中直接拖拽调整货物位置\n左键点击选中，拖动移动，Shift+拖动改变高度\nR键旋转货物，方向键微调(1cm)，Ctrl+方向键(10cm)")
        drag_layout.addWidget(self.drag_mode_btn)
        action_layout.addLayout(drag_layout)
        
        # 拖拽模式辅助控制按钮
        drag_control_layout = QHBoxLayout()
        
        # 旋转按钮
        self.rotate_cargo_btn = ModernButton("🔄 旋转")
        self.rotate_cargo_btn.clicked.connect(self.rotate_selected_cargo_from_btn)
        self.rotate_cargo_btn.setToolTip("旋转选中的货物 (快捷键: R)")
        self.rotate_cargo_btn.setEnabled(False)
        drag_control_layout.addWidget(self.rotate_cargo_btn)
        
        # 碰撞检测开关
        self.collision_check = QCheckBox("碰撞检测")
        self.collision_check.setChecked(True)
        self.collision_check.stateChanged.connect(self.toggle_collision_detection)
        self.collision_check.setToolTip("开启后移动货物时防止与其他货物重叠")
        drag_control_layout.addWidget(self.collision_check)
        action_layout.addLayout(drag_control_layout)
        
        # 微调按钮组
        fine_tune_layout = QHBoxLayout()
        fine_tune_label = QLabel("微调:")
        fine_tune_layout.addWidget(fine_tune_label)
        
        # 1cm 微调按钮
        self.step_1cm_btns = {}
        btn_x_minus = ModernButton("X-")
        btn_x_minus.setFixedWidth(35)
        btn_x_minus.clicked.connect(lambda: self.fine_tune_cargo(-1, 0, 0))
        fine_tune_layout.addWidget(btn_x_minus)
        
        btn_x_plus = ModernButton("X+")
        btn_x_plus.setFixedWidth(35)
        btn_x_plus.clicked.connect(lambda: self.fine_tune_cargo(1, 0, 0))
        fine_tune_layout.addWidget(btn_x_plus)
        
        btn_y_minus = ModernButton("Y-")
        btn_y_minus.setFixedWidth(35)
        btn_y_minus.clicked.connect(lambda: self.fine_tune_cargo(0, -1, 0))
        fine_tune_layout.addWidget(btn_y_minus)
        
        btn_y_plus = ModernButton("Y+")
        btn_y_plus.setFixedWidth(35)
        btn_y_plus.clicked.connect(lambda: self.fine_tune_cargo(0, 1, 0))
        fine_tune_layout.addWidget(btn_y_plus)
        
        btn_z_minus = ModernButton("Z-")
        btn_z_minus.setFixedWidth(35)
        btn_z_minus.clicked.connect(lambda: self.fine_tune_cargo(0, 0, -1))
        fine_tune_layout.addWidget(btn_z_minus)
        
        btn_z_plus = ModernButton("Z+")
        btn_z_plus.setFixedWidth(35)
        btn_z_plus.clicked.connect(lambda: self.fine_tune_cargo(0, 0, 1))
        fine_tune_layout.addWidget(btn_z_plus)
        
        action_layout.addLayout(fine_tune_layout)
        
        # 步进大小选择
        step_layout = QHBoxLayout()
        step_layout.addWidget(QLabel("步进:"))
        self.step_size_combo = QComboBox()
        self.step_size_combo.addItems(["1 cm", "5 cm", "10 cm", "20 cm"])
        self.step_size_combo.setCurrentIndex(0)
        self.step_size_combo.setToolTip("设置微调按钮的移动距离")
        step_layout.addWidget(self.step_size_combo)
        step_layout.addStretch()
        action_layout.addLayout(step_layout)
        
        manual_btn = ModernButton("✋ 精确调整")
        manual_btn.clicked.connect(self.enable_manual_edit)
        manual_btn.setToolTip("配载后通过对话框精确调整货物位置")
        action_layout.addWidget(manual_btn)
        
        clear_result_btn = ModernButton("清除结果")
        clear_result_btn.clicked.connect(self.clear_loading)
        action_layout.addWidget(clear_result_btn)
        
        # 导出选项
        export_layout = QHBoxLayout()
        export_plan_btn = ModernButton("📋 导出方案")
        export_plan_btn.clicked.connect(self.export_loading_plan)
        export_layout.addWidget(export_plan_btn)
        
        export_image_btn = ModernButton("🖼️ 导出图片")
        export_image_btn.clicked.connect(self.export_loading_images)
        export_image_btn.setToolTip("导出装载图（俯视图、正视图、侧视图）")
        export_layout.addWidget(export_image_btn)
        action_layout.addLayout(export_layout)
        
        # 加固建议按钮
        securing_btn = ModernButton("🔧 查看加固建议")
        securing_btn.clicked.connect(self.show_securing_advice_dialog)
        securing_btn.setToolTip("根据配载结果分析，给出智能加固建议")
        action_layout.addWidget(securing_btn)
        
        scroll_layout.addWidget(action_group)
        
        # ==================== 两步装载 ====================
        twostep_group = QGroupBox("📦 两步装载（先组托再装柜）")
        twostep_layout = QVBoxLayout(twostep_group)
        
        palletize_btn = ModernButton("第一步: 货物组托")
        palletize_btn.clicked.connect(self.palletize_cargos)
        palletize_btn.setToolTip("将小箱先组到托盘上")
        twostep_layout.addWidget(palletize_btn)
        
        load_pallets_btn = ModernButton("第二步: 托盘装柜")
        load_pallets_btn.clicked.connect(self.load_pallets_to_container)
        load_pallets_btn.setToolTip("将托盘装入集装箱")
        twostep_layout.addWidget(load_pallets_btn)
        
        scroll_layout.addWidget(twostep_group)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_content)
        left_layout.addWidget(scroll)
        
        # ==================== 右侧面板 ====================
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(12)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        # 多集装箱选择器
        self.container_selector_group = QGroupBox("📦 集装箱选择")
        container_selector_layout = QHBoxLayout(self.container_selector_group)
        
        container_selector_layout.addWidget(QLabel("当前查看:"))
        self.container_selector = QComboBox()
        self.container_selector.addItem("全部概览")
        self.container_selector.currentIndexChanged.connect(self.on_container_selector_changed)
        container_selector_layout.addWidget(self.container_selector, 1)
        
        self.container_selector_group.setVisible(False)  # 默认隐藏，多集装箱模式时显示
        right_layout.addWidget(self.container_selector_group)
        
        # 3D视图
        view_group = QGroupBox("🎮 3D配载视图 (左键旋转 | 滚轮缩放 | 右键平移)")
        view_layout = QVBoxLayout(view_group)
        
        self.gl_widget = Container3DView()
        # 设置拖拽回调
        self.gl_widget.on_cargo_selected = self.on_cargo_drag_selected
        self.gl_widget.on_cargo_moved = self.on_cargo_drag_moved
        view_layout.addWidget(self.gl_widget)
        
        # 视图控制按钮
        view_btn_layout = QHBoxLayout()
        
        views = [("正视", "front"), ("后视", "back"), ("左视", "left"), 
                 ("右视", "right"), ("俯视", "top"), ("等轴", "iso")]
        for name, preset in views:
            btn = ModernButton(name)
            btn.setFixedWidth(60)
            btn.clicked.connect(lambda checked, p=preset: self.gl_widget.set_view(p))
            view_btn_layout.addWidget(btn)
        
        view_btn_layout.addStretch()
        
        reset_btn = ModernButton("🔄 重置视图")
        reset_btn.clicked.connect(self.gl_widget.reset_view)
        view_btn_layout.addWidget(reset_btn)
        
        # 全屏按钮
        fullscreen_btn = ModernButton("⛶ 全屏")
        fullscreen_btn.setStyleSheet("background-color: #1565C0; color: white; font-weight: bold;")
        fullscreen_btn.clicked.connect(self.show_fullscreen_3d_view)
        view_btn_layout.addWidget(fullscreen_btn)
        
        # 使用手册按钮
        help_btn = ModernButton("❓ 使用手册")
        help_btn.setStyleSheet("background-color: #6A1B9A; color: white; font-weight: bold;")
        help_btn.clicked.connect(self.show_user_manual)
        view_btn_layout.addWidget(help_btn)
        
        view_layout.addLayout(view_btn_layout)
        
        # 拖拽模式提示
        self.drag_hint_label = QLabel("")
        self.drag_hint_label.setStyleSheet("color: #FFEB3B; font-size: 12px;")
        self.drag_hint_label.setVisible(False)
        view_layout.addWidget(self.drag_hint_label)
        
        # 选中货物信息面板
        self.selected_cargo_group = QGroupBox("📦 选中货物信息 (点击3D视图中的货物查看)")
        self.selected_cargo_group.setStyleSheet("""
            QGroupBox {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3d5a80, stop:1 #2c3e50);
                border: 1px solid #4a90d9;
                border-radius: 6px;
                margin-top: 8px;
                font-weight: bold;
                color: #81D4FA;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        selected_cargo_layout = QHBoxLayout(self.selected_cargo_group)
        
        # 左侧：基本信息
        left_info = QVBoxLayout()
        self.cargo_name_label = QLabel("名称: -")
        self.cargo_size_label = QLabel("尺寸: -")
        self.cargo_weight_label = QLabel("重量: -")
        self.cargo_stackable_label = QLabel("可堆叠: -")
        
        for label in [self.cargo_name_label, self.cargo_size_label, 
                      self.cargo_weight_label, self.cargo_stackable_label]:
            label.setStyleSheet("color: #E0E0E0; font-size: 11px;")
            left_info.addWidget(label)
        selected_cargo_layout.addLayout(left_info)
        
        # 中间：位置信息
        mid_info = QVBoxLayout()
        self.cargo_pos_label = QLabel("位置: -")
        self.cargo_rotation_label = QLabel("旋转: -")
        self.cargo_layer_label = QLabel("层次: -")
        self.cargo_volume_label = QLabel("体积: -")
        
        for label in [self.cargo_pos_label, self.cargo_rotation_label,
                      self.cargo_layer_label, self.cargo_volume_label]:
            label.setStyleSheet("color: #E0E0E0; font-size: 11px;")
            mid_info.addWidget(label)
        selected_cargo_layout.addLayout(mid_info)
        
        # 右侧：加固建议和托盘详情按钮
        right_info = QVBoxLayout()
        self.cargo_securing_label = QLabel("加固建议: -")
        self.cargo_securing_label.setWordWrap(True)
        self.cargo_securing_label.setStyleSheet("color: #FFD54F; font-size: 11px;")
        right_info.addWidget(self.cargo_securing_label)
        
        # 查看托盘详情按钮（初始隐藏）
        self.view_pallet_btn = QPushButton("🔍 查看组托详情")
        self.view_pallet_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #66BB6A;
            }
        """)
        self.view_pallet_btn.setVisible(False)
        self.view_pallet_btn.clicked.connect(self.show_selected_pallet_details)
        right_info.addWidget(self.view_pallet_btn)
        
        right_info.addStretch()
        selected_cargo_layout.addLayout(right_info)
        
        self.selected_cargo_group.setMaximumHeight(120)
        view_layout.addWidget(self.selected_cargo_group)
        
        right_layout.addWidget(view_group)
        
        # 统计信息
        stats_group = QGroupBox("📊 配载统计")
        stats_layout = QVBoxLayout(stats_group)
        
        self.stats_label = QLabel("请先添加货物并开始配载")
        self.stats_label.setStyleSheet("font-size: 13px; color: #81D4FA;")
        self.stats_label.setWordWrap(True)
        stats_layout.addWidget(self.stats_label)
        
        # 空间利用率
        volume_layout = QHBoxLayout()
        volume_layout.addWidget(QLabel("空间利用率:"))
        self.volume_progress = QProgressBar()
        self.volume_progress.setRange(0, 100)
        self.volume_progress.setValue(0)
        self.volume_progress.setFormat("%p%")
        volume_layout.addWidget(self.volume_progress)
        self.volume_label = QLabel("0%")
        self.volume_label.setFixedWidth(50)
        volume_layout.addWidget(self.volume_label)
        stats_layout.addLayout(volume_layout)
        
        # 载重利用率
        weight_layout = QHBoxLayout()
        weight_layout.addWidget(QLabel("载重利用率:"))
        self.weight_progress = QProgressBar()
        self.weight_progress.setRange(0, 100)
        self.weight_progress.setValue(0)
        self.weight_progress.setFormat("%p%")
        self.weight_progress.setStyleSheet("""
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #FF9800, stop:1 #FFEB3B);
            }
        """)
        weight_layout.addWidget(self.weight_progress)
        self.weight_label = QLabel("0%")
        self.weight_label.setFixedWidth(50)
        weight_layout.addWidget(self.weight_label)
        stats_layout.addLayout(weight_layout)
        
        # 重心偏移
        cog_layout = QHBoxLayout()
        cog_layout.addWidget(QLabel("重心偏移:"))
        self.cog_label = QLabel("X: 0% | Y: 0%")
        self.cog_label.setStyleSheet("color: #4CAF50;")
        cog_layout.addWidget(self.cog_label)
        cog_layout.addStretch()
        stats_layout.addLayout(cog_layout)
        
        right_layout.addWidget(stats_group)
        
        # ==================== 装箱步骤 ====================
        steps_group = QGroupBox("📝 装箱步骤")
        steps_layout = QVBoxLayout(steps_group)
        
        self.steps_table = QTableWidget()
        self.steps_table.setColumnCount(6)
        self.steps_table.setHorizontalHeaderLabels(["序号", "集装箱", "货物名称", "尺寸(cm)", "位置坐标", "加固"])
        
        # 设置列宽比例
        header = self.steps_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # 序号
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)  # 集装箱
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # 货物名称
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # 尺寸
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)  # 位置坐标
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)  # 加固
        
        self.steps_table.setColumnWidth(0, 50)   # 序号
        self.steps_table.setColumnWidth(1, 70)   # 集装箱
        self.steps_table.setColumnWidth(3, 120)  # 尺寸
        self.steps_table.setColumnWidth(4, 130)  # 位置坐标
        self.steps_table.setColumnWidth(5, 80)   # 加固
        
        self.steps_table.setMaximumHeight(180)
        self.steps_table.setAlternatingRowColors(True)
        self.steps_table.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #2a3441;
                gridline-color: #3d4f5f;
            }
            QTableWidget::item {
                padding: 4px;
            }
            QHeaderView::section {
                background-color: #3d5a80;
                color: white;
                padding: 5px;
                border: none;
                font-weight: bold;
            }
        """)
        steps_layout.addWidget(self.steps_table)
        
        right_layout.addWidget(steps_group)
        
        # 添加到主布局
        main_layout.addWidget(left_panel)
        main_layout.addWidget(right_panel, 1)
    
    def setup_rules_table(self):
        """设置规则表格"""
        self.rules_list.setRowCount(len(self.loading_rules))
        for i, rule in enumerate(self.loading_rules):
            # 启用复选框
            cb = QCheckBox()
            cb.setChecked(rule.enabled)
            cb.stateChanged.connect(lambda state, r=rule: setattr(r, 'enabled', state == 2))
            self.rules_list.setCellWidget(i, 0, cb)
            
            # 规则名称
            name_item = QTableWidgetItem(rule.name)
            name_item.setToolTip(rule.description)
            self.rules_list.setItem(i, 1, name_item)
            
            # 优先级
            priority_item = QTableWidgetItem(str(rule.priority))
            self.rules_list.setItem(i, 2, priority_item)
    
    def on_category_changed(self, category):
        """容器类别变化"""
        self.container_combo.clear()
        if category == "海运集装箱":
            self.container_combo.addItems(CONTAINERS_SHIPPING.keys())
        elif category == "公路货车":
            self.container_combo.addItems(CONTAINERS_TRUCK.keys())
        elif category == "托盘/周转箱":
            self.container_combo.addItems(CONTAINERS_PALLET.keys())
        elif category == "自定义":
            self.container_combo.addItems(self.custom_containers.keys())
    
    def show_custom_container_dialog(self):
        """显示自定义容器对话框"""
        from PyQt6.QtWidgets import QDialog, QFormLayout, QDialogButtonBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("自定义容器")
        dialog.setMinimumWidth(350)
        
        layout = QFormLayout(dialog)
        
        name_edit = QLineEdit("自定义容器1")
        length_spin = QDoubleSpinBox()
        length_spin.setRange(1, 50000)
        length_spin.setValue(1200)
        length_spin.setSuffix(" cm")
        
        width_spin = QDoubleSpinBox()
        width_spin.setRange(1, 10000)
        width_spin.setValue(240)
        width_spin.setSuffix(" cm")
        
        height_spin = QDoubleSpinBox()
        height_spin.setRange(1, 10000)
        height_spin.setValue(260)
        height_spin.setSuffix(" cm")
        
        weight_spin = QDoubleSpinBox()
        weight_spin.setRange(1, 1000000)
        weight_spin.setValue(25000)
        weight_spin.setSuffix(" kg")
        
        type_combo = QComboBox()
        type_combo.addItems(["集装箱", "货车", "托盘"])
        
        layout.addRow("名称:", name_edit)
        layout.addRow("内部长度:", length_spin)
        layout.addRow("内部宽度:", width_spin)
        layout.addRow("内部高度:", height_spin)
        layout.addRow("最大载重:", weight_spin)
        layout.addRow("类型:", type_combo)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            type_map = {"集装箱": "container", "货车": "truck", "托盘": "pallet"}
            container = Container(
                name=name_edit.text(),
                length=length_spin.value(),
                width=width_spin.value(),
                height=height_spin.value(),
                max_weight=weight_spin.value(),
                container_type=type_map[type_combo.currentText()]
            )
            self.custom_containers[name_edit.text()] = container
            STANDARD_CONTAINERS[name_edit.text()] = container
            
            # 切换到自定义类别
            self.container_category.setCurrentText("自定义")
            self.on_category_changed("自定义")
            self.container_combo.setCurrentText(name_edit.text())
            
            QMessageBox.information(self, "成功", f"已添加自定义容器: {name_edit.text()}")
    
    def setup_default_container(self):
        """设置默认集装箱"""
        self.container_category.setCurrentText("海运集装箱")
        self.on_category_changed("海运集装箱")
        if self.container_combo.count() > 1:
            self.container_combo.setCurrentIndex(1)  # 40英尺标准箱
    
    def on_container_selected(self, name):
        """容器选择事件"""
        if not name:
            return
        self.container = STANDARD_CONTAINERS.get(name) or self.custom_containers.get(name)
        if self.container:
            type_names = {"container": "集装箱", "truck": "货车", "pallet": "托盘"}
            type_name = type_names.get(self.container.container_type, "容器")
            info = f"类型: {type_name}\n"
            info += f"内部尺寸: {self.container.length} × {self.container.width} × {self.container.height} cm\n"
            info += f"容积: {self.container.volume_cbm:.1f} m³ | 最大载重: {self.container.max_weight:,} kg"
            if self.container.description:
                info += f"\n{self.container.description}"
            self.container_info.setText(info)
            
            self.gl_widget.container = self.container
            self.gl_widget.placed_cargos = self.placed_cargos
            self.gl_widget.update()
    
    def get_next_color(self):
        """获取下一个颜色"""
        color = CARGO_COLORS[self.color_index % len(CARGO_COLORS)]
        self.color_index += 1
        return color
    
    def add_cargo(self):
        """添加货物"""
        cargo = Cargo(
            name=self.cargo_name.text() or f"货物{len(self.cargos)+1}",
            length=self.cargo_length.value(),
            width=self.cargo_width.value(),
            height=self.cargo_height.value(),
            weight=self.cargo_weight.value(),
            quantity=self.cargo_quantity.value(),
            stackable=self.cargo_stackable.isChecked(),
            allow_rotate=self.cargo_rotatable.isChecked(),
            bottom_only=self.cargo_bottom_only.isChecked(),
            priority=self.cargo_priority.value(),
            color=self.get_next_color()
        )
        
        self.cargos.append(cargo)
        self.update_cargo_table()
        self.cargo_name.setText(f"货物{len(self.cargos)+1}")
    
    def update_cargo_table(self):
        """更新货物表格"""
        # 暂时阻止信号，避免触发 cellChanged
        self.cargo_table.blockSignals(True)
        
        self.cargo_table.setRowCount(len(self.cargos))
        for i, cargo in enumerate(self.cargos):
            # 名称列 - 如果是托盘，添加标记
            name_text = cargo.name
            if cargo.is_pallet:
                name_text = f"📦 {cargo.name}"
            name_item = QTableWidgetItem(name_text)
            if cargo.is_pallet:
                name_item.setBackground(QColor(255, 243, 224))  # 浅橙色背景
            self.cargo_table.setItem(i, 0, name_item)
            
            # 尺寸显示为整数，更紧凑
            size_item = QTableWidgetItem(f"{int(cargo.length)}×{int(cargo.width)}×{int(cargo.height)}")
            if cargo.is_pallet:
                size_item.setBackground(QColor(255, 243, 224))
            self.cargo_table.setItem(i, 1, size_item)
            
            weight_item = QTableWidgetItem(f"{cargo.weight}kg")
            if cargo.is_pallet:
                weight_item.setBackground(QColor(255, 243, 224))
            self.cargo_table.setItem(i, 2, weight_item)
            
            qty_item = QTableWidgetItem(str(cargo.quantity))
            if cargo.is_pallet:
                qty_item.setBackground(QColor(255, 243, 224))
            self.cargo_table.setItem(i, 3, qty_item)
            
            # 选项列 - 显示图标表示各种属性
            options = []
            if cargo.is_pallet:
                options.append(f"[{len(cargo.pallet_contents)}件]")  # 托盘内货物数
            if cargo.allow_rotate:
                options.append("🔄")  # 可旋转
            if cargo.bottom_only:
                options.append("⬇")  # 仅底层
            if cargo.priority > 0:
                options.append(f"P{cargo.priority}")  # 优先级
            if cargo.group_id:
                options.append(f"{cargo.group_id}")  # 分组
            options_item = QTableWidgetItem("".join(options))
            if cargo.is_pallet:
                options_item.setBackground(QColor(255, 243, 224))
            self.cargo_table.setItem(i, 4, options_item)

            # 新增一列：组托托盘编号
            pallet_info = ""
            if hasattr(cargo, "pallet_no") and cargo.pallet_no:
                pallet_info = f"托盘{cargo.pallet_no}"
            elif hasattr(cargo, "pallet_of") and cargo.pallet_of:
                pallet_info = f"托盘{cargo.pallet_of}"
            pallet_item = QTableWidgetItem(pallet_info)
            self.cargo_table.setItem(i, 5, pallet_item)

            # 体积列（后移一列）
            volume_item = QTableWidgetItem(f"{cargo.total_volume/1000000:.2f}")
            self.cargo_table.setItem(i, 6, volume_item)
        
        # 恢复信号
        self.cargo_table.blockSignals(False)
    
    def on_cargo_table_cell_changed(self, row: int, column: int):
        """处理货物表格单元格编辑"""
        if row < 0 or row >= len(self.cargos):
            return
        
        cargo = self.cargos[row]
        item = self.cargo_table.item(row, column)
        if not item:
            return
        
        text = item.text().strip()
        
        try:
            if column == 0:  # 名称
                cargo.name = text
            elif column == 2:  # 重量
                # 移除 "kg" 后缀
                weight_str = text.replace("kg", "").replace("Kg", "").replace("KG", "").strip()
                cargo.weight = float(weight_str)
            elif column == 3:  # 数量
                new_qty = int(text)
                if new_qty > 0:
                    cargo.quantity = new_qty
                    # 更新体积列
                    self.cargo_table.blockSignals(True)
                    self.cargo_table.setItem(row, 5, QTableWidgetItem(
                        f"{cargo.total_volume/1000000:.2f}"))
                    self.cargo_table.blockSignals(False)
                else:
                    # 恢复原值
                    self.cargo_table.blockSignals(True)
                    self.cargo_table.setItem(row, 3, QTableWidgetItem(str(cargo.quantity)))
                    self.cargo_table.blockSignals(False)
        except ValueError:
            # 输入无效，恢复原值
            self.update_cargo_table()
    
    def delete_cargo(self):
        """删除选中货物"""
        row = self.cargo_table.currentRow()
        if row >= 0:
            del self.cargos[row]
            self.update_cargo_table()
    
    def clear_cargos(self):
        """清空货物"""
        if self.cargos:
            reply = QMessageBox.question(self, "确认", "确定要清空货物列表吗？")
            if reply == QMessageBox.StandardButton.Yes:
                self.cargos.clear()
                self.color_index = 0
                self.update_cargo_table()
    
    def import_cargos(self):
        """导入货物"""
        file_filter = "Excel文件 (*.xlsx);;JSON文件 (*.json)" if EXCEL_SUPPORT else "JSON文件 (*.json)"
        filename, selected_filter = QFileDialog.getOpenFileName(
            self, "导入货物", "", file_filter)
        if filename:
            try:
                if filename.endswith('.xlsx'):
                    self.import_from_excel(filename)
                else:
                    with open(filename, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    
                    self.cargos = []
                    self.cargo_groups = []
                    group_map = {}
                    
                    # 处理货物数据
                    cargo_list = data.get('cargos', data) if isinstance(data, dict) else data
                    for item in cargo_list:
                        if 'color' in item and isinstance(item['color'], list):
                            item['color'] = tuple(item['color'])
                        else:
                            item['color'] = self.get_next_color()
                        cargo = Cargo(**item)
                        self.cargos.append(cargo)
                        
                        # 记录分组
                        if cargo.group_id:
                            if cargo.group_id not in group_map:
                                group_map[cargo.group_id] = []
                            group_map[cargo.group_id].append(cargo.id)
                    
                    # 创建分组对象
                    for gid, cargo_ids in group_map.items():
                        group = CargoGroup(id=gid, name=f"分组{gid}", cargo_ids=cargo_ids)
                        self.cargo_groups.append(group)
                    
                    self.update_cargo_table()
                    group_info = f"，{len(self.cargo_groups)}个分组" if self.cargo_groups else ""
                    QMessageBox.information(self, "成功", f"成功导入 {len(self.cargos)} 种货物{group_info}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导入失败: {e}")
    
    def import_from_excel(self, filename):
        """从Excel导入货物"""
        wb = load_workbook(filename)
        ws = wb.active
        
        self.cargos = []
        self.cargo_groups = []
        self.color_index = 0
        group_map = {}  # 记录分组ID到货物ID的映射
        
        # 跳过标题行，从第2行开始读取
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 空行跳过
                continue
            
            name = str(row[0]) if row[0] else f"货物{len(self.cargos)+1}"
            length = float(row[1]) if row[1] else 100
            width = float(row[2]) if row[2] else 80
            height = float(row[3]) if row[3] else 60
            weight = float(row[4]) if row[4] else 50
            quantity = int(row[5]) if row[5] else 1
            stackable = True
            if len(row) > 6 and row[6] is not None:
                stackable = str(row[6]).lower() in ('true', '是', '1', 'yes')
            
            # 读取分组信息 (第11列，索引10)
            group_id = None
            if len(row) > 10 and row[10]:
                group_id = str(row[10]).strip()
            
            cargo = Cargo(
                name=name,
                length=length,
                width=width,
                height=height,
                weight=weight,
                quantity=quantity,
                stackable=stackable,
                group_id=group_id,
                color=self.get_next_color()
            )
            self.cargos.append(cargo)
            
            # 记录分组
            if group_id:
                if group_id not in group_map:
                    group_map[group_id] = []
                group_map[group_id].append(cargo.id)
        
        # 创建分组对象
        for gid, cargo_ids in group_map.items():
            group = CargoGroup(
                id=gid,
                name=f"分组{gid}",
                cargo_ids=cargo_ids
            )
            self.cargo_groups.append(group)
        
        self.update_cargo_table()
        group_info = f"，{len(self.cargo_groups)}个分组" if self.cargo_groups else ""
        QMessageBox.information(self, "成功", f"成功从Excel导入 {len(self.cargos)} 种货物{group_info}")
    
    def export_cargos(self):
        """导出货物"""
        if not self.cargos:
            QMessageBox.warning(self, "警告", "没有货物可导出")
            return
        
        file_filter = "Excel文件 (*.xlsx);;JSON文件 (*.json)" if EXCEL_SUPPORT else "JSON文件 (*.json)"
        filename, selected_filter = QFileDialog.getSaveFileName(
            self, "导出货物", "", file_filter)
        if filename:
            try:
                if filename.endswith('.xlsx'):
                    self.export_to_excel(filename)
                else:
                    data = []
                    for cargo in self.cargos:
                        d = asdict(cargo)
                        d['color'] = list(d['color'])
                        data.append(d)
                    with open(filename, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    QMessageBox.information(self, "成功", "货物导出成功")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {e}")
    
    def export_to_excel(self, filename):
        """导出货物到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "货物清单"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        headers = ["货物名称", "长度(cm)", "宽度(cm)", "高度(cm)", "重量(kg)", "数量", "可堆叠", "单件体积(m³)", "总体积(m³)", "总重量(kg)", "分组"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row, cargo in enumerate(self.cargos, 2):
            ws.cell(row=row, column=1, value=cargo.name).border = thin_border
            ws.cell(row=row, column=2, value=cargo.length).border = thin_border
            ws.cell(row=row, column=3, value=cargo.width).border = thin_border
            ws.cell(row=row, column=4, value=cargo.height).border = thin_border
            ws.cell(row=row, column=5, value=cargo.weight).border = thin_border
            ws.cell(row=row, column=6, value=cargo.quantity).border = thin_border
            ws.cell(row=row, column=7, value="是" if cargo.stackable else "否").border = thin_border
            ws.cell(row=row, column=8, value=round(cargo.volume / 1000000, 4)).border = thin_border
            ws.cell(row=row, column=9, value=round(cargo.total_volume / 1000000, 4)).border = thin_border
            ws.cell(row=row, column=10, value=cargo.total_weight).border = thin_border
            ws.cell(row=row, column=11, value=cargo.group_id or "").border = thin_border
        
        # 调整列宽
        column_widths = [15, 12, 12, 12, 12, 10, 10, 14, 14, 14, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(filename)
        QMessageBox.information(self, "成功", "货物已导出到Excel文件")
    
    def start_loading(self):
        """开始配载"""
        if not self.container:
            QMessageBox.warning(self, "警告", "请先选择集装箱")
            return
        
        if not self.cargos:
            QMessageBox.warning(self, "警告", "请先添加货物")
            return
        
        # 创建进度对话框
        self.loading_progress = QProgressDialog("正在配载中...", "取消", 0, 100, self)
        self.loading_progress.setWindowTitle("配载进度")
        self.loading_progress.setWindowModality(Qt.WindowModality.WindowModal)
        self.loading_progress.setMinimumDuration(0)
        self.loading_progress.setAutoClose(True)
        self.loading_progress.setAutoReset(True)
        self.loading_progress.setStyleSheet("""
            QProgressDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QProgressBar {
                border: 1px solid #555;
                border-radius: 5px;
                background-color: #1e1e1e;
                text-align: center;
                color: white;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2196F3, stop:1 #4CAF50);
                border-radius: 4px;
            }
            QPushButton {
                background-color: #c62828;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
        """)
        self.loading_progress.setValue(5)
        self.loading_progress.setLabelText("正在准备配载规则...")
        QApplication.processEvents()
        
        # 收集启用的规则
        active_rules = []
        for row in range(self.rules_list.rowCount()):
            checkbox = self.rules_list.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                rule_name = self.rules_list.item(row, 1).text()
                priority = int(self.rules_list.item(row, 2).text())
                
                if rule_name == "相同尺寸优先":
                    active_rules.append((priority, RuleSameSizeFirst()))
                elif rule_name == "重货在下":
                    active_rules.append((priority, RuleHeavyBottom()))
                elif rule_name == "相似尺寸堆叠":
                    active_rules.append((priority, RuleSimilarSizeStack()))
                elif rule_name == "体积大优先":
                    active_rules.append((priority, RuleVolumeFirst()))
                elif rule_name == "优先级排序":
                    active_rules.append((priority, RulePriorityFirst()))
        
        self.loading_progress.setValue(10)
        self.loading_progress.setLabelText("正在排序配载规则...")
        QApplication.processEvents()
        
        # 按优先级排序规则
        active_rules.sort(key=lambda x: x[0], reverse=True)
        rules = [r[1] for r in active_rules]
        
        self.loading_progress.setValue(15)
        QApplication.processEvents()
        
        # 多集装箱模式
        if self.multi_container_mode:
            self.start_multi_container_loading(rules)
        else:
            self.start_single_container_loading(rules)
    
    def start_single_container_loading(self, rules: list):
        """单集装箱配载"""
        # 更新进度条
        self.loading_progress.setValue(20)
        self.loading_progress.setLabelText("正在初始化配载算法...")
        QApplication.processEvents()
        
        # 执行配载
        algorithm = LoadingAlgorithm(self.container, rules=rules, cargo_groups=self.cargo_groups)
        
        self.loading_progress.setValue(30)
        self.loading_progress.setLabelText("正在计算货物放置位置...")
        QApplication.processEvents()
        
        loaded, not_loaded = algorithm.load_all(self.cargos)
        
        self.loading_progress.setValue(70)
        self.loading_progress.setLabelText("正在生成3D视图...")
        QApplication.processEvents()

        self.placed_cargos = loaded
        self.container_results = []  # 清空多集装箱结果
        self.gl_widget.placed_cargos = loaded
        self.gl_widget.update()
        
        # 隐藏集装箱选择器
        self.container_selector_group.setVisible(False)
        
        # 更新统计
        stats = algorithm.get_statistics()
        
        stats_text = f"已装载: {stats['loaded_count']} 件 | "
        stats_text += f"未装载: {len(not_loaded)} 件 | "
        stats_text += f"总体积: {stats['total_volume']/1000000:.2f} m³ | "
        stats_text += f"总重量: {stats['total_weight']:.1f} kg"
        
        self.stats_label.setText(stats_text)
        self.volume_progress.setValue(int(stats['volume_utilization']))
        self.volume_label.setText(f"{stats['volume_utilization']:.1f}%")
        self.weight_progress.setValue(int(stats['weight_utilization']))
        self.weight_label.setText(f"{stats['weight_utilization']:.1f}%")
        
        # 更新重心显示
        cog_tuple = stats.get('center_of_gravity', (0, 0, 0))
        offset_tuple = stats.get('center_offset', (0, 0, 0))
        
        # 判断重心状态
        max_offset = min(self.container.length, self.container.width) * 0.1
        if abs(offset_tuple[0]) < max_offset and abs(offset_tuple[1]) < max_offset:
            cog_status = "良好"
        else:
            cog_status = "偏移较大"
        
        cog_text = f"重心位置: X={cog_tuple[0]:.1f}, Y={cog_tuple[1]:.1f}, Z={cog_tuple[2]:.1f} cm\n"
        cog_text += f"偏移: 横向 {offset_tuple[0]:.1f}cm, 纵向 {offset_tuple[1]:.1f}cm | 状态: {cog_status}"
        self.cog_label.setText(cog_text)
        
        self.loading_progress.setValue(85)
        self.loading_progress.setLabelText("正在生成装载步骤...")
        QApplication.processEvents()
        
        # 更新装载步骤表格
        self.update_steps_table(algorithm.get_loading_steps())
        
        self.loading_progress.setValue(100)
        self.loading_progress.setLabelText("配载完成！")
        QApplication.processEvents()
        
        if not_loaded:
            cargo_names = ", ".join(set(c.name for c in not_loaded))
            QMessageBox.information(self, "配载完成",
                f"配载完成！\n\n"
                f"空间利用率: {stats['volume_utilization']:.1f}%\n"
                f"载重利用率: {stats['weight_utilization']:.1f}%\n"
                f"重心状态: {cog_status}\n\n"
                f"有 {len(not_loaded)} 件货物无法装入:\n{cargo_names}")
        else:
            QMessageBox.information(self, "配载完成",
                f"所有货物已成功装载！\n\n"
                f"空间利用率: {stats['volume_utilization']:.1f}%\n"
                f"载重利用率: {stats['weight_utilization']:.1f}%\n"
                f"重心状态: {cog_status}")
    
    def start_multi_container_loading(self, rules: list):
        """多集装箱配载"""
        container_count = self.container_count_spin.value()
        
        self.container_results = []
        remaining_cargos = []
        
        # 展开所有货物
        for cargo in self.cargos:
            for i in range(cargo.quantity):
                single_cargo = copy.copy(cargo)
                single_cargo.quantity = 1
                single_cargo.id = f"{cargo.id}_{i}"
                remaining_cargos.append(single_cargo)
        
        self.loading_progress.setValue(20)
        self.loading_progress.setLabelText("正在展开货物列表...")
        QApplication.processEvents()
        
        # 依次填充每个集装箱
        for container_idx in range(container_count):
            if not remaining_cargos:
                break
            
            # 更新进度
            progress = 20 + int(60 * (container_idx + 1) / container_count)
            self.loading_progress.setValue(progress)
            self.loading_progress.setLabelText(f"正在配载集装箱 {container_idx + 1}/{container_count}...")
            QApplication.processEvents()
            
            if self.loading_progress.wasCanceled():
                QMessageBox.information(self, "提示", "配载已取消")
                return
            
            # 创建算法实例
            algorithm = LoadingAlgorithm(self.container, rules=rules)
            
            # 尝试装载剩余货物
            loaded_in_this = []
            still_remaining = []
            
            for cargo in remaining_cargos:
                if algorithm.place_cargo(cargo):
                    placed = algorithm.placed_cargos[-1]
                    placed.container_index = container_idx
                    loaded_in_this.append(placed)
                else:
                    still_remaining.append(cargo)
            
            # 创建结果对象
            result = ContainerLoadingResult(
                container=copy.copy(self.container),
                container_index=container_idx,
                placed_cargos=loaded_in_this
            )
            self.container_results.append(result)
            
            remaining_cargos = still_remaining
        
        self.loading_progress.setValue(85)
        self.loading_progress.setLabelText("正在生成3D视图...")
        QApplication.processEvents()
        
        # 更新集装箱选择器
        self.container_selector.blockSignals(True)
        self.container_selector.clear()
        self.container_selector.addItem("全部概览")
        for i, result in enumerate(self.container_results):
            util = result.volume_utilization
            self.container_selector.addItem(f"集装箱 #{i+1} ({util:.1f}%)")
        self.container_selector.blockSignals(False)
        
        # 显示集装箱选择器
        self.container_selector_group.setVisible(True)
        self.container_selector.setCurrentIndex(0)  # 默认选择"全部概览"
        
        # 设置3D视图为多集装箱模式
        self.gl_widget.set_multi_container_results(self.container_results)
        
        self.loading_progress.setValue(100)
        self.loading_progress.setLabelText("配载完成！")
        QApplication.processEvents()
        
        # 合并所有装载的货物
        self.placed_cargos = []
        for result in self.container_results:
            self.placed_cargos.extend(result.placed_cargos)
        
        # 更新统计
        self.update_stats_for_container(-1)
        
        # 更新装载步骤表格（显示所有集装箱）
        all_steps = []
        step_num = 0
        for result in self.container_results:
            for placed in result.placed_cargos:
                step_num += 1
                all_steps.append({
                    'step': step_num,
                    'container': f"#{result.container_index+1}",
                    'cargo_name': placed.cargo.name,
                    'dimensions': f"{placed.actual_length}×{placed.actual_width}×{placed.cargo.height}",
                    'position': f"({placed.x:.0f}, {placed.y:.0f}, {placed.z:.0f})",
                    'securing': '标准'
                })
        self.update_steps_table(all_steps)
        
        # 显示结果
        total_loaded = len(self.placed_cargos)
        total_remaining = len(remaining_cargos)
        used_containers = len([r for r in self.container_results if r.placed_cargos])
        
        msg = f"多集装箱配载完成！\n\n"
        msg += f"使用集装箱: {used_containers} 个\n"
        msg += f"总装载: {total_loaded} 件\n"
        
        if remaining_cargos:
            cargo_names = ", ".join(set(c.name for c in remaining_cargos))
            msg += f"未装载: {total_remaining} 件\n"
            msg += f"未装载货物: {cargo_names}"
        else:
            msg += f"所有货物已成功装载！"
        
        QMessageBox.information(self, "多集装箱配载完成", msg)
    
    def update_steps_table(self, steps: list):
        """更新装载步骤表格"""
        self.steps_table.setRowCount(len(steps))
        for i, step in enumerate(steps):
            # 序号
            item0 = QTableWidgetItem(str(step.get('step', i+1)))
            item0.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.steps_table.setItem(i, 0, item0)
            
            # 集装箱
            item1 = QTableWidgetItem(step.get('container', '-'))
            item1.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.steps_table.setItem(i, 1, item1)
            
            # 货物名称
            self.steps_table.setItem(i, 2, QTableWidgetItem(step.get('cargo_name', '')))
            
            # 尺寸
            item3 = QTableWidgetItem(step.get('dimensions', ''))
            item3.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.steps_table.setItem(i, 3, item3)
            
            # 位置坐标
            item4 = QTableWidgetItem(step.get('position', ''))
            item4.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.steps_table.setItem(i, 4, item4)
            
            # 加固建议
            item5 = QTableWidgetItem(step.get('securing', '标准'))
            item5.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.steps_table.setItem(i, 5, item5)
    
    def create_cargo_group(self):
        """创建货物分组"""
        selected_rows = set()
        for item in self.cargo_table.selectedItems():
            selected_rows.add(item.row())
        
        if len(selected_rows) < 2:
            QMessageBox.warning(self, "警告", "请至少选择2个货物来创建分组")
            return
        
        # 生成新的分组ID
        group_id = f"G{len(self.cargo_groups) + 1}"
        
        # 获取选中的货物ID列表
        cargo_ids = []
        for row in selected_rows:
            cargo = self.cargos[row]
            cargo.group_id = group_id
            cargo_ids.append(cargo.id)
        
        # 创建分组对象
        group = CargoGroup(
            id=group_id,
            name=f"分组{len(self.cargo_groups) + 1}",
            cargo_ids=cargo_ids
        )
        self.cargo_groups.append(group)
        
        self.update_cargo_table()
        QMessageBox.information(self, "成功", f"已创建分组 {group_id}，包含 {len(cargo_ids)} 个货物")
    
    def ungroup_cargo(self):
        """取消货物分组"""
        selected_rows = set()
        for item in self.cargo_table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请选择要取消分组的货物")
            return
        
        ungrouped_count = 0
        for row in selected_rows:
            cargo = self.cargos[row]
            if cargo.group_id:
                # 从分组中移除
                for group in self.cargo_groups:
                    if cargo.id in group.cargo_ids:
                        group.cargo_ids.remove(cargo.id)
                        if not group.cargo_ids:  # 如果分组为空，删除分组
                            self.cargo_groups.remove(group)
                        break
                cargo.group_id = None
                ungrouped_count += 1
        
        self.update_cargo_table()
        if ungrouped_count > 0:
            QMessageBox.information(self, "成功", f"已取消 {ungrouped_count} 个货物的分组")
        else:
            QMessageBox.information(self, "提示", "选中的货物没有分组")
    
    def enable_manual_edit(self):
        """启用手动编辑模式"""
        if not self.placed_cargos:
            QMessageBox.warning(self, "警告", "没有配载结果可编辑，请先执行配载")
            return
        
        # 创建手动编辑对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("手动编辑配载")
        dialog.setMinimumSize(800, 600)
        layout = QVBoxLayout(dialog)
        
        # 说明标签
        hint_label = QLabel("选择货物并调整其位置，可拖动滑块或直接输入坐标值")
        hint_label.setStyleSheet("color: #888; font-size: 12px;")
        layout.addWidget(hint_label)
        
        # 货物选择
        cargo_combo = QComboBox()
        for i, pc in enumerate(self.placed_cargos):
            cargo_combo.addItem(f"{i+1}. {pc.cargo.name} @ ({pc.x:.0f}, {pc.y:.0f}, {pc.z:.0f})")
        layout.addWidget(cargo_combo)
        
        # 位置编辑
        pos_group = QGroupBox("位置调整")
        pos_layout = QGridLayout(pos_group)
        
        x_label = QLabel("X (长度方向):")
        x_spin = QSpinBox()
        x_spin.setRange(0, int(self.container.length))
        x_spin.setSingleStep(10)
        
        y_label = QLabel("Y (宽度方向):")
        y_spin = QSpinBox()
        y_spin.setRange(0, int(self.container.width))
        y_spin.setSingleStep(10)
        
        z_label = QLabel("Z (高度方向):")
        z_spin = QSpinBox()
        z_spin.setRange(0, int(self.container.height))
        z_spin.setSingleStep(10)
        
        rotate_check = QCheckBox("旋转90度")
        
        pos_layout.addWidget(x_label, 0, 0)
        pos_layout.addWidget(x_spin, 0, 1)
        pos_layout.addWidget(y_label, 1, 0)
        pos_layout.addWidget(y_spin, 1, 1)
        pos_layout.addWidget(z_label, 2, 0)
        pos_layout.addWidget(z_spin, 2, 1)
        pos_layout.addWidget(rotate_check, 3, 0, 1, 2)
        layout.addWidget(pos_group)
        
        def on_cargo_selected(index):
            if index >= 0 and index < len(self.placed_cargos):
                pc = self.placed_cargos[index]
                x_spin.setValue(int(pc.x))
                y_spin.setValue(int(pc.y))
                z_spin.setValue(int(pc.z))
                rotate_check.setChecked(pc.rotated)
        
        def apply_position():
            index = cargo_combo.currentIndex()
            if index >= 0 and index < len(self.placed_cargos):
                pc = self.placed_cargos[index]
                pc.x = x_spin.value()
                pc.y = y_spin.value()
                pc.z = z_spin.value()
                pc.rotated = rotate_check.isChecked()
                self.gl_widget.update()
                cargo_combo.setItemText(index, 
                    f"{index+1}. {pc.cargo.name} @ ({pc.x:.0f}, {pc.y:.0f}, {pc.z:.0f})")
        
        cargo_combo.currentIndexChanged.connect(on_cargo_selected)
        on_cargo_selected(0)  # 初始化第一个
        
        # 应用按钮
        apply_btn = QPushButton("应用更改")
        apply_btn.clicked.connect(apply_position)
        apply_btn.setStyleSheet("background-color: #4CAF50; font-weight: bold;")
        layout.addWidget(apply_btn)
        
        # 删除货物按钮
        def remove_cargo():
            index = cargo_combo.currentIndex()
            if index >= 0 and index < len(self.placed_cargos):
                del self.placed_cargos[index]
                cargo_combo.removeItem(index)
                self.gl_widget.update()
                # 更新组合框中的编号
                for i in range(cargo_combo.count()):
                    pc = self.placed_cargos[i]
                    cargo_combo.setItemText(i, 
                        f"{i+1}. {pc.cargo.name} @ ({pc.x:.0f}, {pc.y:.0f}, {pc.z:.0f})")
        
        remove_btn = QPushButton("删除此货物")
        remove_btn.clicked.connect(remove_cargo)
        remove_btn.setStyleSheet("background-color: #f44336;")
        layout.addWidget(remove_btn)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec()
        
        # 更新统计
        if self.placed_cargos:
            total_volume = sum(p.cargo.volume for p in self.placed_cargos)
            total_weight = sum(p.cargo.weight for p in self.placed_cargos)
            vol_util = (total_volume / self.container.volume) * 100
            wt_util = (total_weight / self.container.max_weight) * 100
            
            self.volume_progress.setValue(int(vol_util))
            self.volume_label.setText(f"{vol_util:.1f}%")
            self.weight_progress.setValue(int(wt_util))
            self.weight_label.setText(f"{wt_util:.1f}%")
    
    def palletize_cargos(self):
        """小件组托 - 将小货物组合成托盘，使用3D装箱算法"""
        if not self.cargos:
            QMessageBox.warning(self, "警告", "请先添加货物")
            return
        
        # 创建组托对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("小件组托")
        dialog.setMinimumWidth(500)
        layout = QVBoxLayout(dialog)
        
        # 托盘尺寸选择
        pallet_group = QGroupBox("托盘规格")
        pallet_layout = QFormLayout(pallet_group)
        
        pallet_type = QComboBox()
        pallet_type.addItems(["标准托盘 (120×100×15)", "欧标托盘 (120×80×15)", "自定义"])
        pallet_layout.addRow("托盘类型:", pallet_type)
        
        pallet_length = QSpinBox()
        pallet_length.setRange(50, 200)
        pallet_length.setValue(120)
        pallet_layout.addRow("长度(cm):", pallet_length)
        
        pallet_width = QSpinBox()
        pallet_width.setRange(50, 200)
        pallet_width.setValue(100)
        pallet_layout.addRow("宽度(cm):", pallet_width)
        
        pallet_base_height = QSpinBox()
        pallet_base_height.setRange(10, 30)
        pallet_base_height.setValue(15)
        pallet_layout.addRow("托盘底座高度(cm):", pallet_base_height)
        
        max_height = QSpinBox()
        max_height.setRange(50, 300)
        max_height.setValue(150)
        max_height.setToolTip("包含托盘底座的总高度限制")
        pallet_layout.addRow("最大堆叠高度(cm):", max_height)
        
        max_weight = QSpinBox()
        max_weight.setRange(100, 2000)
        max_weight.setValue(1000)
        pallet_layout.addRow("最大载重(kg):", max_weight)
        
        def on_pallet_type_changed(index):
            if index == 0:  # 标准托盘
                pallet_length.setValue(120)
                pallet_width.setValue(100)
            elif index == 1:  # 欧标托盘
                pallet_length.setValue(120)
                pallet_width.setValue(80)
        
        pallet_type.currentIndexChanged.connect(on_pallet_type_changed)
        layout.addWidget(pallet_group)
        
        # 选择要组托的货物
        cargo_group = QGroupBox("选择货物")
        cargo_layout = QVBoxLayout(cargo_group)
        
        cargo_list = QListWidget()
        cargo_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        for cargo in self.cargos:
            # 标记已经是托盘的货物
            prefix = "📦 " if cargo.is_pallet else ""
            cargo_list.addItem(f"{prefix}{cargo.name} - {cargo.length}×{cargo.width}×{cargo.height}cm, {cargo.weight}kg × {cargo.quantity}")
        cargo_layout.addWidget(cargo_list)
        
        btn_row = QHBoxLayout()
        select_all_btn = QPushButton("全选小件(体积<0.1m³)")
        def select_small():
            for i, cargo in enumerate(self.cargos):
                if cargo.volume < 100000 and not cargo.is_pallet:  # 0.1m³ = 100000 cm³
                    cargo_list.item(i).setSelected(True)
        select_all_btn.clicked.connect(select_small)
        btn_row.addWidget(select_all_btn)
        
        select_none_btn = QPushButton("取消全选")
        select_none_btn.clicked.connect(lambda: cargo_list.clearSelection())
        btn_row.addWidget(select_none_btn)
        cargo_layout.addLayout(btn_row)
        layout.addWidget(cargo_group)
        
        # 按钮
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("开始组托")
        ok_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; padding: 8px 20px;")
        cancel_btn = QPushButton("取消")
        
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            selected_indices = [cargo_list.row(item) for item in cargo_list.selectedItems()]
            if not selected_indices:
                QMessageBox.warning(self, "警告", "请选择要组托的货物")
                return
            
            # 获取托盘参数
            pallet_l = pallet_length.value()
            pallet_w = pallet_width.value()
            base_h = pallet_base_height.value()
            max_h = max_height.value()
            max_wt = max_weight.value()
            content_max_h = max_h - base_h  # 货物可用高度
            
            # 收集选中的货物（展开数量）
            selected_cargos = []
            for i, cargo in enumerate(self.cargos):
                if i in selected_indices and not cargo.is_pallet:
                    # 展开数量
                    for _ in range(cargo.quantity):
                        single_cargo = copy.copy(cargo)
                        single_cargo.quantity = 1
                        selected_cargos.append(single_cargo)

            if not selected_cargos:
                QMessageBox.warning(self, "警告", "没有可组托的货物")
                return
            
            # 组托算法
            palletized_cargos = self._palletize_with_3d_algorithm(
                selected_cargos, pallet_l, pallet_w, base_h, content_max_h, max_wt
            )

            # 标记原货物所属托盘
            for pallet in palletized_cargos:
                for content in pallet.pallet_contents:
                    # 在原货物列表中查找对应货物并标记
                    for cargo in self.cargos:
                        if (not cargo.is_pallet and cargo.id == content.cargo.id):
                            cargo.pallet_of = pallet.name.replace("托盘","")  # 托盘编号

            # 显示组托结果，保留原货物信息
            self._show_palletize_result(palletized_cargos, self.cargos)
    
    def _palletize_with_3d_algorithm(self, cargos: List[Cargo], 
                                      pallet_l: float, pallet_w: float, 
                                      base_h: float, content_max_h: float, 
                                      max_wt: float) -> List[Cargo]:
        """使用3D装箱算法进行组托"""
        palletized = []
        remaining = cargos.copy()
        pallet_count = 0
        
        # 按体积从大到小排序
        remaining.sort(key=lambda c: c.volume, reverse=True)
        
        while remaining:
            pallet_count += 1
            pallet_contents = []
            current_weight = 0
            placed_items = []  # (x, y, z, length, width, height)
            
            # 多次遍历剩余货物，直到没有货物能放入当前托盘
            made_progress = True
            while made_progress:
                made_progress = False
                still_remaining = []
                
                for cargo in remaining:
                    placed = False
                    
                    # 尝试不同的放置方式（先不旋转，再旋转）
                    for rotated in [False, True]:
                        if not cargo.allow_rotate and rotated:
                            continue
                        
                        c_l = cargo.width if rotated else cargo.length
                        c_w = cargo.length if rotated else cargo.width
                        c_h = cargo.height
                        
                        # 检查尺寸是否适合托盘
                        if c_l > pallet_l or c_w > pallet_w or c_h > content_max_h:
                            continue
                        
                        # 检查重量
                        if current_weight + cargo.weight > max_wt:
                            continue
                        
                        # 寻找可放置位置
                        position = self._find_position_on_pallet(
                            placed_items, c_l, c_w, c_h, 
                            pallet_l, pallet_w, content_max_h
                        )
                        
                        if position:
                            x, y, z = position
                            placed_items.append((x, y, z, c_l, c_w, c_h))
                            current_weight += cargo.weight
                            
                            # 记录放置内容
                            content = PalletContent(
                                cargo=cargo,
                                x=x, y=y, z=z,
                                rotated=rotated,
                                quantity=1
                            )
                            pallet_contents.append(content)
                            placed = True
                            made_progress = True  # 有进展，继续循环
                            break
                    
                    if not placed:
                        still_remaining.append(cargo)
                
                remaining = still_remaining
            
            # 创建托盘货物
            if pallet_contents:
                # 计算实际使用的高度
                actual_height = base_h
                for item in placed_items:
                    item_top = item[2] + item[5]  # z + height
                    actual_height = max(actual_height, base_h + item_top)
                
                pallet_cargo = Cargo(
                    name=f"托盘{pallet_count}",
                    length=pallet_l,
                    width=pallet_w,
                    height=actual_height,
                    weight=current_weight,
                    quantity=1,
                    stackable=True,
                    color=self.get_next_color(),
                    is_pallet=True,
                    pallet_base_height=base_h,
                    pallet_contents=pallet_contents,
                    original_cargos=[c.cargo for c in pallet_contents]
                )
                palletized.append(pallet_cargo)
            
            # 防止无限循环
            if not pallet_contents and remaining:
                # 有货物放不进任何托盘
                QMessageBox.warning(self, "警告", 
                    f"有 {len(remaining)} 件货物尺寸超过托盘限制，无法组托")
                break
        
        return palletized
    
    def _find_position_on_pallet(self, placed_items, c_l, c_w, c_h, 
                                  pallet_l, pallet_w, max_h) -> Optional[Tuple[float, float, float]]:
        """在托盘上寻找可放置位置 - 使用底部左下角优先策略"""
        
        # 策略：优先填满底层，从左下角开始，逐行逐列扫描
        # 生成所有可能的放置位置
        
        # 1. 收集所有"极限点"（Extreme Points）
        extreme_points = set()
        extreme_points.add((0, 0, 0))  # 起始点
        
        # 从已放置物品生成极限点
        for item in placed_items:
            x, y, z, l, w, h = item
            # 物品右侧
            extreme_points.add((x + l, y, z))
            # 物品前侧  
            extreme_points.add((x, y + w, z))
            # 物品顶部
            extreme_points.add((x, y, z + h))
            # 组合点
            extreme_points.add((x + l, y, 0))
            extreme_points.add((x, y + w, 0))
            extreme_points.add((x + l, y + w, 0))
            extreme_points.add((x + l, y + w, z))
        
        # 2. 按固定间隔扫描整个托盘底面，确保不遗漏空间
        # 使用更小的步长确保覆盖所有可能位置
        scan_step_x = min(c_l / 2, 5)  # 步长为货物长度一半或5cm
        scan_step_y = min(c_w / 2, 5)  # 步长为货物宽度一半或5cm
        
        x = 0
        while x <= pallet_l - c_l + 0.01:
            y = 0
            while y <= pallet_w - c_w + 0.01:
                extreme_points.add((x, y, 0))
                y += scan_step_y
            x += scan_step_x
        
        # 3. 对于每个已放物品的顶部，也扫描可堆叠位置
        for item in placed_items:
            ix, iy, iz, il, iw, ih = item
            top_z = iz + ih
            if top_z + c_h <= max_h + 0.01:
                # 在该物品顶部扫描
                sx = ix
                while sx <= ix + il - c_l + 0.01:
                    sy = iy
                    while sy <= iy + iw - c_w + 0.01:
                        extreme_points.add((sx, sy, top_z))
                        sy += scan_step_y
                    sx += scan_step_x
        
        # 4. 按优先级排序：先底层，再按y坐标（前后），再按x坐标（左右）
        candidates = list(extreme_points)
        candidates.sort(key=lambda p: (p[2], p[1], p[0]))
        
        for cx, cy, cz in candidates:
            # 检查边界
            if cx < -0.01 or cy < -0.01 or cz < -0.01:
                continue
            if cx + c_l > pallet_l + 0.01 or cy + c_w > pallet_w + 0.01 or cz + c_h > max_h + 0.01:
                continue
            
            # 检查与已放置物品的碰撞
            collision = False
            for item in placed_items:
                x, y, z, l, w, h = item
                # 使用严格的碰撞检测
                if (cx < x + l and cx + c_l > x and
                    cy < y + w and cy + c_w > y and
                    cz < z + h and cz + c_h > z):
                    collision = True
                    break
            
            if collision:
                continue
            
            # 检查底部支撑（如果不在底层）
            if cz > 0.01:
                support = self._check_support(cx, cy, cz, c_l, c_w, placed_items)
                if not support:
                    continue
            
            return (cx, cy, cz)
        
        return None
    
    def _check_support(self, x, y, z, l, w, placed_items) -> bool:
        """检查底部是否有足够支撑"""
        support_area = 0
        required_area = l * w * 0.7  # 需要70%的支撑面积
        
        for item in placed_items:
            ix, iy, iz, il, iw, ih = item
            # 检查是否在正下方
            if abs(iz + ih - z) < 0.1:  # 顶部接触
                # 计算重叠面积
                overlap_x = max(0, min(x + l, ix + il) - max(x, ix))
                overlap_y = max(0, min(y + w, iy + iw) - max(y, iy))
                support_area += overlap_x * overlap_y
        
        return support_area >= required_area
    
    def _show_palletize_result(self, palletized_cargos: List[Cargo], remaining_cargos: List[Cargo]):
        """显示组托结果对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("组托结果")
        dialog.setMinimumSize(800, 600)
        layout = QVBoxLayout(dialog)
        
        # 统计信息
        info_label = QLabel(f"✅ 已生成 {len(palletized_cargos)} 个托盘")
        info_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2196F3;")
        layout.addWidget(info_label)
        
        # 托盘列表
        tab_widget = QTabWidget()
        
        for i, pallet in enumerate(palletized_cargos):
            tab = QWidget()
            tab_layout = QVBoxLayout(tab)
            
            # 托盘信息
            info = QLabel(f"尺寸: {pallet.length}×{pallet.width}×{pallet.height}cm  |  "
                         f"重量: {pallet.weight:.1f}kg  |  "
                         f"内含 {len(pallet.pallet_contents)} 件货物")
            tab_layout.addWidget(info)
            
            # 货物清单
            content_table = QTableWidget()
            content_table.setColumnCount(6)
            content_table.setHorizontalHeaderLabels(["货物名称", "尺寸(cm)", "位置(x,y,z)", "旋转", "重量(kg)", "体积(cm³)"])
            content_table.setRowCount(len(pallet.pallet_contents))
            
            for j, content in enumerate(pallet.pallet_contents):
                cargo = content.cargo
                content_table.setItem(j, 0, QTableWidgetItem(cargo.name))
                content_table.setItem(j, 1, QTableWidgetItem(f"{cargo.length}×{cargo.width}×{cargo.height}"))
                content_table.setItem(j, 2, QTableWidgetItem(f"({content.x:.0f}, {content.y:.0f}, {content.z:.0f})"))
                content_table.setItem(j, 3, QTableWidgetItem("是" if content.rotated else "否"))
                content_table.setItem(j, 4, QTableWidgetItem(f"{cargo.weight:.1f}"))
                content_table.setItem(j, 5, QTableWidgetItem(f"{cargo.volume:.0f}"))
            
            content_table.horizontalHeader().setStretchLastSection(True)
            tab_layout.addWidget(content_table)
            
            # 添加查看3D视图按钮
            view_btn = QPushButton("🔍 查看托盘3D视图")
            view_btn.clicked.connect(lambda checked, p=pallet: self._show_pallet_3d_view(p))
            tab_layout.addWidget(view_btn)
            
            tab_widget.addTab(tab, f"托盘{i+1}")
        
        layout.addWidget(tab_widget)
        
        # 按钮
        btn_layout = QHBoxLayout()
        
        confirm_btn = QPushButton("✅ 确认并应用")
        confirm_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px 30px;")
        
        cancel_btn = QPushButton("❌ 取消")
        cancel_btn.setStyleSheet("padding: 10px 30px;")
        
        def apply_result():
            # 更新货物列表
            self.cargos = remaining_cargos + palletized_cargos
            self.update_cargo_table()
            dialog.accept()
            QMessageBox.information(self, "组托完成", 
                f"已将选中货物组成 {len(palletized_cargos)} 个托盘")
        
        confirm_btn.clicked.connect(apply_result)
        cancel_btn.clicked.connect(dialog.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(confirm_btn)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def _show_pallet_3d_view(self, pallet: Cargo):
        """显示托盘的3D视图"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"{pallet.name} - 3D视图")
        dialog.setMinimumSize(700, 500)
        layout = QVBoxLayout(dialog)
        
        # 创建临时容器用于显示托盘内容
        temp_container = Container(
            name=pallet.name,
            length=pallet.length,
            width=pallet.width,
            height=pallet.content_height,  # 不含底座的高度
            max_weight=pallet.weight * 2
        )
        
        # 创建PlacedCargo列表
        temp_placed = []
        for i, content in enumerate(pallet.pallet_contents):
            placed = PlacedCargo(
                cargo=content.cargo,
                x=content.x,
                y=content.y,
                z=content.z,
                rotated=content.rotated,
                step_number=i + 1
            )
            temp_placed.append(placed)
        
        # 创建3D视图
        pallet_view = Container3DView()
        pallet_view.container = temp_container
        pallet_view.placed_cargos = temp_placed
        pallet_view.setMinimumSize(650, 400)
        layout.addWidget(pallet_view)

        # 视图控制
        ctrl_layout = QHBoxLayout()
        for name, preset in [("正视", "front"), ("俯视", "top"), ("等轴", "iso")]:
            btn = QPushButton(name)
            btn.clicked.connect(lambda checked, p=preset: pallet_view.set_view(p))
            ctrl_layout.addWidget(btn)

        # 全屏按钮
        fullscreen_btn = QPushButton("全屏")
        fullscreen_btn.setStyleSheet("background-color: #222; color: #fff; font-weight: bold;")
        ctrl_layout.addWidget(fullscreen_btn)
        layout.addLayout(ctrl_layout)

        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)

        # 全屏切换逻辑
        def toggle_fullscreen():
            if dialog.isFullScreen():
                dialog.showNormal()
                fullscreen_btn.setText("全屏")
            else:
                dialog.showFullScreen()
                fullscreen_btn.setText("退出全屏")
        fullscreen_btn.clicked.connect(toggle_fullscreen)

        dialog.exec()

    def show_fullscreen_3d_view(self):
        """显示全屏3D视图"""
        if not self.container:
            QMessageBox.warning(self, "警告", "请先选择集装箱并执行配载")
            return
        
        # 创建全屏对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("3D配载视图 - 全屏模式")
        dialog.setStyleSheet("background-color: #1e1e1e;")
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 创建新的3D视图
        fullscreen_view = Container3DView()
        fullscreen_view.container = self.gl_widget.container
        fullscreen_view.placed_cargos = list(self.gl_widget.placed_cargos) if self.gl_widget.placed_cargos else []
        fullscreen_view.all_container_results = list(self.gl_widget.all_container_results) if self.gl_widget.all_container_results else []
        fullscreen_view.current_container_index = self.gl_widget.current_container_index
        
        # 多集装箱模式检测 - 只要有container_results就显示切换选项
        has_container_results = len(fullscreen_view.all_container_results) >= 1
        
        layout.addWidget(fullscreen_view, 1)
        
        # 控制栏
        ctrl_layout = QHBoxLayout()
        
        # 视图切换按钮
        views = [("正视", "front"), ("后视", "back"), ("左视", "left"), 
                 ("右视", "right"), ("俯视", "top"), ("等轴", "iso")]
        for name, preset in views:
            btn = QPushButton(name)
            btn.setStyleSheet("background-color: #333; color: white; padding: 8px 15px; font-weight: bold;")
            btn.clicked.connect(lambda checked, p=preset: fullscreen_view.set_view(p))
            ctrl_layout.addWidget(btn)
        
        ctrl_layout.addStretch()
        
        # 内容切换下拉框
        container_combo = QComboBox()
        container_combo.setStyleSheet("""
            QComboBox {
                background-color: #333; 
                color: white; 
                padding: 8px 15px;
                min-width: 180px;
                font-weight: bold;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox QAbstractItemView {
                background-color: #333;
                color: white;
                selection-background-color: #2196F3;
            }
        """)
        
        # 添加选项
        if has_container_results:
            # 有集装箱结果模式
            container_combo.addItem("📦 全部概览")
            for i, result in enumerate(fullscreen_view.all_container_results):
                util = result.volume_utilization
                container_combo.addItem(f"📦 集装箱 {i+1} ({util:.1f}%)")
            container_combo.setCurrentIndex(fullscreen_view.current_container_index + 1)
        else:
            # 单集装箱模式（没有container_results的情况）
            container_combo.addItem(f"📦 {self.container.name}")
            # 如果有托盘，添加托盘选项
            pallet_cargos = [p for p in self.placed_cargos if p.cargo.is_pallet]
            for i, pallet_placed in enumerate(pallet_cargos):
                container_combo.addItem(f"🎁 {pallet_placed.cargo.name}")
        
        def on_content_changed(index):
            if has_container_results:
                # 多集装箱模式：切换集装箱
                fullscreen_view.show_container(index - 1)
            else:
                # 单集装箱模式
                if index == 0:
                    # 显示集装箱
                    fullscreen_view.container = self.gl_widget.container
                    fullscreen_view.placed_cargos = self.gl_widget.placed_cargos.copy()
                    fullscreen_view.all_container_results = []
                    fullscreen_view.current_container_index = -1
                    fullscreen_view.update()
                else:
                    # 显示托盘
                    pallet_cargos = [p for p in self.placed_cargos if p.cargo.is_pallet]
                    if index - 1 < len(pallet_cargos):
                        pallet = pallet_cargos[index - 1].cargo
                        # 创建临时容器显示托盘内容
                        temp_container = Container(
                            name=pallet.name,
                            length=pallet.length,
                            width=pallet.width,
                            height=pallet.content_height,
                            max_weight=pallet.weight * 2
                        )
                        temp_placed = []
                        for j, content in enumerate(pallet.pallet_contents):
                            placed = PlacedCargo(
                                cargo=content.cargo,
                                x=content.x,
                                y=content.y,
                                z=content.z,
                                rotated=content.rotated,
                                step_number=j + 1
                            )
                            temp_placed.append(placed)
                        fullscreen_view.container = temp_container
                        fullscreen_view.placed_cargos = temp_placed
                        fullscreen_view.all_container_results = []
                        fullscreen_view.current_container_index = -1
                        fullscreen_view.reset_view()
        
        container_combo.currentIndexChanged.connect(on_content_changed)
        
        ctrl_layout.addWidget(QLabel("<span style='color:white;'>查看:</span>"))
        ctrl_layout.addWidget(container_combo)
        
        # 重置视图按钮
        reset_btn = QPushButton("🔄 重置")
        reset_btn.setStyleSheet("background-color: #333; color: white; padding: 8px 15px;")
        reset_btn.clicked.connect(fullscreen_view.reset_view)
        ctrl_layout.addWidget(reset_btn)
        
        # 退出全屏按钮
        exit_btn = QPushButton("✕ 退出全屏")
        exit_btn.setStyleSheet("background-color: #c62828; color: white; padding: 8px 15px; font-weight: bold;")
        exit_btn.clicked.connect(dialog.accept)
        ctrl_layout.addWidget(exit_btn)
        
        layout.addLayout(ctrl_layout)
        
        # 全屏显示
        dialog.showFullScreen()
        dialog.exec()
    
    def load_pallets_to_container(self):
        """装载托盘到集装箱"""
        # 筛选托盘货物
        pallet_cargos = [c for c in self.cargos if c.name.startswith("托盘")]
        
        if not pallet_cargos:
            QMessageBox.warning(self, "警告", "没有托盘可装载，请先执行'小件组托'")
            return
        
        if not self.container:
            QMessageBox.warning(self, "警告", "请先选择集装箱")
            return
        
        # 直接执行配载
        self.start_loading()
        
        QMessageBox.information(self, "提示", 
            f"已将 {len(pallet_cargos)} 个托盘装入集装箱")
    
    def clear_loading(self):
        """清除配载结果"""
        self.placed_cargos.clear()
        self.gl_widget.placed_cargos = []
        self.gl_widget.update()
        
        self.stats_label.setText("请先添加货物并开始配载")
        self.volume_progress.setValue(0)
        self.volume_label.setText("0%")
        self.weight_progress.setValue(0)
        self.weight_label.setText("0%")
    
    def export_loading_plan(self):
        """导出配载方案"""
        if not self.placed_cargos:
            QMessageBox.warning(self, "警告", "没有配载结果可导出")
            return
        
        filename, filter_used = QFileDialog.getSaveFileName(
            self, "导出配载方案", "", 
            "PDF文件 (*.pdf);;文本文件 (*.txt);;JSON文件 (*.json)")
        
        if filename:
            try:
                # 检查是否为多集装箱模式
                is_multi = self.multi_container_mode and len(self.container_results) > 0
                
                if is_multi:
                    self.export_multi_container_plan(filename)
                else:
                    self.export_single_container_plan(filename)
                    
                QMessageBox.information(self, "成功", "配载方案导出成功")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {e}")
    
    def export_single_container_plan(self, filename: str):
        """导出单集装箱配载方案"""
        # 计算重心信息
        total_volume = sum(p.cargo.volume for p in self.placed_cargos)
        total_weight = sum(p.cargo.weight for p in self.placed_cargos)
        
        # 计算重心
        if total_weight > 0:
            cog_x = sum(p.center_x * p.cargo.weight for p in self.placed_cargos) / total_weight
            cog_y = sum(p.center_y * p.cargo.weight for p in self.placed_cargos) / total_weight
            cog_z = sum(p.center_z * p.cargo.weight for p in self.placed_cargos) / total_weight
            center_x = self.container.length / 2
            center_y = self.container.width / 2
            offset_x = cog_x - center_x
            offset_y = cog_y - center_y
        else:
            cog_x = cog_y = cog_z = 0
            offset_x = offset_y = 0
        
        if filename.endswith(".pdf"):
            self.export_loading_plan_pdf(filename, total_volume, total_weight,
                                         cog_x, cog_y, cog_z, offset_x, offset_y)
        elif filename.endswith(".json"):
            self.export_single_container_json(filename, total_volume, total_weight,
                                              cog_x, cog_y, cog_z, offset_x, offset_y)
        elif filename.endswith(".txt"):
            self.export_single_container_txt(filename, total_volume, total_weight,
                                             cog_x, cog_y, cog_z, offset_x, offset_y)
    
    def export_multi_container_plan(self, filename: str):
        """导出多集装箱配载方案"""
        if filename.endswith(".pdf"):
            self.export_multi_container_pdf(filename)
        elif filename.endswith(".json"):
            self.export_multi_container_json(filename)
        elif filename.endswith(".txt"):
            self.export_multi_container_txt(filename)
    
    def export_multi_container_txt(self, filename: str):
        """导出多集装箱方案为文本文件"""
        with open(filename, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("                     多集装箱配载方案\n")
            f.write("=" * 70 + "\n\n")
            
            f.write(f"使用集装箱数量: {len(self.container_results)}\n")
            f.write(f"总装载件数: {len(self.placed_cargos)}\n\n")
            
            for idx, result in enumerate(self.container_results):
                f.write("-" * 70 + "\n")
                f.write(f"集装箱 #{idx + 1}: {result.container.name}\n")
                f.write("-" * 70 + "\n")
                f.write(f"内部尺寸: {result.container.length} × {result.container.width} × {result.container.height} cm\n")
                f.write(f"装载件数: {len(result.placed_cargos)}\n")
                f.write(f"空间利用率: {result.volume_utilization:.1f}%\n")
                f.write(f"载重利用率: {result.weight_utilization:.1f}%\n\n")
                
                f.write("装载明细:\n")
                for i, p in enumerate(result.placed_cargos, 1):
                    f.write(f"  {i:3d}. {p.cargo.name}\n")
                    f.write(f"       尺寸: {p.actual_length}×{p.actual_width}×{p.cargo.height} cm\n")
                    f.write(f"       位置: ({p.x:.0f}, {p.y:.0f}, {p.z:.0f})\n")
                f.write("\n")
            
            f.write("=" * 70 + "\n")
    
    def export_multi_container_json(self, filename: str):
        """导出多集装箱方案为JSON文件"""
        data = {
            "multi_container": True,
            "container_count": len(self.container_results),
            "total_loaded": len(self.placed_cargos),
            "containers": []
        }
        
        for idx, result in enumerate(self.container_results):
            container_data = {
                "index": idx + 1,
                "container": {
                    "name": result.container.name,
                    "length": result.container.length,
                    "width": result.container.width,
                    "height": result.container.height
                },
                "statistics": {
                    "loaded_count": len(result.placed_cargos),
                    "volume_utilization": round(result.volume_utilization, 1),
                    "weight_utilization": round(result.weight_utilization, 1)
                },
                "cargos": [
                    {
                        "name": p.cargo.name,
                        "dimensions": f"{p.actual_length}×{p.actual_width}×{p.cargo.height}",
                        "position": {"x": round(p.x, 1), "y": round(p.y, 1), "z": round(p.z, 1)},
                        "rotated": p.rotated
                    }
                    for p in result.placed_cargos
                ]
            }
            data["containers"].append(container_data)
        
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def export_multi_container_pdf(self, filename: str):
        """导出多集装箱方案为PDF文件"""
        if not PDF_SUPPORT:
            QMessageBox.warning(self, "警告", "PDF导出功能不可用，请安装 reportlab 库")
            return
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER
        
        doc = SimpleDocTemplate(filename, pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('ChineseTitle', parent=styles['Title'],
                                     fontName='ChineseFont', fontSize=24, alignment=TA_CENTER, spaceAfter=30)
        heading_style = ParagraphStyle('ChineseHeading', parent=styles['Heading2'],
                                       fontName='ChineseFont', fontSize=14,
                                       textColor=colors.HexColor('#2c5282'), spaceBefore=15, spaceAfter=10)
        normal_style = ParagraphStyle('ChineseNormal', parent=styles['Normal'],
                                      fontName='ChineseFont', fontSize=10, leading=14)
        
        elements = []
        elements.append(Paragraph("多集装箱配载方案", title_style))
        elements.append(Spacer(1, 20))
        
        # 总体统计
        elements.append(Paragraph("一、总体统计", heading_style))
        summary_data = [
            ['统计项', '数值'],
            ['使用集装箱数', f'{len(self.container_results)} 个'],
            ['总装载件数', f'{len(self.placed_cargos)} 件'],
        ]
        summary_table = Table(summary_data, colWidths=[6*cm, 9*cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # 每个集装箱的详情
        for idx, result in enumerate(self.container_results):
            elements.append(Paragraph(f"集装箱 #{idx + 1}: {result.container.name}", heading_style))
            
            # 集装箱信息
            info_data = [
                ['项目', '数值'],
                ['内部尺寸', f'{result.container.length} × {result.container.width} × {result.container.height} cm'],
                ['装载件数', f'{len(result.placed_cargos)} 件'],
                ['空间利用率', f'{result.volume_utilization:.1f}%'],
                ['载重利用率', f'{result.weight_utilization:.1f}%'],
            ]
            info_table = Table(info_data, colWidths=[5*cm, 10*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 10))
            
            # 装载明细
            cargo_header = ['序号', '货物名称', '尺寸(cm)', '位置(X,Y,Z)']
            cargo_data = [cargo_header]
            for i, p in enumerate(result.placed_cargos, 1):
                cargo_data.append([
                    str(i),
                    p.cargo.name[:12],
                    f'{p.actual_length}×{p.actual_width}×{p.cargo.height:.0f}',
                    f'({p.x:.0f}, {p.y:.0f}, {p.z:.0f})'
                ])
            
            cargo_table = Table(cargo_data, colWidths=[1.5*cm, 5*cm, 4*cm, 4.5*cm])
            cargo_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#805ad5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf5ff')]),
                ('PADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(cargo_table)
            elements.append(Spacer(1, 15))
            
            # 添加等轴测视图
            if PIL_SUPPORT:
                try:
                    from reportlab.platypus import Image as RLImage
                    
                    # 临时更新3D视图数据以获取正确的截图
                    old_container = self.gl_widget.container
                    old_placed = self.gl_widget.placed_cargos
                    old_index = self.gl_widget.current_container_index
                    
                    self.gl_widget.container = result.container
                    self.gl_widget.placed_cargos = result.placed_cargos
                    self.gl_widget.current_container_index = idx  # 非-1表示单个集装箱模式
                    
                    generator = LoadingImageGenerator(result.container, result.placed_cargos, self.gl_widget)
                    iso_img = generator.generate_isometric_view(450, 350)
                    
                    # 恢复原来的数据
                    self.gl_widget.container = old_container
                    self.gl_widget.placed_cargos = old_placed
                    self.gl_widget.current_container_index = old_index
                    
                    if iso_img:
                        import tempfile
                        tmp_dir = os.path.dirname(filename) or tempfile.gettempdir()
                        tmp_path = os.path.join(tmp_dir, f"_temp_container_{idx}_{id(self)}.png")
                        iso_img.save(tmp_path)
                        
                        elements.append(Paragraph(f"装载示意图", normal_style))
                        elements.append(Spacer(1, 5))
                        elements.append(RLImage(tmp_path, width=14*cm, height=11*cm))
                        
                        # 记录临时文件以便后续清理
                        if not hasattr(self, '_temp_files'):
                            self._temp_files = []
                        self._temp_files.append(tmp_path)
                except Exception as e:
                    elements.append(Paragraph(f"装载图生成失败: {str(e)}", normal_style))
            
            elements.append(Spacer(1, 20))
            
            # 如果不是最后一个集装箱，添加分页
            if idx < len(self.container_results) - 1:
                elements.append(PageBreak())
        
        doc.build(elements)
        
        # 清理临时文件
        if hasattr(self, '_temp_files'):
            for tmp_path in self._temp_files:
                try:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except:
                    pass
            self._temp_files = []
    
    def export_single_container_json(self, filename: str, total_volume: float, total_weight: float,
                                      cog_x: float, cog_y: float, cog_z: float,
                                      offset_x: float, offset_y: float):
        """导出单集装箱方案为JSON"""
        data = {
            "container": {
                "name": self.container.name,
                "type": self.container.container_type,
                "length": self.container.length,
                "width": self.container.width,
                "height": self.container.height,
                "max_weight": self.container.max_weight
            },
            "statistics": {
                "loaded_count": len(self.placed_cargos),
                "total_volume_m3": round(total_volume / 1000000, 3),
                "total_weight_kg": round(total_weight, 1),
                "volume_utilization": round((total_volume/self.container.volume)*100, 1),
                "weight_utilization": round((total_weight/self.container.max_weight)*100, 1)
            },
            "center_of_gravity": {
                "x": round(cog_x, 1),
                "y": round(cog_y, 1),
                "z": round(cog_z, 1),
                "offset_x": round(offset_x, 1),
                "offset_y": round(offset_y, 1)
            },
            "loading_steps": [
                {
                    "step": i + 1,
                    "cargo_name": p.cargo.name,
                    "dimensions": f"{p.cargo.length}×{p.cargo.width}×{p.cargo.height}",
                    "weight": p.cargo.weight,
                    "position": {"x": round(p.x, 1), "y": round(p.y, 1), "z": round(p.z, 1)},
                    "rotated": p.rotated
                }
                for i, p in enumerate(self.placed_cargos)
            ]
        }
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def export_single_container_txt(self, filename: str, total_volume: float, total_weight: float,
                                     cog_x: float, cog_y: float, cog_z: float,
                                     offset_x: float, offset_y: float):
        """导出单集装箱方案为文本文件"""
        with open(filename, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("                     集装箱配载方案\n")
            f.write("=" * 70 + "\n\n")
            
            f.write(f"集装箱类型: {self.container.name}\n")
            f.write(f"容器类别: {self.container.container_type}\n")
            f.write(f"内部尺寸: {self.container.length} × {self.container.width} × {self.container.height} cm\n")
            f.write(f"容积: {self.container.volume_cbm:.1f} m³\n")
            f.write(f"最大载重: {self.container.max_weight:,} kg\n\n")
            
            f.write("-" * 70 + "\n")
            f.write("重心分析:\n")
            f.write("-" * 70 + "\n")
            f.write(f"  重心位置: X={cog_x:.1f}cm, Y={cog_y:.1f}cm, Z={cog_z:.1f}cm\n")
            f.write(f"  横向偏移: {offset_x:.1f}cm {'(偏左)' if offset_x < 0 else '(偏右)' if offset_x > 0 else '(居中)'}\n")
            f.write(f"  纵向偏移: {offset_y:.1f}cm {'(偏前)' if offset_y < 0 else '(偏后)' if offset_y > 0 else '(居中)'}\n")
            
            max_offset = min(self.container.length, self.container.width) * 0.1
            if abs(offset_x) < max_offset and abs(offset_y) < max_offset:
                f.write("  评估: ✓ 重心分布良好\n\n")
            else:
                f.write("  评估: ⚠ 重心偏移较大，建议调整\n\n")
            
            f.write("-" * 70 + "\n")
            f.write("装载步骤 (按顺序装载):\n")
            f.write("-" * 70 + "\n\n")
            
            for i, p in enumerate(self.placed_cargos, 1):
                f.write(f"步骤 {i:3d}: {p.cargo.name}\n")
                f.write(f"  尺寸: {p.cargo.length} × {p.cargo.width} × {p.cargo.height} cm\n")
                f.write(f"  重量: {p.cargo.weight} kg\n")
                f.write(f"  位置: X={p.x:.1f}, Y={p.y:.1f}, Z={p.z:.1f} cm\n")
                f.write(f"  旋转: {'是' if p.rotated else '否'}\n")
                f.write(f"  加固: {self.get_securing_advice(p, i-1, len(self.placed_cargos))}\n\n")
            
            f.write("-" * 70 + "\n")
            f.write("尾部加固建议:\n")
            f.write("-" * 70 + "\n")
            f.write(self.get_tail_securing_advice())
            f.write("\n")
            
            f.write("-" * 70 + "\n")
            f.write("统计信息:\n")
            f.write(f"  装载件数: {len(self.placed_cargos)}\n")
            f.write(f"  总体积: {total_volume/1000000:.2f} m³\n")
            f.write(f"  空间利用率: {(total_volume/self.container.volume)*100:.1f}%\n")
            f.write(f"  总重量: {total_weight:.1f} kg\n")
            f.write(f"  载重利用率: {(total_weight/self.container.max_weight)*100:.1f}%\n")
            f.write("=" * 70 + "\n")

    def export_loading_plan_pdf(self, filename: str, total_volume: float, total_weight: float,
                                 cog_x: float, cog_y: float, cog_z: float, 
                                 offset_x: float, offset_y: float):
        """导出配载方案为PDF格式"""
        if not PDF_SUPPORT:
            QMessageBox.warning(self, "警告", "PDF导出功能不可用，请安装 reportlab 库:\npip install reportlab")
            return
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # 创建PDF文档
        doc = SimpleDocTemplate(
            filename,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # 样式设置
        styles = getSampleStyleSheet()
        
        # 中文标题样式
        title_style = ParagraphStyle(
            'ChineseTitle',
            parent=styles['Title'],
            fontName='ChineseFont',
            fontSize=24,
            alignment=TA_CENTER,
            spaceAfter=30
        )
        
        heading_style = ParagraphStyle(
            'ChineseHeading',
            parent=styles['Heading2'],
            fontName='ChineseFont',
            fontSize=14,
            textColor=colors.HexColor('#2c5282'),
            spaceBefore=15,
            spaceAfter=10
        )
        
        normal_style = ParagraphStyle(
            'ChineseNormal',
            parent=styles['Normal'],
            fontName='ChineseFont',
            fontSize=10,
            leading=14
        )
        
        elements = []
        
        # 标题
        elements.append(Paragraph("集装箱配载方案", title_style))
        elements.append(Spacer(1, 20))
        
        # 容器信息部分
        elements.append(Paragraph("一、容器信息", heading_style))
        container_data = [
            ['项目', '数值'],
            ['容器类型', self.container.name],
            ['容器类别', self.container.container_type],
            ['内部尺寸', f'{self.container.length} × {self.container.width} × {self.container.height} cm'],
            ['容积', f'{self.container.volume_cbm:.1f} m³'],
            ['最大载重', f'{self.container.max_weight:,} kg'],
        ]
        
        container_table = Table(container_data, colWidths=[5*cm, 10*cm])
        container_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f7fafc')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(container_table)
        elements.append(Spacer(1, 20))
        
        # 装载统计部分
        elements.append(Paragraph("二、装载统计", heading_style))
        vol_util = (total_volume / self.container.volume) * 100 if self.container.volume > 0 else 0
        wt_util = (total_weight / self.container.max_weight) * 100 if self.container.max_weight > 0 else 0
        
        stats_data = [
            ['统计项目', '数值'],
            ['装载件数', f'{len(self.placed_cargos)} 件'],
            ['总体积', f'{total_volume/1000000:.2f} m³'],
            ['空间利用率', f'{vol_util:.1f}%'],
            ['总重量', f'{total_weight:.1f} kg'],
            ['载重利用率', f'{wt_util:.1f}%'],
        ]
        
        stats_table = Table(stats_data, colWidths=[5*cm, 10*cm])
        stats_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f0fff4')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))
        
        # 重心分析部分
        elements.append(Paragraph("三、重心分析", heading_style))
        max_offset = min(self.container.length, self.container.width) * 0.1
        cog_status = "良好" if abs(offset_x) < max_offset and abs(offset_y) < max_offset else "需注意"
        
        cog_data = [
            ['分析项目', '数值', '评估'],
            ['重心X坐标', f'{cog_x:.1f} cm', ''],
            ['重心Y坐标', f'{cog_y:.1f} cm', ''],
            ['重心Z坐标', f'{cog_z:.1f} cm', ''],
            ['横向偏移', f'{offset_x:.1f} cm', '偏左' if offset_x < 0 else '偏右' if offset_x > 0 else '居中'],
            ['纵向偏移', f'{offset_y:.1f} cm', '偏前' if offset_y < 0 else '偏后' if offset_y > 0 else '居中'],
            ['整体评估', cog_status, '✓' if cog_status == "良好" else '⚠'],
        ]
        
        cog_table = Table(cog_data, colWidths=[4*cm, 5*cm, 6*cm])
        cog_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3182ce')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#ebf8ff')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(cog_table)
        elements.append(Spacer(1, 20))
        
        # 装载明细部分
        elements.append(Paragraph("四、装载明细", heading_style))
        
        # 装载步骤表头
        loading_header = ['序号', '货物名称', '尺寸 (cm)', '重量 (kg)', '位置 (X,Y,Z)', '旋转', '加固建议']
        loading_data = [loading_header]
        
        for i, p in enumerate(self.placed_cargos, 1):
            row = [
                str(i),
                p.cargo.name[:10],  # 截断过长的名称
                f'{p.cargo.length}×{p.cargo.width}×{p.cargo.height}',
                f'{p.cargo.weight:.1f}',
                f'{p.x:.0f},{p.y:.0f},{p.z:.0f}',
                '是' if p.rotated else '否',
                self.get_securing_advice(p, i-1, len(self.placed_cargos))[:15]
            ]
            loading_data.append(row)
        
        loading_table = Table(loading_data, colWidths=[1*cm, 2.5*cm, 3*cm, 2*cm, 2.5*cm, 1.2*cm, 3*cm])
        loading_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#805ad5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf5ff')]),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(loading_table)
        elements.append(Spacer(1, 20))
        
        # 尾部加固建议
        elements.append(Paragraph("五、尾部加固建议", heading_style))
        securing_advice = self.get_tail_securing_advice()
        for line in securing_advice.split('\n'):
            if line.strip():
                elements.append(Paragraph(line.strip(), normal_style))
        
        elements.append(Spacer(1, 30))
        
        # 添加组托方案详情
        pallet_cargos = [p for p in self.placed_cargos if p.cargo.is_pallet and p.cargo.pallet_contents]
        if pallet_cargos:
            elements.append(PageBreak())
            elements.append(Paragraph("六、组托方案详情", heading_style))
            elements.append(Spacer(1, 10))
            
            for pallet_placed in pallet_cargos:
                pallet = pallet_placed.cargo
                elements.append(Paragraph(
                    f"📦 {pallet.name} - 尺寸: {pallet.length}×{pallet.width}×{pallet.height}cm, "
                    f"重量: {pallet.weight:.1f}kg, 包含 {len(pallet.pallet_contents)} 件货物",
                    normal_style
                ))
                elements.append(Spacer(1, 5))
                
                # 组托内容表格
                pallet_header = ['序号', '货物名称', '尺寸(cm)', '位置(X,Y,Z)', '重量(kg)']
                pallet_data = [pallet_header]
                
                for i, content in enumerate(pallet.pallet_contents, 1):
                    cargo = content.cargo
                    pallet_data.append([
                        str(i),
                        cargo.name[:12],
                        f'{cargo.length}×{cargo.width}×{cargo.height}',
                        f'({content.x:.0f},{content.y:.0f},{content.z:.0f})',
                        f'{cargo.weight:.1f}'
                    ])
                
                pallet_table = Table(pallet_data, colWidths=[1*cm, 3.5*cm, 3.5*cm, 3.5*cm, 2.5*cm])
                pallet_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ed8936')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fed7aa')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fffaf0')]),
                    ('PADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(pallet_table)
                elements.append(Spacer(1, 15))
            
            # 生成组托等轴测视图
            if PIL_SUPPORT:
                for pallet_placed in pallet_cargos[:3]:  # 最多显示3个托盘的视图
                    pallet = pallet_placed.cargo
                    try:
                        # 创建临时容器用于生成托盘视图
                        temp_container = Container(
                            name=pallet.name,
                            length=pallet.length,
                            width=pallet.width,
                            height=pallet.content_height,
                            max_weight=pallet.weight * 2
                        )
                        
                        temp_placed = []
                        for i, content in enumerate(pallet.pallet_contents):
                            placed = PlacedCargo(
                                cargo=content.cargo,
                                x=content.x,
                                y=content.y,
                                z=content.z,
                                rotated=content.rotated,
                                step_number=i + 1
                            )
                            temp_placed.append(placed)
                        
                        pallet_generator = LoadingImageGenerator(temp_container, temp_placed)
                        pallet_img = pallet_generator._generate_isometric_view_pil(400, 300)
                        
                        if pallet_img:
                            import tempfile
                            tmp_dir = os.path.dirname(filename) or tempfile.gettempdir()
                            pallet_tmp_path = os.path.join(tmp_dir, f"_temp_pallet_{pallet.name}_{id(self)}.png")
                            pallet_img.save(pallet_tmp_path)
                            
                            elements.append(Paragraph(f"{pallet.name} 组托示意图:", normal_style))
                            elements.append(RLImage(pallet_tmp_path, width=12*cm, height=9*cm))
                            elements.append(Spacer(1, 10))
                            
                            # 记录临时文件用于清理
                            if not hasattr(self, '_temp_pallet_files'):
                                self._temp_pallet_files = []
                            self._temp_pallet_files.append(pallet_tmp_path)
                    except Exception as e:
                        elements.append(Paragraph(f"托盘视图生成失败: {str(e)}", normal_style))
        
        # 尝试添加装载图
        section_num = "七" if pallet_cargos else "六"
        tmp_path = None
        if PIL_SUPPORT:
            elements.append(PageBreak())
            elements.append(Paragraph(f"{section_num}、装载示意图", heading_style))
            
            try:
                # 生成等轴测视图
                generator = LoadingImageGenerator(self.container, self.placed_cargos, self.gl_widget)
                iso_img = generator.generate_isometric_view(500, 400)
                
                if iso_img:
                    # 保存临时图片到与目标PDF相同的目录
                    import tempfile
                    tmp_dir = os.path.dirname(filename) or tempfile.gettempdir()
                    tmp_path = os.path.join(tmp_dir, f"_temp_loading_diagram_{id(self)}.png")
                    iso_img.save(tmp_path)
                    
                    # 添加到PDF
                    elements.append(RLImage(tmp_path, width=15*cm, height=12*cm))
            except Exception as e:
                elements.append(Paragraph(f"装载图生成失败: {str(e)}", normal_style))
        
        # 生成PDF
        doc.build(elements)
        
        # 清理临时文件
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass
        
        # 清理组托临时文件
        if hasattr(self, '_temp_pallet_files'):
            for pallet_path in self._temp_pallet_files:
                if os.path.exists(pallet_path):
                    try:
                        os.remove(pallet_path)
                    except:
                        pass
            self._temp_pallet_files = []

    def get_securing_advice(self, placed_cargo, index: int, total: int) -> str:
        """获取单个货物的加固建议"""
        advice = []
        
        # 根据位置给出建议
        if placed_cargo.z == 0:  # 底层
            advice.append("底层固定")
        
        if placed_cargo.cargo.weight > 500:  # 重货
            advice.append("使用绑带固定")
        
        if index >= total - 3:  # 最后几件
            advice.append("尾部加固")
        
        # 根据是否可堆叠
        if not placed_cargo.cargo.stackable:
            advice.append("顶部勿压")
        
        return ", ".join(advice) if advice else "标准加固"
    
    def analyze_tail_space(self) -> dict:
        """分析集装箱尾部空间情况，用于生成加固建议"""
        if not self.placed_cargos or not self.container:
            return {}
        
        # 找到最后一排货物的 X 坐标
        max_x_end = 0
        last_row_cargos = []
        
        for p in self.placed_cargos:
            x_end = p.x + p.actual_length
            if x_end > max_x_end:
                max_x_end = x_end
        
        # 尾部剩余空间
        tail_gap = self.container.length - max_x_end
        
        # 找最后一排的货物（X坐标最大的那些）
        threshold = max_x_end - 50  # 50cm 范围内的都算最后一排
        for p in self.placed_cargos:
            if p.x + p.actual_length >= threshold:
                last_row_cargos.append(p)
        
        # 分析宽度方向的空隙
        width_gaps = []
        if last_row_cargos:
            # 按 Y 坐标排序
            sorted_by_y = sorted(last_row_cargos, key=lambda p: p.y)
            # 检查左边空隙
            if sorted_by_y[0].y > 5:
                width_gaps.append(('左侧', sorted_by_y[0].y))
            # 检查货物之间的空隙
            for i in range(len(sorted_by_y) - 1):
                gap = sorted_by_y[i+1].y - (sorted_by_y[i].y + sorted_by_y[i].actual_width)
                if gap > 5:
                    width_gaps.append(('货物间', gap))
            # 检查右边空隙
            last_cargo = sorted_by_y[-1]
            right_gap = self.container.width - (last_cargo.y + last_cargo.actual_width)
            if right_gap > 5:
                width_gaps.append(('右侧', right_gap))
        
        # 分析高度方向的空隙（最后一排货物上方的空间）
        height_gaps = []
        for p in last_row_cargos:
            top_gap = self.container.height - (p.z + p.cargo.height)
            if top_gap > 10:
                height_gaps.append((p.cargo.name, top_gap, p.z + p.cargo.height))
        
        # 分析最后一排是否稳定
        bottom_cargos = [p for p in last_row_cargos if p.z < 1]
        stacked_cargos = [p for p in last_row_cargos if p.z >= 1]
        
        return {
            'tail_gap': tail_gap,
            'last_row_count': len(last_row_cargos),
            'width_gaps': width_gaps,
            'height_gaps': height_gaps,
            'bottom_cargos': bottom_cargos,
            'stacked_cargos': stacked_cargos,
            'max_x_end': max_x_end
        }
    
    def get_tail_securing_advice(self) -> str:
        """获取智能尾部加固建议，根据实际空间分析"""
        analysis = self.analyze_tail_space()
        
        if not analysis:
            return "  无货物，无需加固建议"
        
        advice = []
        advice.append("━━━━━━━━━━ 集装箱尾部加固建议 ━━━━━━━━━━")
        advice.append("")
        
        # 1. 尾部纵向空隙处理
        tail_gap = analysis.get('tail_gap', 0)
        if tail_gap > 0:
            advice.append(f"【纵向空隙】尾部剩余空间: {tail_gap:.0f} cm")
            if tail_gap > 100:
                advice.append("  ⚠️ 空隙较大 (>100cm)，建议:")
                advice.append("    • 使用木方框架搭建隔板固定")
                advice.append("    • 配合充气袋填充大空间")
                advice.append("    • 考虑使用货物网或绑带横向固定")
            elif tail_gap > 50:
                advice.append("  ⚠️ 中等空隙 (50-100cm)，建议:")
                advice.append("    • 使用2-3个充气袋填充")
                advice.append("    • 或使用木条/木块搭建支撑")
            elif tail_gap > 20:
                advice.append("  • 使用充气袋填充 (1-2个)")
                advice.append("  • 或使用泡沫块/纸箱填充")
            else:
                advice.append("  ✓ 空隙较小，使用泡沫条或气泡膜填充即可")
        else:
            advice.append("【纵向空隙】✓ 货物贴紧柜门，无纵向空隙")
        
        advice.append("")
        
        # 2. 宽度方向空隙处理
        width_gaps = analysis.get('width_gaps', [])
        if width_gaps:
            advice.append("【横向空隙】检测到宽度方向存在空隙:")
            for position, gap in width_gaps:
                if gap > 30:
                    advice.append(f"  ⚠️ {position}空隙 {gap:.0f}cm - 建议使用充气袋填充")
                elif gap > 10:
                    advice.append(f"  • {position}空隙 {gap:.0f}cm - 建议使用木块或泡沫块填充")
                else:
                    advice.append(f"  • {position}空隙 {gap:.0f}cm - 可用填充物填塞")
        else:
            advice.append("【横向空隙】✓ 货物紧密排列，无明显横向空隙")
        
        advice.append("")
        
        # 3. 高度方向处理
        height_gaps = analysis.get('height_gaps', [])
        stacked = analysis.get('stacked_cargos', [])
        bottom = analysis.get('bottom_cargos', [])
        
        if height_gaps:
            advice.append("【垂直空隙】最后一排货物上方空间:")
            # 只显示前3个
            for cargo_name, gap, top_z in height_gaps[:3]:
                if gap > 50:
                    advice.append(f"  • {cargo_name[:10]}: 顶部{gap:.0f}cm空隙 - 建议使用木条固定防止顶部货物移动")
        
        if stacked:
            advice.append("【堆叠货物】检测到多层堆叠的货物:")
            advice.append(f"  • 共 {len(stacked)} 件堆叠货物")
            advice.append("  • 建议使用绑带将上下层货物绑定")
            advice.append("  • 高处重货需特别注意，使用钢丝绳加固")
        
        advice.append("")
        
        # 4. 底部固定建议
        if bottom:
            advice.append("【底部固定】底层货物加固建议:")
            heavy_bottom = [p for p in bottom if p.cargo.weight > 200]
            if heavy_bottom:
                advice.append(f"  • 底层有 {len(heavy_bottom)} 件重货 (>200kg)")
                advice.append("  • 建议在货物底部放置防滑垫")
                advice.append("  • 使用木块或楔子在货物前后固定")
            else:
                advice.append("  • 使用防滑垫或木条固定底层货物")
        
        advice.append("")
        
        # 5. 根据容器类型的特定建议
        if hasattr(self, 'container') and self.container:
            advice.append("【特别注意事项】")
            if self.container.container_type == "truck":
                advice.append("  🚛 货车运输注意事项:")
                advice.append("    • 确保重心尽量靠近车轴，避免头重或尾重")
                advice.append("    • 使用防滑垫防止刹车时货物前冲")
                advice.append("    • 货物固定需能承受急刹车的惯性力")
            elif self.container.container_type == "shipping":
                advice.append("  🚢 海运集装箱注意事项:")
                advice.append("    • 预留膨胀空间，防止温度变化导致货物变形")
                advice.append("    • 柜门端加固需特别注意，防止开门时货物倾倒")
                advice.append("    • 建议在门端使用木方或钢管横向固定")
                advice.append("    • 考虑海上颠簸，所有加固材料需加强")
            else:
                advice.append("  • 确保所有货物固定牢靠")
                advice.append("  • 检查绑带/绳索是否系紧")
        
        advice.append("")
        advice.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        
        return "\n".join(advice)
    
    def show_securing_advice_dialog(self):
        """显示智能加固建议对话框"""
        if not self.placed_cargos:
            QMessageBox.information(self, "提示", "请先进行配载，然后再查看加固建议。")
            return
        
        # 创建对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("🔧 智能加固建议")
        dialog.setMinimumSize(700, 600)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1e1e1e;
            }
            QTextEdit {
                background-color: #2d2d2d;
                color: #e0e0e0;
                border: 1px solid #3d3d3d;
                border-radius: 5px;
                padding: 10px;
                font-family: 'Consolas', 'Microsoft YaHei', monospace;
                font-size: 12px;
            }
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1084d8;
            }
            QLabel {
                color: #e0e0e0;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        
        # 标题
        title_label = QLabel("📦 基于配载结果的智能加固分析")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        layout.addWidget(title_label)
        
        # 统计信息
        analysis = self.analyze_tail_space()
        stats_text = f"装载货物: {len(self.placed_cargos)} 件 | "
        stats_text += f"尾部空隙: {analysis.get('tail_gap', 0):.0f} cm | "
        stats_text += f"最后一排: {analysis.get('last_row_count', 0)} 件"
        stats_label = QLabel(stats_text)
        stats_label.setStyleSheet("padding: 5px; color: #9cdcfe;")
        layout.addWidget(stats_label)
        
        # 加固建议内容
        advice_text = QTextEdit()
        advice_text.setReadOnly(True)
        advice_text.setText(self.get_tail_securing_advice())
        layout.addWidget(advice_text)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        copy_btn = QPushButton("📋 复制到剪贴板")
        copy_btn.clicked.connect(lambda: (
            QApplication.clipboard().setText(advice_text.toPlainText()),
            QMessageBox.information(dialog, "提示", "已复制到剪贴板！")
        ))
        btn_layout.addWidget(copy_btn)
        
        btn_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()

    # ==================== 多集装箱功能 ====================
    
    def toggle_multi_container_mode(self, state):
        """切换多集装箱模式"""
        self.multi_container_mode = state == 2
        self.container_count_spin.setEnabled(self.multi_container_mode)
        self.container_selector_group.setVisible(False)  # 开始配载后才显示
    
    def on_container_selector_changed(self, index):
        """集装箱选择器变化"""
        if not self.container_results:
            return
        
        if index == 0:
            # 全部概览 - 使用 -1 表示概览模式
            self.gl_widget.show_container(-1)
            self.update_stats_for_container(-1)  # 显示总体统计
        else:
            # 显示特定集装箱
            container_index = index - 1
            if container_index < len(self.container_results):
                self.gl_widget.show_container(container_index)
                self.update_stats_for_container(container_index)
    
    def update_stats_for_container(self, container_index: int):
        """更新特定集装箱的统计信息"""
        if container_index < 0:
            # 显示总体统计
            total_loaded = sum(len(r.placed_cargos) for r in self.container_results)
            total_volume = sum(r.total_volume for r in self.container_results)
            total_weight = sum(r.total_weight for r in self.container_results)
            
            # 计算平均利用率
            avg_vol_util = sum(r.volume_utilization for r in self.container_results) / len(self.container_results) if self.container_results else 0
            avg_wt_util = sum(r.weight_utilization for r in self.container_results) / len(self.container_results) if self.container_results else 0
            
            self.stats_label.setText(
                f"共 {len(self.container_results)} 个集装箱 | "
                f"总装载: {total_loaded} 件 | "
                f"总体积: {total_volume/1000000:.2f} m³ | "
                f"总重量: {total_weight:.1f} kg"
            )
            self.volume_progress.setValue(int(avg_vol_util))
            self.volume_label.setText(f"{avg_vol_util:.1f}%")
            self.weight_progress.setValue(int(avg_wt_util))
            self.weight_label.setText(f"{avg_wt_util:.1f}%")
        else:
            # 显示单个集装箱统计
            result = self.container_results[container_index]
            self.stats_label.setText(
                f"集装箱 #{container_index + 1} | "
                f"装载: {len(result.placed_cargos)} 件 | "
                f"体积: {result.total_volume/1000000:.2f} m³ | "
                f"重量: {result.total_weight:.1f} kg"
            )
            self.volume_progress.setValue(int(result.volume_utilization))
            self.volume_label.setText(f"{result.volume_utilization:.1f}%")
            self.weight_progress.setValue(int(result.weight_utilization))
            self.weight_label.setText(f"{result.weight_utilization:.1f}%")
    
    # ==================== 拖拽调整功能 ====================
    
    def toggle_drag_mode(self, checked: bool):
        """切换拖拽调整模式"""
        self.gl_widget.set_drag_mode(checked)
        
        # 启用/禁用旋转按钮
        self.rotate_cargo_btn.setEnabled(checked)
        
        if checked:
            self.drag_mode_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FF9800;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
            """)
            self.drag_hint_label.setText("拖拽模式已开启：左键选中 → 拖动移动 → Shift+拖动调高度 → R键旋转 → 方向键微调")
            self.drag_hint_label.setVisible(True)
        else:
            self.drag_mode_btn.setStyleSheet("""
                QPushButton {
                    background-color: #37474F;
                    color: white;
                    border: 1px solid #546E7A;
                    border-radius: 6px;
                    padding: 8px 16px;
                }
            """)
            self.drag_hint_label.setVisible(False)
    
    def rotate_selected_cargo_from_btn(self):
        """从按钮触发旋转选中的货物"""
        if self.gl_widget.rotate_selected_cargo():
            # 更新显示信息
            index = self.gl_widget.selected_cargo_index
            if 0 <= index < len(self.placed_cargos):
                self.update_selected_cargo_info(index)
                placed = self.placed_cargos[index]
                self.drag_hint_label.setText(f"已旋转: {placed.cargo.name} → 新位置: ({placed.x:.0f}, {placed.y:.0f}, {placed.z:.0f})")
        else:
            self.drag_hint_label.setText("无法旋转：货物不允许旋转或会与其他货物碰撞")
    
    def toggle_collision_detection(self, state):
        """切换碰撞检测开关"""
        self.gl_widget.collision_enabled = (state == 2)
    
    def fine_tune_cargo(self, dx: int, dy: int, dz: int):
        """微调货物位置"""
        # 获取步进大小
        step_text = self.step_size_combo.currentText()
        step = float(step_text.split()[0])  # 提取数字部分
        
        if self.gl_widget.move_selected_cargo(dx * step, dy * step, dz * step):
            # 更新显示信息
            index = self.gl_widget.selected_cargo_index
            if 0 <= index < len(self.placed_cargos):
                self.update_selected_cargo_info(index)
        else:
            self.drag_hint_label.setText("无法移动：碰到边界或其他货物")

    def on_cargo_drag_selected(self, index: int):
        """货物被拖拽选中"""
        if 0 <= index < len(self.placed_cargos):
            cargo = self.placed_cargos[index]
            self.drag_hint_label.setText(
                f"已选中: {cargo.cargo.name} | 位置: ({cargo.x:.0f}, {cargo.y:.0f}, {cargo.z:.0f})"
            )
            # 更新选中货物详情面板
            self.update_selected_cargo_info(index)
    
    def update_selected_cargo_info(self, index: int):
        """更新选中货物的详细信息显示"""
        if not hasattr(self, 'cargo_name_label') or index < 0 or index >= len(self.placed_cargos):
            return
        
        placed = self.placed_cargos[index]
        cargo = placed.cargo
        
        # 基本信息
        pallet_indicator = " 📦[组托]" if cargo.is_pallet else ""
        self.cargo_name_label.setText(f"名称: {cargo.name}{pallet_indicator}")
        if placed.rotated:
            self.cargo_size_label.setText(f"尺寸: {cargo.width} × {cargo.length} × {cargo.height} cm (已旋转)")
        else:
            self.cargo_size_label.setText(f"尺寸: {cargo.length} × {cargo.width} × {cargo.height} cm")
        self.cargo_weight_label.setText(f"重量: {cargo.weight:.1f} kg")
        
        # 如果是托盘，显示包含的货物信息
        if cargo.is_pallet and cargo.pallet_contents:
            self.cargo_stackable_label.setText(f"包含货物: {len(cargo.pallet_contents)} 件")
        else:
            self.cargo_stackable_label.setText(f"可堆叠: {'是' if cargo.stackable else '否'}")
        
        # 位置信息
        self.cargo_pos_label.setText(f"位置: X={placed.x:.0f}, Y={placed.y:.0f}, Z={placed.z:.0f} cm")
        self.cargo_rotation_label.setText(f"旋转: {'是 (长宽互换)' if placed.rotated else '否'}")
        
        # 计算层次 (根据 Z 坐标)
        layer = 1
        for i, p in enumerate(self.placed_cargos):
            if p.z < placed.z:
                layer = max(layer, 2)
            if p.z > placed.z:
                layer = max(layer, 1)
        z_height = placed.z
        if z_height == 0:
            layer_text = "底层 (地面)"
        elif z_height < self.container.height / 3:
            layer_text = "下层"
        elif z_height < self.container.height * 2 / 3:
            layer_text = "中层"
        else:
            layer_text = "上层"
        self.cargo_layer_label.setText(f"层次: {layer_text} (Z={z_height:.0f}cm)")
        
        # 体积信息
        volume_m3 = cargo.volume / 1000000
        self.cargo_volume_label.setText(f"体积: {volume_m3:.3f} m³")
        
        # 加固建议
        securing = self.get_securing_advice(placed, index, len(self.placed_cargos))
        self.cargo_securing_label.setText(f"加固建议: {securing}")
        
        # 更新托盘详情按钮可见性
        if hasattr(self, 'view_pallet_btn'):
            self.view_pallet_btn.setVisible(cargo.is_pallet and len(cargo.pallet_contents) > 0)
            self._selected_pallet_index = index  # 保存选中的托盘索引
        
        # 更新标题
        pallet_mark = " [组托]" if cargo.is_pallet else ""
        self.selected_cargo_group.setTitle(f"📦 选中货物信息{pallet_mark} - 第 {index + 1} 件 / 共 {len(self.placed_cargos)} 件")
    
    def show_selected_pallet_details(self):
        """显示选中托盘的详情"""
        if hasattr(self, '_selected_pallet_index') and 0 <= self._selected_pallet_index < len(self.placed_cargos):
            pallet = self.placed_cargos[self._selected_pallet_index].cargo
            if pallet.is_pallet:
                self._show_pallet_3d_view(pallet)
    
    def on_cargo_drag_moved(self, index: int):
        """货物被拖拽移动后"""
        if 0 <= index < len(self.placed_cargos):
            cargo = self.placed_cargos[index]
            self.drag_hint_label.setText(
                f"已移动: {cargo.cargo.name} | 新位置: ({cargo.x:.0f}, {cargo.y:.0f}, {cargo.z:.0f})"
            )
            # 更新统计信息
            self.update_loading_stats()
    
    def update_loading_stats(self):
        """更新装载统计信息"""
        if not self.placed_cargos:
            return
        
        total_volume = sum(p.cargo.volume for p in self.placed_cargos)
        total_weight = sum(p.cargo.weight for p in self.placed_cargos)
        vol_util = (total_volume / self.container.volume) * 100 if self.container.volume > 0 else 0
        wt_util = (total_weight / self.container.max_weight) * 100 if self.container.max_weight > 0 else 0
        
        self.volume_progress.setValue(int(vol_util))
        self.volume_label.setText(f"{vol_util:.1f}%")
        self.weight_progress.setValue(int(wt_util))
        self.weight_label.setText(f"{wt_util:.1f}%")
    
    # ==================== 导出装载图片 ====================
    
    def export_loading_images(self):
        """导出装载图片"""
        if not self.placed_cargos:
            QMessageBox.warning(self, "警告", "没有配载结果可导出")
            return
        
        if not PIL_SUPPORT:
            QMessageBox.warning(self, "警告", 
                "未安装 Pillow 库，无法生成图片。\n请运行: pip install Pillow")
            return
        
        # 选择保存目录
        directory = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if not directory:
            return
        
        try:
            # 生成图片
            generator = LoadingImageGenerator(self.container, self.placed_cargos, self.gl_widget)
            base_name = os.path.join(directory, "loading_plan")
            saved_files = generator.save_images(base_name)
            
            if saved_files:
                QMessageBox.information(self, "成功", 
                    f"已保存 {len(saved_files)} 张装载图：\n" + 
                    "\n".join([os.path.basename(f) for f in saved_files]))
            else:
                QMessageBox.warning(self, "警告", "图片生成失败")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {e}")

    def show_user_manual(self):
        """显示使用手册"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📖 使用手册")
        dialog.setMinimumSize(700, 600)
        dialog.setStyleSheet("background-color: #1e1e1e;")
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 创建滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: #252525;
            }
            QScrollBar:vertical {
                background-color: #1e1e1e;
                width: 12px;
            }
            QScrollBar::handle:vertical {
                background-color: #3d3d3d;
                border-radius: 6px;
            }
        """)
        
        # 手册内容
        content = QLabel()
        content.setWordWrap(True)
        content.setOpenExternalLinks(True)
        content.setStyleSheet("""
            QLabel {
                color: #e0e0e0;
                font-size: 14px;
                line-height: 1.6;
                padding: 20px;
                background-color: #252525;
            }
        """)
        
        manual_html = """
        <h1 style="color: #81D4FA; text-align: center;">📦 集装箱配载软件使用手册</h1>
        <hr style="border-color: #3d3d3d;">
        
        <h2 style="color: #4FC3F7;">🚀 快速开始</h2>
        <ol>
            <li><b>选择集装箱</b>：左侧面板选择容器类别和型号</li>
            <li><b>添加货物</b>：输入货物名称、尺寸、重量、数量，点击"添加到列表"</li>
            <li><b>执行配载</b>：点击"执行配载"按钮，自动计算最优装载方案</li>
            <li><b>查看结果</b>：在3D视图中查看装载效果</li>
        </ol>
        
        <h2 style="color: #4FC3F7;">📋 功能说明</h2>
        
        <h3 style="color: #81D4FA;">1. 容器选择</h3>
        <ul>
            <li><b>标准集装箱</b>：20GP、40GP、40HC等国际标准集装箱</li>
            <li><b>托盘</b>：标准托盘、欧标托盘等</li>
            <li><b>自定义容器</b>：可自定义任意尺寸的容器</li>
        </ul>
        
        <h3 style="color: #81D4FA;">2. 货物管理</h3>
        <ul>
            <li><b>添加货物</b>：手动输入或从Excel导入</li>
            <li><b>编辑货物</b>：双击货物列表中的单元格可直接编辑</li>
            <li><b>删除货物</b>：选中行后点击"删除选中"按钮</li>
            <li><b>可旋转</b>：勾选后货物可在XY平面旋转90度</li>
        </ul>
        
        <h3 style="color: #81D4FA;">3. 配载规则</h3>
        <ul>
            <li><b>允许堆叠</b>：货物是否可以堆叠放置</li>
            <li><b>重不压轻</b>：重货放下层，轻货放上层</li>
            <li><b>堆叠层数限制</b>：限制最大堆叠层数</li>
            <li><b>支撑比例</b>：上层货物需要的底部支撑面积比例</li>
        </ul>
        
        <h3 style="color: #81D4FA;">4. 两步装载（推荐用于小件）</h3>
        <ol>
            <li><b>第一步：货物组托</b> - 将小箱先组合到托盘上</li>
            <li><b>第二步：托盘装柜</b> - 将托盘装入集装箱</li>
        </ol>
        <p style="color: #FFEB3B;">💡 提示：组托功能适合大量小件货物，可提高装载效率和叉车操作便利性</p>
        
        <h3 style="color: #81D4FA;">5. 多集装箱模式</h3>
        <ul>
            <li>当货物超出单个集装箱容量时，自动使用多个集装箱</li>
            <li>可设置使用的集装箱数量</li>
            <li>使用下拉框切换查看不同集装箱或全部概览</li>
        </ul>
        
        <h2 style="color: #4FC3F7;">🎮 3D视图操作</h2>
        <table style="width:100%; border-collapse: collapse;">
            <tr style="background-color: #333;">
                <td style="padding: 10px; border: 1px solid #555;"><b>鼠标左键拖动</b></td>
                <td style="padding: 10px; border: 1px solid #555;">旋转视图</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #555;"><b>鼠标右键拖动</b></td>
                <td style="padding: 10px; border: 1px solid #555;">平移视图</td>
            </tr>
            <tr style="background-color: #333;">
                <td style="padding: 10px; border: 1px solid #555;"><b>鼠标滚轮</b></td>
                <td style="padding: 10px; border: 1px solid #555;">缩放视图</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #555;"><b>点击货物</b></td>
                <td style="padding: 10px; border: 1px solid #555;">选中并查看货物信息</td>
            </tr>
            <tr style="background-color: #333;">
                <td style="padding: 10px; border: 1px solid #555;"><b>预设视图</b></td>
                <td style="padding: 10px; border: 1px solid #555;">正视/后视/左视/右视/俯视/等轴</td>
            </tr>
        </table>
        
        <h2 style="color: #4FC3F7;">📤 导出功能</h2>
        <ul>
            <li><b>导出方案</b>：导出PDF或Excel格式的装载报告</li>
            <li><b>导出图片</b>：导出3D视图截图</li>
        </ul>
        
        <h2 style="color: #4FC3F7;">💾 数据导入</h2>
        <ul>
            <li><b>从Excel导入</b>：支持批量导入货物数据</li>
            <li>Excel格式要求：名称、长度(cm)、宽度(cm)、高度(cm)、重量(kg)、数量</li>
        </ul>
        
        <h2 style="color: #4FC3F7;">⌨️ 快捷键</h2>
        <table style="width:100%; border-collapse: collapse;">
            <tr style="background-color: #333;">
                <td style="padding: 8px; border: 1px solid #555;"><b>Enter</b></td>
                <td style="padding: 8px; border: 1px solid #555;">添加货物到列表</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #555;"><b>Delete</b></td>
                <td style="padding: 8px; border: 1px solid #555;">删除选中货物</td>
            </tr>
        </table>
        
        <hr style="border-color: #3d3d3d; margin-top: 30px;">
        <p style="text-align: center; color: #9e9e9e;">集装箱配载软件 v0.5 - by Henry Xue</p>
        """
        
        content.setText(manual_html)
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 30px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        dialog.exec()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    
    # 设置深色调色板
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(224, 224, 224))
    palette.setColor(QPalette.ColorRole.Base, QColor(45, 45, 45))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(35, 35, 35))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(224, 224, 224))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor(30, 30, 30))
    palette.setColor(QPalette.ColorRole.Text, QColor(224, 224, 224))
    palette.setColor(QPalette.ColorRole.Button, QColor(45, 45, 45))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(224, 224, 224))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(33, 150, 243))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
    app.setPalette(palette)
    
    window = ContainerLoadingApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
